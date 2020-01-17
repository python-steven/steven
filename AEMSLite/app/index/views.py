from django.shortcuts import render, redirect
# from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from app.login.models import User,Department,Customer,BudgetCodeForm,PartItem,PartItemResult,MaintenanceLog,Configuration,Project,AccountTitle,ExchangeRate
from app.login.views import Update_User_IsActivated
from django.views.generic.base import View
from django.db import connection
from django.http import HttpResponseRedirect,HttpResponse
from app import restful,mail
from app.access_control import access_control,FeeLimit_count,MergerFeeLimit
from datetime import datetime,timedelta,date
from django.conf import settings
import random
import string
import os
import time
import calendar
from openpyxl import load_workbook,Workbook
import json
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
UpdatedTime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#登入首页的时候需要验证信息
class IndexView(View):
    @csrf_exempt
    def get(self,request):
        try:
            id = request.session['user_Id']
            user = User.objects.get(Id=id)
            request.session.set_expiry(0)
            num = BudgetCodeForm.objects.filter(Status="Process",SignerId=id,MergeId=None).count()\
                  + BudgetCodeForm.objects.exclude(MergeId=None).filter(Status="Process",SignerId=id).distinct("MergeId").count()
            return render(request, "./index/main.html", {'user': user,'num':num})
        except:
            return HttpResponseRedirect("/login/")


#获取部门的信息和客户的信息
@csrf_exempt
#@access_control
def Budget_info_get(request):
    if request.method == "GET":
        try:
            data_dict={}
            cus_info = list(Customer.objects.exclude(IsActivated='False').values())
            depart_info = list(Department.objects.exclude(IsActivated='False').values())
            project_info = list(Project.objects.exclude(IsActivated='False').values())
            account_info = list(AccountTitle.objects.exclude(IsActivated='False').values())
            exchangeRate_info =list(ExchangeRate.objects.exclude(IsActivated='False').values())
            t = time.time()
            t = time.localtime(t)
            t = time.strftime("%y%m%d%H%M%S", t)
            n = random.randint(0, 99)
            resn = str(n) if n >= 10 else "0" + str(n)
            data_dict['Number'] = t + resn
            data_dict['Account'] =account_info
            data_dict['projects'] =project_info
            data_dict['customer'] =cus_info
            data_dict['department'] =depart_info
            data_dict['exchangeRate'] =exchangeRate_info
            return restful.ok(data=data_dict)
        except Exception as e:
            return restful.params_error(message=repr(e))

#检查用户和负责人的合法性
@csrf_exempt
#@access_control
def Budget_check_user(request):
    if request.method == "POST":
        try:
            check_user = request.POST['user_approve']
            user_c = User.objects.exclude(IsActivated=False).get(EmployeeId=check_user)
            if user_c:
                return restful.ok()
        except:
            return restful.params_error(message='the Signer not exist need admin check')

#检查需求人是否存在性
@csrf_exempt
#@access_control
def Budget_check_principal(request):
    if request.method == "POST":
        try:
            principal = request.POST['principal']
            user_c =list(User.objects.exclude(IsActivated=False).filter(Name__icontains=principal).values("Id","Name"))
            if len(user_c)>0:
                return restful.ok(data=user_c[0])
            else:
                return restful.params_error(data={'data':""})
        except:
            # return restful.ok()
            # try:
            #     user_num = User.objects.get(EmployeeId=principal)
            #     if user_num:
            #         return restful.ok()
            # except:
            return restful.params_error(data="")


#表单信息的获取
class BudgetCodeApply(View):
    @csrf_exempt
    def get(self,request):
        try:
            page = int(request.GET.get('page'))
            number = request.GET.get('num')
            id = request.session['user_Id']
            dict_data = {}
            count = BudgetCodeForm.objects.filter(PicId=id).count()  # 总共多少条数据
            if number == "All":
                budgetcode_info = BudgetCodeForm.objects.filter(PicId=id).order_by("-UpdatedTime","MergeId") \
                    .values("Id", "BillingType", "Department", "ApplyDate", "Pic", "ProductName", "Signer", "Status",
                            "BudgetCode", "MergeId","FormId")
                budgetcode_info = list(budgetcode_info)
                dict_data['data'] = budgetcode_info
                dict_data['page_count'] = count
                return restful.ok(data=dict_data)
            if number != "All":
                number = int(number)
                page_num = count // number  # 总共多少页
                if count % number > 0:
                    page_num = page_num + 1
                if page_num >= page:
                    budgetcode_info = BudgetCodeForm.objects.filter(PicId=id).order_by("-Id")\
                        .values("Id","BillingType","Department","ApplyDate","Pic","ProductName","Signer","Status","BudgetCode","MergeId","FormId")[(page-1)*number:number*page]
                    budgetcode_info = list(budgetcode_info)
                    dict_data['data'] = budgetcode_info
                    dict_data['page_count'] =page_num
                    return restful.ok(data=dict_data)
                else:
                    return restful.params_error(message="it had no other pages",data={'count':count})
        except Exception as e:
            return restful.params_error(message=repr(e))

    #提交信息送至签核人process
    @csrf_exempt
    @FeeLimit_count
    def post(self, request):
        UpdatedTime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        bud_id = request.POST.get('bud_id',"")                             #判别是否是修改的动作的id
        bud_formId = request.POST.get('bud_formId')                     #表单号的获取
        bud_depart = request.POST.get('Department')                     #部门
        bud_req = request.POST.get('Remark')                            #判断表单是批量还是特殊
        bud_num_type = request.POST.get('bud_num_type')                 #判断是否是201，还是pmcs 表单
        bud_num = request.POST.get('bud_num')                           #单号
        bud_time = request.POST.get('bud_time')                         #需求設備到廠日期
        bud_principal = request.POST.get('bud_principal',"")            #需求人
        bud_machine_name = request.POST.get('bud_machine_name')         #设备名字
        bud_machine_type = request.POST.get('bud_machine_type')         #设备类型
        # bud_request_type = request.POST.get('bud_request_type')         #计算类型方式的number值
        bud_account_type_name = request.POST.get('account_type')        #获取计算公式的名字
        bud_price = request.POST.get('bud_price')                       #单价
        bud_qty = request.POST.get('bud_qty')                           #数量
        bud_qty_type = request.POST.get('bud_qty_type')                 #设备是按台合适个
        bud_money_type = request.POST.get('bud_money_type')             #获取使用是RMB还是USD
        bud_customer = request.POST.get('bud_customer')                 #获取客户
        bud_mach_type= request.POST.get('bud_mach_type')                #获取机种名字
        bud_pn = request.POST.get('bud_pn')                             #PN获取
        bud_user = request.POST.get('bud_user')                         #获取签核人
        bud_reason = request.POST.get('bud_reason')                     #获取申请原因
        created_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')     #获取创建时间
        own_id = request.session.get('user_Id')                         #获取创建人的ID
        try:
        #獲取機種信息和科目信息
            Project_obj = list(Project.objects.filter(Name=bud_mach_type).values("Id","Name","Code"))
            Account_obj = list(AccountTitle.objects.filter(Type=bud_account_type_name).values("Id","Type"))
            ProjectId = Project_obj[0]['Id']
            TypeOfMachine = Project_obj[0]["Name"]
            ProjectCode = Project_obj[0]["Code"]
            AccountTitleId = Account_obj[0]["Id"]
            PurchaseType = Account_obj[0]["Type"]
            time_num = int(time.time())
            time_num = str(time_num)
            file = request.FILES.get('upload_file')                         #附件的获取
            if len(bud_time) == 0:
                bud_time =None
            if file:
                file_name = file.name
                file_sp_name = file_name.split('.')[0]
                file_ven_name = file_sp_name + time_num + '.' + file_name.split('.')[1]
                file_path = os.path.join(settings.MEDIA_ROOT, file_ven_name)
                with open(file_path, 'wb') as f:
                    for chunk in file.chunks():
                        f.write(chunk)
                f.close()
                file_url = request.build_absolute_uri(settings.MEDIA_URL + file_ven_name)
            else:
                file_sp_name=""
                file_url=""
            if bud_id != "":
                try:
                    depart = Department.objects.get(Department=bud_depart)
                    depart_id = depart.Id
                    user = User.objects.get(EmployeeId=bud_user)
                    user_id = user.Id
                    cus = Customer.objects.get(Customer=bud_customer)
                    cus_id = cus.Id
                    pic_user = User.objects.get(Id=own_id)
                    pic_user_EmployeeId = pic_user.EmployeeId
                    if bud_principal == "":
                        requiredPICId=None
                    elif len(User.objects.filter(Name=bud_principal).values("Id")) == 0:
                        requiredPICId=None
                    else:
                        requiredPICId = (User.objects.filter(Name=bud_principal).values("Id"))[0]["Id"]
                    BudgetCodeForm.objects.filter(Id=bud_id).update(DepartmentId=depart_id, Department=bud_depart, Remark=bud_req
                                                  , Attachment=file_sp_name, ApplyDate=created_time
                                                  , equipmentToFactoryDate=bud_time
                                                  , ExternalNumberType=bud_num_type, ExternalNumber=bud_num
                                                  , PicId=own_id, Pic=pic_user_EmployeeId, ProductName=bud_machine_name
                                                  , Model=bud_machine_type, PurchaseType=PurchaseType
                                                  , UnitPrice=bud_price, Quantity=bud_qty, Unit=bud_qty_type
                                                  , Currency=bud_money_type, CustomerId=cus_id
                                                  , Customer=bud_customer, TypeOfMachine=TypeOfMachine
                                                  , ProjectCode=ProjectCode, ApplyReason=bud_reason, SignerId=user_id
                                                  , Signer=bud_user, Status='Process', CreatedTime=created_time
                                                  , UpdatedTime=UpdatedTime, OwnerId=own_id, AttachmentPath=file_url
                                                  ,PN=bud_pn,ProjectId=ProjectId,AccountTitleId=AccountTitleId
                                                  ,FormId=bud_formId,requiredPICId=requiredPICId
                                                  )
                    # 邮件发送创建的表单给签核的人去签核表单信息
                    subject = "Apply Budge Code eForm to You Review"
                    email_1 = user.Email
                    email_2 = pic_user.Email
                    apply_er = User.objects.get(Id=own_id).Name
                    mail_data = {'signer': user.Name, 'applicant': apply_er}
                    mail.send_apply_form_mail([email_1, email_2], subject, mail_data)
                    return restful.ok(message='BudgetCodeForm modify success')
                except Exception as e:
                    return restful.params_error(message=repr(e))
            else:
                try:
                    depart = Department.objects.get(Department=bud_depart)
                    depart_id = depart.Id
                    user = User.objects.get(EmployeeId=bud_user)
                    user_id = user.Id
                    cus = Customer.objects.get(Customer=bud_customer)
                    cus_id = cus.Id
                    pic_user = User.objects.get(Id=own_id)
                    pic_user_EmployeeId = pic_user.EmployeeId

                    if bud_principal == "":
                        requiredPICId=None
                    elif len(User.objects.filter(Name=bud_principal).values("Id")) == 0:
                        requiredPICId=None
                    else:
                        requiredPICId = (User.objects.filter(Name=bud_principal).values("Id"))[0]["Id"]
                    BudgetCodeForm.objects.create(DepartmentId=depart_id, Department=bud_depart, Remark=bud_req
                                                  , Attachment=file_sp_name, ApplyDate=created_time
                                                  , equipmentToFactoryDate=bud_time
                                                  , ExternalNumberType=bud_num_type, ExternalNumber=bud_num
                                                  , PicId=own_id, Pic=pic_user_EmployeeId, ProductName=bud_machine_name
                                                  , Model=bud_machine_type, PurchaseType=PurchaseType
                                                  , UnitPrice=bud_price, Quantity=bud_qty, Unit=bud_qty_type
                                                  , Currency=bud_money_type, CustomerId=cus_id
                                                  , Customer=bud_customer, TypeOfMachine=TypeOfMachine
                                                  , ProjectCode=ProjectCode, ApplyReason=bud_reason, SignerId=user_id
                                                  , Signer=bud_user, Status='Process', CreatedTime=created_time
                                                  , UpdatedTime=UpdatedTime, OwnerId=own_id, AttachmentPath=file_url
                                                  ,PN=bud_pn,ProjectId=ProjectId,AccountTitleId=AccountTitleId
                                                  ,FormId=bud_formId,requiredPICId=requiredPICId
                                                  )
                    # # 邮件发送创建的表单给签核的人去签核表单信息
                    subject = "Apply Budge Code eForm to You Review"
                    email_1 = user.Email
                    email_2 = pic_user.Email
                    apply_er = User.objects.get(Id=own_id).Name
                    mail_data = {'signer': user.Name, 'applicant': apply_er}
                    mail.send_apply_form_mail([email_1, email_2], subject, mail_data)
                    return restful.ok(message='BudgetCodeForm create success')
                except Exception as e:
                    return restful.params_error(message=repr(e))
        except Exception as e:
            return restful.params_error(repr(e))
            # return restful.params_error("you need add manage info")

# 保存信息表单状态为Draft
@csrf_exempt
#@access_control
def Budget_form_save(request):
    if request.method == "POST":
        bud_id = request.POST.get('bud_id')
        bud_formId = request.POST.get('bud_formId')  # 表单号的获取
        bud_depart = request.POST.get('Department')
        bud_req = request.POST.get('Remark')
        bud_num_type = request.POST.get('bud_num_type')
        bud_num = request.POST.get('bud_num')
        bud_time = request.POST.get('bud_time')
        bud_principal = request.POST.get('bud_principal',"")
        bud_machine_name = request.POST.get('bud_machine_name')
        bud_machine_type = request.POST.get('bud_machine_type')
        # bud_request_type = request.POST.get('bud_request_type')
        bud_account_type_name = request.POST.get('account_type')  # 获取计算公式的名字
        bud_price = request.POST.get('bud_price')
        bud_qty = request.POST.get('bud_qty')
        bud_qty_type = request.POST.get('bud_qty_type')
        bud_money_type = request.POST.get('bud_money_type')
        bud_customer = request.POST.get('bud_customer')
        bud_mach_type = request.POST.get('bud_mach_type')
        bud_pn = request.POST.get('bud_pn')                        # PN获取
        bud_user = request.POST.get('bud_user')
        bud_reason = request.POST.get('bud_reason')
        created_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        own_id = request.session.get('user_Id')

        # try:
        # 獲取機種信息和科目信息
        Project_obj = list(Project.objects.filter(Name=bud_mach_type).values("Id", "Name", "Code"))
        Account_obj = list(AccountTitle.objects.filter(Type=bud_account_type_name).values("Id", "Type"))
        ProjectId = Project_obj[0]['Id']
        TypeOfMachine = Project_obj[0]["Name"]
        ProjectCode = Project_obj[0]["Code"]
        AccountTitleId = Account_obj[0]["Id"]
        PurchaseType = Account_obj[0]["Type"]
        time_num = int(time.time())
        time_num = str(time_num)
        if len(bud_time) ==0:
            bud_time=None
        file = request.FILES.get('upload_file')
        if file:
            file_name = file.name
            file_sp_name = file_name.split('.')[0]
            file_ven_name = file_sp_name + time_num + '.' + file_name.split('.')[1]
            file_path = os.path.join(settings.MEDIA_ROOT, file_ven_name)
            with open(file_path, 'wb') as f:
                for chunk in file.chunks():
                    f.write(chunk)
            f.close()
            file_url = request.build_absolute_uri(settings.MEDIA_URL + file_ven_name)
        else:
            file_sp_name=""
            file_url=""
        if len(bud_id) !=0:
            try:
                depart = Department.objects.get(Department=bud_depart)
                depart_id = depart.Id
                user = User.objects.get(EmployeeId=bud_user)
                user_id = user.Id
                cus = Customer.objects.get(Customer=bud_customer)
                cus_id = cus.Id
                pic_user = User.objects.get(Id=own_id)
                pic_user_EmployeeId = pic_user.EmployeeId
                if bud_principal == "":
                    requiredPICId = None
                elif len(User.objects.filter(Name=bud_principal).values("Id")) == 0:
                    requiredPICId = None
                else:
                    requiredPICId = (User.objects.filter(Name=bud_principal).values("Id"))[0]["Id"]
                BudgetCodeForm.objects.filter(Id=bud_id).update(DepartmentId=depart_id, Department=bud_depart, Remark=bud_req
                                                  , Attachment=file_sp_name, ApplyDate=created_time
                                                  , equipmentToFactoryDate=bud_time
                                                  , ExternalNumberType=bud_num_type, ExternalNumber=bud_num
                                                  , PicId=own_id, Pic=pic_user_EmployeeId, ProductName=bud_machine_name
                                                  , Model=bud_machine_type, PurchaseType=PurchaseType
                                                  , UnitPrice=bud_price, Quantity=bud_qty, Unit=bud_qty_type
                                                  , Currency=bud_money_type, CustomerId=cus_id
                                                  , Customer=bud_customer, TypeOfMachine=TypeOfMachine
                                                  , ProjectCode=ProjectCode, ApplyReason=bud_reason, SignerId=user_id
                                                  , Signer=bud_user, Status='Draft', CreatedTime=created_time
                                                  , UpdatedTime=UpdatedTime, OwnerId=own_id, AttachmentPath=file_url
                                                  ,PN=bud_pn,ProjectId=ProjectId,AccountTitleId=AccountTitleId
                                                  ,FormId=bud_formId,requiredPICId=requiredPICId
                                                  )
                return restful.ok(message="BudgetCodeForm modify success")
            except Exception as e:
                return restful.params_error(message=repr(e))
        else:
            try:
                depart = Department.objects.get(Department=bud_depart)
                depart_id = depart.Id
                user = User.objects.get(EmployeeId=bud_user)
                user_id = user.Id
                cus = Customer.objects.get(Customer=bud_customer)
                cus_id = cus.Id
                pic_user = User.objects.get(Id=own_id)
                pic_user_EmployeeId = pic_user.EmployeeId

                if bud_principal == "":
                    requiredPICId = None
                elif len(User.objects.filter(Name=bud_principal).values("Id")) == 0:
                    requiredPICId = None
                else:
                    requiredPICId = (User.objects.filter(Name=bud_principal).values("Id"))[0]["Id"]

                BudgetCodeForm.objects.create(DepartmentId=depart_id, Department=bud_depart, Remark=bud_req
                                              , Attachment=file_sp_name, ApplyDate=created_time
                                              , equipmentToFactoryDate=bud_time
                                              , ExternalNumberType=bud_num_type, ExternalNumber=bud_num
                                              , PicId=own_id, Pic=pic_user_EmployeeId, ProductName=bud_machine_name
                                              , Model=bud_machine_type, PurchaseType=PurchaseType
                                              , UnitPrice=bud_price, Quantity=bud_qty, Unit=bud_qty_type
                                              , Currency=bud_money_type, CustomerId=cus_id
                                              , Customer=bud_customer, TypeOfMachine=TypeOfMachine
                                              , ProjectCode=ProjectCode, ApplyReason=bud_reason, SignerId=user_id
                                              , Signer=bud_user, Status='Draft', CreatedTime=created_time
                                              , UpdatedTime=UpdatedTime, OwnerId=own_id, AttachmentPath=file_url
                                              ,PN=bud_pn,ProjectId=ProjectId,AccountTitleId=AccountTitleId
                                              ,FormId=bud_formId,requiredPICId=requiredPICId
                                              )
                return restful.ok(message="BudgetCodeForm create success")
            except Exception as e:
                return restful.params_error(message=repr(e))
        # except:
        #     return restful.params_error("you need add manage info")
#合并表单为Draft的信息获取
@csrf_exempt
#@access_control
def Budget_merge_order(request):
    if request.method == "GET":
        try:
            id = request.session['user_Id']
            budgetcode_megre_info = BudgetCodeForm.objects.filter(Status='Draft',OwnerId=id).order_by("-UpdatedTime")\
                .values("Id","FormId","Department", "ApplyDate", "Pic","ProductName","Signer", "Status")
            budgetcode_megre_info =list(budgetcode_megre_info)
            return restful.ok(data=budgetcode_megre_info)
        except Exception as e:
            return restful.params_error(message=repr(e))

#合并开单的信息送至签核
@csrf_exempt
#@access_control
@MergerFeeLimit
def merge_form_sub(request):
    if request.method == "POST":
        try:
            checked_id_array = request.POST.getlist('ids[]')
            own_id = request.session.get('user_Id')
            time_id =int(time.time())
            budget_user = BudgetCodeForm.objects.filter(Id__in=checked_id_array).values("Signer").distinct("Signer")
            singer_num = budget_user.count()
            if singer_num == 1:
                BudgetCodeForm.objects.filter(Id__in=list(checked_id_array)).update(MergeId=time_id, BillingType=1,Status="Process")
                # 邮件发送合并表单的要签核的信息给签核人
                subject = "Apply Budge Code eForm to You Review"
                budget_user =list(budget_user)
                sign_user = User.objects.get(EmployeeId=budget_user[0]['Signer'])
                email_1 = sign_user.Email
                sign_user_name = sign_user.Name
                apply_user = User.objects.get(Id=own_id)
                apply_user = apply_user.Name
                mail_data = {'signer': sign_user_name, 'applicant': apply_user}
                mail.send_apply_form_mail([email_1,], subject, mail_data)
                return restful.ok(message='BudgetCodeForm merged success')
            else:
                return restful.params_error(message="merged form signer different")
        except Exception as e:
            return restful.params_error(message=repr(e))
#修改之前先获取表里面的数据并且填入输入框中
@csrf_exempt
#@access_control
def budget_detail_modify(request):
    if request.method == "POST":
        try:
            detail_id = request.POST.get('id')
            detail_obj = BudgetCodeForm.objects.filter(Id=detail_id).values("ExternalNumberEffectiveDate","ExternalNumber")
            detail_obj=list(detail_obj)
            return restful.ok(data=detail_obj)
        except Exception as e:
            return restful.params_error(message=repr(e))


#修改表单信息
@csrf_exempt
#@access_control
def budget_modify_type(request):
    if request.method == "POST":
        try:
            modify_id = int(request.POST['modify_id'])
            modify_date = request.POST['modify_date']
            modify_number = request.POST['modify_number']
            # modify_budget_ob = BudgetCodeForm.objects.get(Id=modify_id)
            # mer_id =modify_budget_ob.MergeId
            # if mer_id != None:
            #     BudgetCodeForm.objects.filter(MergeId=mer_id).update(ExternalNumberEffectiveDate=modify_date,ExternalNumber=modify_number)
            #     return restful.ok(message="modify form success")
            # else:
            obj = BudgetCodeForm.objects.get(Id=modify_id)
            obj.ExternalNumberEffectiveDate = modify_date
            obj.ExternalNumber = modify_number
            obj.save()
            return restful.ok(message="modify form success")
        except Exception as e:
            return restful.params_error(message=repr(e))

#修改reject和draft的表单的获取信息
@csrf_exempt
#@access_control
def budget_modify_unique(request):
    if request.method == "POST":
        try:
            modify_unique_id = request.POST['modify_unique_id']
            modify_unique_ob = BudgetCodeForm.objects.filter(Id=modify_unique_id)
            modify_unique_ob = list(modify_unique_ob.values())
            for i in range(0,len(modify_unique_ob)):
                if AccountTitle.objects.filter(Type=modify_unique_ob[i]['PurchaseType']).count() !=0:
                    modify_unique_ob[i]['Rule']=list(AccountTitle.objects.filter(Type=modify_unique_ob[i]['PurchaseType']).values("Rule"))[0]
                else:
                    modify_unique_ob[i]['Rule']="1"
            return restful.ok(data=modify_unique_ob)

        except Exception as e:
            return restful.params_error(message=repr(e))

#复制表单信息
@csrf_exempt
#@access_control
def budget_copy_type(request):
    if request.method == "POST":
        try:
            #产生formId的函数
            t = time.time()
            t = time.localtime(t)
            t = time.strftime("%y%m%d%H%M%S", t)
            n = random.randint(0, 99)
            resn = str(n) if n >= 10 else "0" + str(n)
            copy_formId = t + resn
            copy_id = request.POST['copy_id']
            cop_ob = BudgetCodeForm.objects.get(Id=copy_id)
            BudgetCodeForm.objects.create(DepartmentId=cop_ob.DepartmentId, Department=cop_ob.Department
                                          , Remark=cop_ob.Remark, ApplyDate=UpdatedTime
                                          , Attachment=cop_ob.Attachment
                                          , ExternalNumberEffectiveDate=cop_ob.ExternalNumberEffectiveDate
                                          , ExternalNumberType=cop_ob.ExternalNumberType, ExternalNumber=cop_ob.ExternalNumber
                                          , PicId=cop_ob.PicId, Pic=cop_ob.Pic
                                          , ProductName=cop_ob.ProductName
                                          , Model=cop_ob.Model, PurchaseType=cop_ob.PurchaseType
                                          , UnitPrice=cop_ob.UnitPrice, Quantity=cop_ob.Quantity, Unit=cop_ob.Unit
                                          , Currency=cop_ob.Currency, CustomerId=cop_ob.CustomerId
                                          , Customer=cop_ob.Customer, TypeOfMachine=cop_ob.TypeOfMachine
                                          , ProjectCode=cop_ob.ProjectCode, ApplyReason=cop_ob.ApplyReason
                                          , SignerId=cop_ob.SignerId, Signer=cop_ob.Signer, Status='Draft'
                                          , CreatedTime=cop_ob.CreatedTime, UpdatedTime=UpdatedTime
                                          , OwnerId=cop_ob.OwnerId, AttachmentPath=cop_ob.AttachmentPath
                                          , FormId=copy_formId, PN=cop_ob.PN, ProjectId=cop_ob.ProjectId
                                          ,AccountTitleId=cop_ob.AccountTitleId
                                          )
            return restful.ok(message='copy form success')
        except Exception as e:
            return restful.params_error(message=repr(e))

#删除(取消)表单信息
@csrf_exempt
#@access_control
def budget_delete_type(request):
    if request.method == "POST":
        try:
            del_id = request.POST['del_id']
            budget_ob = BudgetCodeForm.objects.get(Id=del_id)
            if budget_ob.Status == "Draft" or budget_ob.Status == "Reject" or budget_ob.Status == "Cancel":
                BudgetCodeForm.objects.filter(Id=del_id).delete()
                return restful.ok(message="form delete success")

            if budget_ob.Status == "Process":
                merged_bud = BudgetCodeForm.objects.filter(MergeId=budget_ob.MergeId).count()
                if merged_bud-1 == 1:
                    BudgetCodeForm.objects.filter(MergeId=budget_ob.MergeId).update(MergeId=None,BillingType=0)
                BudgetCodeForm.objects.filter(Id=del_id).update(Status="Cancel",MergeId=None,BillingType=0)
                cancel_budget = list(BudgetCodeForm.objects.filter(Id=del_id).values())
                cancel_budget_Id =cancel_budget[0]['Signer']
                cancel_budget_Name =User.objects.get(EmployeeId=cancel_budget_Id)
                cancel_budget_principal= cancel_budget[0]['Pic']
                if cancel_budget[0]['BillingType']=='0':
                    cancel_budget[0]['BillingType'] = '單獨開單'
                else:
                    cancel_budget[0]['BillingType'] = '合併開單'
                cancel_budget[0]['rate']=check_rate(cancel_budget[0]['Currency'])
                # if ExchangeRate.objects.filter(CurrencyFrom=cancel_budget[0]['Currency'],CurrencyTo="CNY",IsActivated=True).count() != 0:
                #     cancel_budget[0]['rate'] = list(ExchangeRate.objects.filter(CurrencyFrom=cancel_budget[0]['Currency'],CurrencyTo="CNY",IsActivated=True).order_by("-UpdatedTime").values("ExchangeRate"))[
                #         0]["ExchangeRate"]
                # else:
                #     cancel_budget[0]['rate'] = 1
                cancel_budget[0]['count_fee'] = round(cancel_budget[0]['UnitPrice'] * cancel_budget[0]['Quantity']*cancel_budget[0]['rate'],2)
                cancel_budget[0]['ApplyDate']=cancel_budget[0]['ApplyDate'].strftime('%Y-%m-%d')
                mail_user = User.objects.get(Name=cancel_budget_Name.Name)
                mail_principal = User.objects.get(EmployeeId=cancel_budget_principal)
                email_1 = mail_user.Email
                email_2 = mail_principal.Email
                subject = "Cancel Budge Code eForm"
                mail_data={'signer':mail_principal.Name,'reason':cancel_budget[0]['SignRemarks'],'budget_code_list':cancel_budget,'applicant':cancel_budget_Name.Name}
                mail.send_canceled_form_mail([email_1, email_2], subject, mail_data)
                return restful.ok(message='BudgetCodeForm cancel success',data={})
        except Exception as e:
            return restful.params_error(message=repr(e))

#签核单号的信息的获取函数
@csrf_exempt
#@access_control
def budget_singing_info(request):
    if request.method == "GET":
        try:
            id = request.session['user_Id']
            budget_singing_data = BudgetCodeForm.objects.filter(Status='Process',SignerId=id).order_by("ApplyDate")\
                .values("Id","FormId","Department","OwnerId","Customer","ProjectId","ProductName","Model","UnitPrice","Quantity","ApplyReason","AttachmentPath","BillingType","MergeId","Currency")
            budget_singing_data = list(budget_singing_data)
            for i in range(0,len(budget_singing_data)):
                if budget_singing_data[i]["FormId"] == None:
                    budget_singing_data[i]["FormId"]=""
                if budget_singing_data[i]["OwnerId"] != None:
                    budget_singing_data[i]['applier']=list(User.objects.filter(Id=budget_singing_data[i]["OwnerId"]).values("Name"))[0]["Name"]
                else:
                    budget_singing_data[i]['applier']=""
                if budget_singing_data[i]["ProjectId"] != None:
                    budget_singing_data[i]['Project']=list(Project.objects.filter(Id=budget_singing_data[i]["ProjectId"]).values("Name"))[0]["Name"]
                else:
                    budget_singing_data[i]['Project']=""
                budget_singing_data[i]["rate"]=check_rate(budget_singing_data[i]['Currency'])
                # if ExchangeRate.objects.filter(CurrencyFrom=budget_singing_data[i]['Currency'],CurrencyTo="CNY",IsActivated=True).count() !=0:
                #     budget_singing_data[i]["rate"]=list(ExchangeRate.objects.filter(CurrencyFrom=budget_singing_data[i]['Currency'],CurrencyTo="CNY",IsActivated=True).order_by("-UpdatedTime").values('ExchangeRate'))[0]['ExchangeRate']
                # else:
                #     budget_singing_data[i]["rate"]=1
                budget_singing_data[i]["count_price"]=round(budget_singing_data[i]["UnitPrice"]*budget_singing_data[i]["Quantity"]*budget_singing_data[i]["rate"],2)

            return restful.ok(data=budget_singing_data)
        except Exception as e:
            return restful.params_error(message=repr(e))

#签核表单内容
@csrf_exempt
#@access_control
def merge_signed(request):
    if request.method == "POST":
        try:
            bud_id = request.POST['bud_id']
            bud_merged_id = request.POST['bud_merged_id']
            bud_budgetcode = request.POST['budget_cod_text']
            bud_signremarks = request.POST['budget_text']
            sign_id = request.session['user_Id']
            user = User.objects.get(Id=sign_id)
            sign_budget_ob = BudgetCodeForm.objects.get(Id=bud_id)
            #add check department info
            check_depart_ids = User.objects.filter(DepartmentId=2).values_list("Id",flat=True).order_by("Id")
            if user.EmployeeId != sign_budget_ob.Signer:
                return restful.params_error(message="you need connect admin")
            else:
                if bud_merged_id == "null":
                    bud_obj = BudgetCodeForm.objects.get(Id=bud_id)
                    if bud_obj.Status == "Cancel":
                        return restful.params_error("this form had Canceled")
                    elif bud_obj.Department !="MZVT00":
                        BudgetCodeForm.objects.filter(Id=bud_id).update(BudgetCode=bud_budgetcode,
                                                                        UpdatedTime=datetime.now().strftime(
                                                                            '%Y-%m-%d %H:%M:%S'),
                                                                        SignRemarks=bud_signremarks, Status='Approve')
                    else:
                        BudgetCodeForm.objects.filter(Id=bud_id).update(BudgetCode=bud_budgetcode,UpdatedTime=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                                                                        SignRemarks=bud_signremarks, Status='Ongoing')
                    sign_pincipal = sign_budget_ob.Pic
                    mail_list_data = list(BudgetCodeForm.objects.filter(Id=bud_id).values())
                    mail_list_data[0]['rate']=check_rate(mail_list_data[0]['Currency'])
                    # if ExchangeRate.objects.filter(CurrencyFrom=mail_list_data[0]['Currency'],CurrencyTo="CNY",IsActivated=True).count() !=0:
                    #     mail_list_data[0]['rate']=list(ExchangeRate.objects.filter(CurrencyFrom=mail_list_data[0]['Currency'],CurrencyTo="CNY",IsActivated=True).order_by("-UpdatedTime").values("ExchangeRate"))[0]["ExchangeRate"]
                    # else:
                    #     mail_list_data[0]['rate']=1
                    if mail_list_data[0]['BillingType'] == '0':
                        mail_list_data[0]['BillingType'] = '單獨開單'
                    else:
                        mail_list_data[0]['BillingType'] = '合併開單'
                    mail_list_data[0]['ApplyDate']=mail_list_data[0]['ApplyDate'].strftime('%Y-%m-%d')
                    mail_list_data[0]['count_fee'] = round(mail_list_data[0]['UnitPrice'] * mail_list_data[0]['Quantity']*mail_list_data[0]['rate'],2)
                    mail_user = User.objects.get(EmployeeId=sign_pincipal)
                    mail_signer = User.objects.get(EmployeeId=sign_budget_ob.Signer)
                    email_1 = mail_user.Email
                    subject = "Approve Budge Code eForm"
                    mail_data = {'signer': mail_signer.Name, 'budget_code': bud_budgetcode,
                                 'budget_code_list': mail_list_data, 'applicant': mail_user.Name}
                    mail.send_approved_form_mail([email_1, ], subject, mail_data)

                    return restful.ok(message='BudgetCodeForm approved success')
                else:
                    bud_ob_merged = list(BudgetCodeForm.objects.filter(MergeId=int(bud_merged_id)).values("Id","Department","Status","PicId"))
                    # return restful.params_error(data=bud_ob_merged[0]["PicId"])
                    for bud in bud_ob_merged:
                        if bud["Status"] == "Cancel":
                            return restful.params_error("this form had Cancel")

                        elif bud["PicId"] not in check_depart_ids:
                            BudgetCodeForm.objects.filter(Id=bud["Id"]).update(BudgetCode=bud_budgetcode,
                                                                                        SignRemarks=bud_signremarks,
                                                                                        Status='Approve',UpdatedTime=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                        else:
                            BudgetCodeForm.objects.filter(Id=bud["Id"]).update(BudgetCode=bud_budgetcode,
                                                                                 SignRemarks=bud_signremarks,
                                                                                 Status='Ongoing',
                                                                                 UpdatedTime=datetime.now().strftime(
                                                                                     '%Y-%m-%d %H:%M:%S'))
                        approve_budget = list(BudgetCodeForm.objects.filter(Id=bud["Id"]).values())
                        for i in range(0,len(approve_budget)):
                            approve_budget[i]['ApplyDate']=approve_budget[i]['ApplyDate'].strftime('%Y-%m-%d')
                            approve_budget[i]['rate']=check_rate(approve_budget[i]['Currency'])
                            # if ExchangeRate.objects.filter(CurrencyFrom=approve_budget[i]['Currency'],CurrencyTo="CNY",IsActivated=True).count() != 0:
                            #     approve_budget[i]['rate'] = list(ExchangeRate.objects.filter(CurrencyFrom=approve_budget[i]['Currency'],CurrencyTo="CNY",IsActivated=True).order_by("-UpdatedTime").values(
                            #             "ExchangeRate"))[0]["ExchangeRate"]
                            # else:
                            #     approve_budget[i]['rate'] = 1
                            if approve_budget[i]['BillingType'] == '0':
                                approve_budget[i]['BillingType'] = '單獨開單'
                            else:
                                approve_budget[i]['BillingType'] = '合併開單'
                            approve_budget[i]['count_fee'] = round(approve_budget[i]['UnitPrice']*approve_budget[i]['Quantity']*approve_budget[i]['rate'],2)
                        sign_pincipal = sign_budget_ob.Pic
                        mail_user = User.objects.get(EmployeeId=sign_pincipal)
                        mail_signer = User.objects.get(EmployeeId=sign_budget_ob.Signer)
                        email_1 = mail_user.Email
                        subject = "Approve Budge Code eForm"
                        mail_data = {'signer': mail_signer.Name, 'budget_code': bud_budgetcode,
                                     'budget_code_list': approve_budget, 'applicant': mail_user.Name}
                        mail.send_approved_form_mail([email_1, ], subject, mail_data)
                    return restful.ok(message='BudgetCodeForm signed success')
        except Exception as e:
            return restful.params_error(message=repr(e))

#拒绝签核表单信息内容
@csrf_exempt
#@access_control
def merge_rejected(request):
    if request.method == "POST":
        try:
            bud_id = request.POST['bud_id']
            bud_merged_id = request.POST['bud_merged_id']
            bud_budgetcode = request.POST['budget_cod_text']
            bud_signremarks = request.POST['budget_text']
            sign_id = request.session['user_Id']
            user = User.objects.get(Id=sign_id)
            sign_budget_ob = BudgetCodeForm.objects.get(Id=bud_id)
            if user.EmployeeId != sign_budget_ob.Signer:
                return restful.params_error(message="you need connect admin")
            else:
                if bud_merged_id == "null":
                    if sign_budget_ob.Status == "Cancel":
                        return restful.params_error("this form had Canceled")
                    else:
                        BudgetCodeForm.objects.filter(Id=bud_id).update(BudgetCode=bud_budgetcode,
                                                                        SignRemarks=bud_signremarks, Status='Reject',UpdatedTime=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                        sign_pincipal = sign_budget_ob.Pic
                        mail_user = User.objects.get(EmployeeId=sign_pincipal)
                        mail_signer = User.objects.get(EmployeeId=sign_budget_ob.Signer)
                        email_1 = mail_user.Email
                        mail_list_data = list(BudgetCodeForm.objects.filter(Id=bud_id).values())
                        mail_list_data[0]['ApplyDate']=mail_list_data[0]['ApplyDate'].strftime('%Y-%m-%d')
                        if mail_list_data[0]['BillingType'] == '0':
                            mail_list_data[0]['BillingType'] = '單獨開單'
                        else:
                            mail_list_data[0]['BillingType'] = '合併開單'
                        mail_list_data[0]['rate']=check_rate(mail_list_data[0]['Currency'])
                        # if ExchangeRate.objects.filter(CurrencyFrom=mail_list_data[0]['Currency'],CurrencyTo="CNY",IsActivated=True).count() !=0:
                        #     mail_list_data[0]['rate']=list(ExchangeRate.objects.filter(CurrencyFrom=mail_list_data[0]['Currency'],CurrencyTo="CNY",IsActivated=True).order_by("-UpdatedTime").values("ExchangeRate"))[0]["ExchangeRate"]
                        # else:
                        #     mail_list_data[0]['rate']=1
                        mail_list_data[0]['count_fee']=round(mail_list_data[0]['UnitPrice'] * mail_list_data[0]['Quantity']*mail_list_data[0]['rate'],2)
                        subject = "Reject Budge Code eForm"
                        mail_data = {'applicant': mail_user.Name,'signer': mail_signer.Name, 'reason': bud_signremarks,
                                     'budget_code_list': mail_list_data }
                        mail.send_rejected_form_mail([email_1, ], subject, mail_data)
                        return restful.ok(message='BudgetCodeForm rejected')
                else:
                    if sign_budget_ob.Status == "Cancel":
                        return restful.params_error("this form had Cancel")
                    else:
                        BudgetCodeForm.objects.filter(MergeId=bud_merged_id).update(BudgetCode=bud_budgetcode,
                                                                                    SignRemarks=bud_signremarks,
                                                                                    Status='Reject',UpdatedTime=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
                        mail_list_data = list(BudgetCodeForm.objects.filter(MergeId=bud_merged_id).values())
                        for i in range(0,len(mail_list_data)):
                            mail_list_data[i]['ApplyDate']=mail_list_data[i]['ApplyDate'].strftime('%Y-%m-%d')
                            if mail_list_data[i]['BillingType'] == '0':
                                mail_list_data[i]['BillingType'] = '單獨開單'
                            else:
                                mail_list_data[i]['BillingType'] = '合併開單'
                            mail_list_data[i]['rate']=check_rate(mail_list_data[i]['Currency'])
                            # if ExchangeRate.objects.filter(CurrencyFrom=mail_list_data[i]['Currency'],CurrencyTo="CNY",IsActivated=True).count() != 0:
                            #     mail_list_data[i]['rate'] = list(ExchangeRate.objects.filter(CurrencyFrom=mail_list_data[i]['Currency'],CurrencyTo="CNY",IsActivated=True).order_by("-UpdatedTime").values(
                            #             "ExchangeRate"))[0]["ExchangeRate"]
                            # else:
                            #     mail_list_data[i]['rate'] = 1
                            mail_list_data[i]['count_fee'] = round(mail_list_data[i]['UnitPrice']*mail_list_data[i]['Quantity']*mail_list_data[i]['rate'],2)
                        sign_pincipal = sign_budget_ob.Pic
                        mail_user = User.objects.get(EmployeeId=sign_pincipal)
                        mail_signer = User.objects.get(EmployeeId=sign_budget_ob.Signer)
                        # 此表单为签核同意的表单提交信息 发送email给负责人
                        email_1 = mail_user.Email
                        subject = "Reject Budge Code eForm"
                        mail_data = {'applicant': mail_user.Name
                                    ,'signer': mail_signer.Name
                                    ,'reason': bud_signremarks
                                    ,'budget_code_list': mail_list_data, }
                        mail.send_rejected_form_mail([email_1, ], subject, mail_data)
                        return restful.ok(message='BudgetCodeForm rejected')
        except Exception as e:
            return restful.params_error(message=repr(e))

#获取预算编码的内容get
@csrf_exempt
#@access_control
def merge_signed_finished(request):
    if request.method == "GET":
        try:
            page = int(request.GET.get('page'))
            number = request.GET.get('num')
            id = request.session['user_Id']
            dict_data = {}
            count = BudgetCodeForm.objects.filter(Status__in=('Approve','Reject','Ongoing'),SignerId=id).count()
            if number== "All":
                budget_singed_info = BudgetCodeForm.objects.filter(Status__in=('Approve','Reject','Ongoing'),SignerId=id).order_by("-UpdatedTime","MergeId")
                budget_singed_info = budget_singed_info.values("Id","FormId","BillingType", "Department", "ApplyDate"
                                                               , "Pic","ProductName", "Signer", "Status"
                                                               , "BudgetCode", "MergeId","UpdatedTime")
                budget_singed_info = list(budget_singed_info)
                dict_data['data']=budget_singed_info
                dict_data['page_count']=count
                return restful.ok(data=dict_data)
            if number !="All":
                number = int(number)
                page_num = count // number  # 总共多少页
                if count % number > 0:
                    page_num = page_num + 1
                if page_num >= page:
                    budget_singed_info = BudgetCodeForm.objects.filter(Status__in=('Approve', 'Reject','Ongoing'),
                                                                       SignerId=id).order_by("-UpdatedTime", "MergeId")
                    budget_singed_info = budget_singed_info.values("Id","FormId", "BillingType", "Department", "ApplyDate","UpdatedTime"
                                                                   , "Pic", "ProductName", "Signer", "Status"
                                                                   , "BudgetCode", "MergeId")[(page-1)*number:number*page]
                    budget_singed_info = list(budget_singed_info)
                    budget_singed_info = list(budget_singed_info)
                    dict_data['data'] = budget_singed_info
                    dict_data['page_count'] = page_num
                    return restful.ok(data=dict_data)
                else:
                    return restful.params_error(message="it had no other pages",data={'count':count})
        except Exception as e:
            return restful.params_error(message=repr(e))

#获取预算编码信息并且生成报表的信息
@csrf_exempt
#@access_control
def merge_statement_detail(request):
    if request.method == "GET":
        try:
            page = int(request.GET.get('page'))
            number = request.GET.get('num')
            id = request.session['user_Id']
            dict_data = {}
            count = BudgetCodeForm.objects.filter(Status__in=('Approve','Reject','Process','Close'),SignerId=id).count()
            if number == "All":
                budget_statement_detail = BudgetCodeForm.objects.filter(Status__in=('Approve','Reject','Process','Close'),SignerId=id).order_by("-Id")
                budget_statement_detail = budget_statement_detail.values("Id","FormId","BillingType", "Department", "ApplyDate"
                                                                        ,"Pic","ProductName", "Signer", "Status"
                                                                        ,"BudgetCode", "MergeId","UpdatedTime")
                budget_statement_detail = list(budget_statement_detail)
                dict_data['data'] = budget_statement_detail
                dict_data['page_count'] = count
                return restful.ok(data=dict_data)
            if number != "All":
                number = int(number)
                page_num = count // number  # 总共多少页
                if count % number > 0:
                    page_num = page_num + 1
                if page_num >= page:
                    budget_statement_detail = BudgetCodeForm.objects.filter(Status__in=('Approve','Reject','Process','Close'),SignerId=id).order_by("-Id")
                    budget_statement_detail = budget_statement_detail.values("Id","FormId", "BillingType", "Department","ApplyDate"
                                                                             , "Pic", "ProductName", "Signer", "Status"
                                                                             , "BudgetCode", "MergeId","UpdatedTime")[(page-1)*number:number*page]
                    budget_statement_detail = list(budget_statement_detail)
                    dict_data['data'] = budget_statement_detail
                    dict_data['page_count'] = page_num
                    return restful.ok(data=dict_data)
                else:
                    return restful.params_error(message="it had no other pages")
        except Exception as e:
            return restful.params_error(message=repr(e))

#筛选信息的获取
@csrf_exempt
#@access_control
def statement_query(request):
    if request.method == "POST":
        try:
            page = int(request.POST.get('page'))
            number = request.POST.get('num')
            query_billing_type = request.POST['query_billing_type']
            query_department = request.POST['query_department']
            query_start_date = request.POST['query_start_date']
            query_end_date = request.POST['query_end_date']
            query_pic = request.POST['query_pic']
            query_product_name = request.POST['query_product_name']
            query_signer = request.POST['query_signer']
            query_status = request.POST['query_status']
            id = request.session['user_Id']
            dict_data={}
            sql ='select "Id","BillingType","FormId","Department","ApplyDate","Pic","ProductName","Signer","UpdatedTime","Status","BudgetCode"' \
                 ',"MergeId" FROM "BudgetCodeForm" WHERE "Status" !=\'Cancel\' and "Status" !=\'Draft\' and "Status" !=\'Ongoing\' ' #查询状态为approve的对象
            sql2 = 'select count(*) FROM "BudgetCodeForm" WHERE "Status" !=\'Cancel\' and "Status" !=\'Draft\'  and "Status" !=\'Ongoing\''

            if query_billing_type != "":
                sql = sql + 'AND "BillingType" = \''+query_billing_type +'\''
                sql2 = sql2 + 'AND "BillingType" = \''+query_billing_type +'\''
            if query_department != "":
                sql = sql+'and "Department" ilike \'%'+ query_department +'%\''
                sql2 = sql2+'and "Department" ilike \'%'+ query_department +'%\''
            if query_pic != "":
                sql = sql + 'AND "Pic" ilike \'%'+query_pic +'%\''
                sql2 = sql2 + 'AND "Pic" ilike \'%'+query_pic +'%\''
            if query_product_name != "":
                sql = sql + 'AND "ProductName" ilike \'%'+query_product_name +'%\''
                sql2 = sql2 + 'AND "ProductName" ilike \'%'+query_product_name +'%\''
            if query_signer != "":
                sql = sql + 'AND "Signer" ilike \'%'+query_signer +'%\''
                sql2 = sql2 + 'AND "Signer" ilike \'%'+query_signer +'%\''
            if query_status != "":
                sql = sql + 'AND "Status" = \''+query_status +'\''
                sql2 = sql2 + 'AND "Status" = \''+query_status +'\''
            if query_start_date != "":
                sql = sql + 'and "ApplyDate" >=\'' + query_start_date + '\''
                sql2 = sql2 + 'and "ApplyDate" >=\'' + query_start_date + '\''
            if query_end_date != "":
                sql = sql + 'and "ApplyDate" <=\'' + query_end_date + '\''
                sql2 = sql2 + 'and "ApplyDate" <=\'' + query_end_date + '\''
            cur = connection.cursor()
            cur.execute(sql2)
            count = cur.fetchall()  # 数量的总数
            if number == 'All':
                cur = connection.cursor()
                cur.execute(sql)
                data = cur.fetchall()
                dict_data['data'] = data
                dict_data['page_count'] = count[0][0]
                return restful.ok(data=dict_data)
            if number != "All":
                number = int(number)
                count_page = count[0][0] // number  # 总数除以一页显示多少条，得到总的页数
                if count[0][0] % number > 0:
                    count_page += 1
                if page <= count_page:
                    sql = sql +' order by "ApplyDate" limit '+str(number)+' offset ' + str((page-1)*number)
                    cur = connection.cursor()
                    cur.execute(sql)
                    data = cur.fetchall()
                    dict_data['data']=data
                    dict_data['page_count']=count_page
                    return restful.ok(data=dict_data)
                else:
                    return restful.params_error(message="it had no other pages")
        except Exception as e:
            return restful.params_error(message=repr(e))

#详细信息的展示和加载
@csrf_exempt
#@access_control
def budget_code_detail(request):
    if request.method == "POST":
        detail_id = request.POST['detail_id']
        detail_merged_id = request.POST['detail_merged_id']
        detail_merged_id =str(detail_merged_id)
        try:
            try:
                detail_data = BudgetCodeForm.objects.filter(MergeId=detail_merged_id).values()
                detail_data = list(detail_data)
                for i in range(0,len(detail_data)):
                    if AccountTitle.objects.filter(Type=detail_data[i]["ProjectCode"]).count() !=0:
                        detail_data[i]["Rule"] =list(AccountTitle.objects.filter(Type=detail_data[i]["ProjectCode"]).values("Rule"))[0]["Rule"]
                    else:
                        detail_data[i]["Rule"]=""
                    detail_data[i]["rate"]=check_rate(detail_data[i]["Currency"])
                    # if ExchangeRate.objects.filter(CurrencyFrom=detail_data[i]["Currency"],CurrencyTo="CNY",IsActivated=True).count() !=0:
                    #     detail_data[i]["rate"]=list(ExchangeRate.objects.filter(CurrencyFrom=detail_data[i]["Currency"],CurrencyTo="CNY",IsActivated=True).order_by("-UpdatedTime").values("ExchangeRate"))[0]['ExchangeRate']
                    # else:
                    #     detail_data[i]["rate"]=1
                return restful.ok(data=detail_data)
            except:
                detail_data = BudgetCodeForm.objects.filter(Id=detail_id).values()
                detail_data = list(detail_data)
                for i in range(0,len(detail_data)):
                    if AccountTitle.objects.filter(Type=detail_data[i]["ProjectCode"]).count() !=0:
                        detail_data[i]["Rule"] =list(AccountTitle.objects.filter(Type=detail_data[i]["ProjectCode"]).values("Rule"))[0]["Rule"]
                    else:
                        detail_data[i]["Rule"]=""
                    detail_data[i]["rate"]=check_rate(detail_data[i]["Currency"])
                    # if ExchangeRate.objects.filter(CurrencyFrom=detail_data[i]["Currency"],CurrencyTo="CNY",IsActivated=True).count() !=0:
                    #     detail_data[i]["rate"]=list(ExchangeRate.objects.filter(CurrencyFrom=detail_data[i]["Currency"],CurrencyTo="CNY",IsActivated=True).order_by("-UpdatedTime").values("ExchangeRate"))[0]['ExchangeRate']
                    # else:
                    #     detail_data[i]["rate"]=1
                return restful.ok(data=detail_data)
        except:
            return restful.params_error(message='data error ')

#生成报表文件的处理方式
@csrf_exempt
#@access_control
def statement_bring_info(request):
    if request.method == "POST":
        try:
            statement_ids = request.POST.getlist('statement_ids[]')
            query_billing_type = request.POST.get('query_billing_type','')
            query_department = request.POST.get('query_department','')
            query_start_date = request.POST.get('query_start_date','')
            query_end_date = request.POST.get('query_end_date','')
            query_pic = request.POST.get('query_pic','')
            query_product_name = request.POST.get('query_product_name','')
            query_signer = request.POST.get('query_signer','')
            query_status = request.POST.get('query_status','')
            tb1 =['部门','備註/新增或損耗','鏈接','開單狀況','預算編號','PMCS单号','申請日期','PMCS簽核日期','PIC'
                ,'設備名稱/治具類型','規格/型號/版本','類別','單價','申請數量','單位','總費用(RMB)','實際發生金額(KRMB)'
                ,'RMB/USD','客戶','機種','ProjectCode','申請原因/用途']
            tb2 = ['部门', '備註/新增或損耗', '鏈接', '開單狀況', '預算編號', '201单号', '申請日期', '201簽核日期', 'PIC'
                , '設備名稱/治具類型', '規格/型號/版本', '類別', '單價', '申請數量', '單位', '總費用(RMB)', '實際發生金額(KRMB)'
                , 'RMB/USD', '客戶', '機種', 'ProjectCode', '申請原因/用途']
            sql2 = 'select "Department","BudgetCodeForm"."Remark","AttachmentPath","BillingType","BudgetCode","ExternalNumberType"' \
                   ',"ExternalNumber","ApplyDate","ExternalNumberEffectiveDate","Pic","ProductName","Model"' \
                   ',"PurchaseType","UnitPrice","Quantity","Unit","Currency","Customer","TypeOfMachine"' \
                   ',"ProjectCode","ApplyReason","AccountTitle"."Rule" from "BudgetCodeForm" left join "AccountTitle" on "AccountTitle"."Type"="BudgetCodeForm"."PurchaseType"  where 1=1'
            #ESRZ20,PR新增需求,,0,,1,,2019-09-19T10:09:46,2019-09-28T00:00:00,Z15123001,10,10,租金,10,15,PCS,RMB,WISTRON,CPU,HT01,new%20buy,1
            if query_billing_type != "":
                sql2 = sql2 + ' AND "BillingType" = \'' + query_billing_type + '\''
            if query_department != "":
                sql2 = sql2 + ' AND "Department" ilike \'%' + query_department + '%\''
            if query_pic != "":
                sql2 = sql2 + ' AND "Pic" ilike \'%' + query_pic + '%\''
            if query_product_name != "":
                sql2 = sql2 + ' AND "ProductName" ilike \'%' + query_product_name + '%\''
            if query_signer != "":
                sql2 = sql2 + ' AND "Signer" ilike \'%' + query_signer + '%\''
            if query_status != "":
                sql2 = sql2 + ' AND "Status" = \'' + query_status + '\''
            if query_start_date != "":
                sql2 = sql2 + ' AND "ApplyDate" >=\'' + query_start_date + '\''
            if query_end_date != "":
                sql2 = sql2 + ' AND "ApplyDate" <=\'' + query_end_date + '\''
            if len(statement_ids) == 1:
                statement_ids= statement_ids[0]
                sql2 = sql2 + '"BudgetCodeForm"."Id"=' + statement_ids
            if len(statement_ids) >1:
                statement_ids = tuple(statement_ids)
                statement_ids =str(statement_ids)
                sql2 = sql2 + '"BudgetCodeForm"."Id" in ' + statement_ids
            cur = connection.cursor()
            cur.execute(sql2)
            data = list(cur.fetchall())
            data1=[]
            data2=[]
            for i in data:
                i=list(i)
                if i[7]:
                    i[7]=str(i[7]).split(' ')[0]
                if i[8]:
                    i[8]=str(i[8]).split(' ')[0]
                if i[5] == '1':
                    if i[3] == '0':
                        i[3] = "單獨開單"
                    if i[3] == '1':
                        i[3] = "合併開單"
                    i.pop(5)
                    rate=check_rate(i[15])
                    if rate !=0:
                        sum = round(i[12] * i[13]*rate,2)
                    else:
                        sum =""
                    i.insert(15,str(sum))
                    if i[-1] == "1"and sum !="":
                        i.insert(16,str(sum))
                    elif i[-1] == "2"and sum !="":
                        i.insert(16,str('%.3f' %(sum/12)))
                    elif i[-1] == "3"and sum !="":
                        i.insert(16,str('%.3f' %(sum/24)))
                    elif i[-1] == "4"and sum !="":
                        i.insert(16,str('%.3f' %(sum/36)))
                    elif i[-1] == "5"and sum !="":
                        i.insert(16,str('%.3f' %(sum/72)))
                    else:
                        i.insert(16,str(sum))
                    i.pop()
                    data1.append(i)
                if i[5] == '2':
                    if i[3] == '0':
                        i[3] = "單獨開單"
                    if i[3] == '1':
                        i[3] = "合併開單"
                    i.pop(5)
                    rate=check_rate(i[15])
                    if rate!=0:
                        sum = round(i[12] * i[13]*rate,2)
                    else:
                        sum=""
                    i.insert(15, str(sum))
                    if i[-1] == "1" and sum !="":
                        i.insert(16,str(sum))
                    elif i[-1] == "2"and sum !="":
                        i.insert(16,str('%.3f' %(sum/12)))
                    elif i[-1] == "3"and sum !="":
                        i.insert(16,str('%.3f' %(sum/24)))
                    elif i[-1] == "4"and sum !="":
                        i.insert(16,str('%.3f' %(sum/36)))
                    elif i[-1] == "5"and sum !="":
                        i.insert(16,str('%.3f' %(sum/72)))
                    else:
                        i.insert(16,str(sum))
                    i.pop()
                    data2.append(i)
            data1.insert(0,tb1)
            data2.insert(0,tb2)
            #写入文件
            time_num = int(time.time())
            time_num = str(time_num)
            filename = 'budget_code_report_' + time_num + '.xlsx'
            wb = Workbook()
            index = 0
            sheet_name = "PMCS預算編號記錄"
            wb.create_sheet(sheet_name, index=index)
            sheet = wb[sheet_name]
            for row in data1:
                sheet.append(row)
            index2 = 1
            sheet_name = "201领用"
            wb.create_sheet(sheet_name, index=index2)
            sheet = wb[sheet_name]
            for row2 in data2:
                sheet.append(row2)

            wb.save(os.path.join(settings.MEDIA_CHANGE_ROOT, filename))
            file_url = request.build_absolute_uri(settings.MEDIA_CHANGE_URL + filename)
            data = [file_url]
            return restful.ok(data=data)
        except:
            return restful.params_error(message=" please select one information ")

#添加的跟踪表单信息
def Budget_ongoing_info(request):
    if request.method == "GET":
        try:
            id = request.session['user_Id']
            budgetOngoing = BudgetCodeForm.objects.filter(Status='Ongoing',SignerId=id).order_by("-Id")\
                .values("Id","FormId","Department","OwnerId","Customer","ProjectId","ProductName","Model"
                        ,"UnitPrice","Quantity","ApplyReason","AttachmentPath","BillingType","MergeId"
                        ,"Currency","ExternalNumber","equipmentToFactoryDate","poNumber","toFactoryDate","traceRemark")
            budget_onging_data = list(budgetOngoing)
            for i in range(0,len(budget_onging_data)):
                if budget_onging_data[i]["FormId"] == None:
                    budget_onging_data[i]["FormId"]=""
                if budget_onging_data[i]["OwnerId"] != None:
                    budget_onging_data[i]['applier']=list(User.objects.filter(Id=budget_onging_data[i]["OwnerId"]).values("Name"))[0]["Name"]
                else:
                    budget_onging_data[i]['applier']=""
                if budget_onging_data[i]["ProjectId"] != None:
                    budget_onging_data[i]['Project']=list(Project.objects.filter(Id=budget_onging_data[i]["ProjectId"]).values("Name"))[0]["Name"]
                else:
                    budget_onging_data[i]['Project']=""
                budget_onging_data[i]["rate"]=check_rate(budget_onging_data[i]['Currency'])
                budget_onging_data[i]["count_price"]=round(budget_onging_data[i]["UnitPrice"]*budget_onging_data[i]["Quantity"]*budget_onging_data[i]["rate"],2)

            return restful.ok(data=budget_onging_data)
        except Exception as e:
            return restful.params_error(message=repr(e))
#跟踪信息的取消按钮
def Budget_ongoing_cancel(request):
    if request.method == "POST":
        try:
            budget_id = request.POST.get("id","")
            budget_mergerId = request.POST.get("mergerId","")
            if budget_mergerId != "null":
                BudgetCodeForm.objects.filter(MergeId=budget_mergerId).update(Status="Cancel")
                return restful.ok(message="close Form finished")
            elif budget_id != "":
                BudgetCodeForm.objects.filter(Id=budget_id).update(Status="Cancel")
                return restful.ok(message="close Form finished")
        except Exception as e:
            return restful.params_error(repr(e))
def Budget_ongoing_close(request):
    if request.method == "POST":
        try:
            budget_id = request.POST.get("id", "")
            budget_mergerId = request.POST.get("mergerId", "")
            budget_equipmentDate =request.POST.get("equipmentToFactoryDate","")
            budget_ExternalNumber =request.POST.get("ExternalNumber","")
            budget_po =request.POST.get("poNumber","")
            budget_tofactorydate =request.POST.get("toFactoryDate","")
            budget_traceRemark =request.POST.get("traceRemark","")
            if budget_mergerId != "null":
                BudgetCodeForm.objects.filter(MergeId=budget_mergerId).update(Status="Close"
                    ,equipmentToFactoryDate=budget_equipmentDate,ExternalNumber=budget_ExternalNumber
                    ,poNumber=budget_po,toFactoryDate=budget_tofactorydate,traceRemark=budget_traceRemark)
                return restful.ok(message="close Form finished")
            elif budget_id != "":
                BudgetCodeForm.objects.filter(Id=budget_id).update(Status="Close"
                    ,equipmentToFactoryDate=budget_equipmentDate,ExternalNumber=budget_ExternalNumber
                    ,poNumber=budget_po,toFactoryDate=budget_tofactorydate,traceRemark=budget_traceRemark)
                return restful.ok(message="close Form finished")
        except Exception as e:
            return restful.params_error(repr(e))

def Budget_ongoing_save(request):
    if request.method == "POST":
        try:
            budget_id = request.POST.get("id", "")
            budget_mergerId = request.POST.get("mergerId", "")
            budget_equipmentDate =request.POST.get("equipmentToFactoryDate","")
            budget_ExternalNumber = request.POST.get("ExternalNumber", "")
            budget_po = request.POST.get("poNumber", "")
            budget_tofactorydate = request.POST.get("toFactoryDate", "")
            budget_traceRemark = request.POST.get("traceRemark", "")
            if budget_mergerId != "null":
                if budget_equipmentDate !="":
                    BudgetCodeForm.objects.filter(MergeId=budget_mergerId).update(equipmentToFactoryDate=budget_equipmentDate)
                if budget_ExternalNumber !="":
                    BudgetCodeForm.objects.filter(MergeId=budget_mergerId).update(ExternalNumber=budget_ExternalNumber)
                if budget_po !="":
                    BudgetCodeForm.objects.filter(MergeId=budget_mergerId).update(poNumber=budget_po)
                if budget_tofactorydate !="":
                    BudgetCodeForm.objects.filter(MergeId=budget_mergerId).update(toFactoryDate=budget_tofactorydate)
                if budget_traceRemark !="":
                    BudgetCodeForm.objects.filter(MergeId=budget_mergerId).update(traceRemark=budget_traceRemark)
                return restful.ok(message="save form success")
            elif budget_id != "":
                if budget_equipmentDate !="":
                    BudgetCodeForm.objects.filter(Id=budget_id).update(equipmentToFactoryDate=budget_equipmentDate)
                if budget_ExternalNumber !="":
                    BudgetCodeForm.objects.filter(Id=budget_id).update(ExternalNumber=budget_ExternalNumber)
                if budget_po !="":
                    BudgetCodeForm.objects.filter(Id=budget_id).update(poNumber=budget_po)
                if budget_tofactorydate !="":
                    BudgetCodeForm.objects.filter(Id=budget_id).update(toFactoryDate=budget_tofactorydate)
                if budget_traceRemark !="":
                    BudgetCodeForm.objects.filter(Id=budget_id).update(traceRemark=budget_traceRemark)
                return restful.ok(message="save form success")
        except Exception as e:
            return restful.params_error(repr(e))


#获取需要签核的表单发邮件
def ti_xing():
    data = list(BudgetCodeForm.objects.filter(Status="Process").distinct("SignerId").values("SignerId","Pic"))
    if len(data) !=0:
        for i in range(0, len(data)):
            num = BudgetCodeForm.objects.filter(Status="Process", SignerId=data[0]['SignerId'], MergeId=None).count()
            num += BudgetCodeForm.objects.exclude(MergeId=None).filter(Status="Process", SignerId=data[0]['SignerId']).distinct("MergeId").count()
            user_mail = list(User.objects.filter(Id=data[i]['SignerId']).values("Email", "Name"))
            # apply_mail = list(User.objects.filter(EmployeeId=data[i]['Pic']).values("Email", "Name"))
            subject = "There are some forms waiting for your verification"
            mail_data = {'signer': user_mail[0]["Name"], 'applicant_count': num}
            mail.send_remind_form_mail([user_mail[0]["Email"],], subject, mail_data)
#邮件提醒追踪表单信息
def applyTiXing():
    data = list(BudgetCodeForm.objects.filter(Status="Ongoing").distinct("Pic").values("Pic","PicId"))
    if len(data) !=0:
        for i in range(0, len(data)):
            num = BudgetCodeForm.objects.filter(Status="Ongoing", Pic=data[0]['Pic'], MergeId=None).count()
            num += BudgetCodeForm.objects.exclude(MergeId=None).filter(Status="Ongoing", Pic=data[0]['Pic']).distinct("MergeId").count()
            user_mail = list(User.objects.filter(Id=data[i]['PicId']).values("Email", "Name","DepartmentId"))
            # apply_mail = list(User.objects.filter(EmployeeId=data[i]['Pic']).values("Email", "Name"))
            subject = "There are some forms waiting for your edit"
            mail_data = {'signer': user_mail[0]["Name"], 'applicant_count': num}
            #check department info
            depart = Department.objects.filter(Id=user_mail[0]["DepartmentId"],Department="MZVT00",IsActivated=True).count()
            if len(user_mail[0]) !=0 and depart>0:
                mail.send_remind_applyer_mail([user_mail[0]["Email"],], subject, mail_data)

@csrf_exempt
def BudgetCode_set_num(request):
    if request.method == "GET":
        try:
            data ={}
            t = time.time()
            t = time.localtime(t)
            t = time.strftime("%y%m%d%H%M%S", t)
            n = random.randint(0, 99)
            resn = str(n) if n >= 10 else "0" + str(n)
            data['Number'] =t+resn
            return restful.ok(data=data)
        except Exception as e:
            return restful.params_error(repr(e))

@csrf_exempt
def BudgetCode_rate(request):
    if request.method == "POST":
        try:
            CurrencyFrom = request.POST.get("machao")
            CurrencyFromto = request.POST.get("shuang")
            #调用函数查汇率
            rate_data=check_rate(CurrencyFrom,CurrencyFromto)
            if rate_data ==0:
                return restful.params_error(message="please double check exchange rate")
            return restful.ok(data={"rate":rate_data})
        except Exception as e:
            return restful.params_error(repr(e))

def check_rate(currency,currencyto="CNY"):
    check_obj = ExchangeRate.objects.filter(CurrencyFrom=currency, CurrencyTo=currencyto, IsActivated=True)
    if check_obj.count() > 0:
        rate_data = list(check_obj.order_by("-UpdatedTime").values())[0]["ExchangeRate"]
    elif currency =="RMB":
        rate_data = 1
    else:
        rate_data =0
    return rate_data


#"获取当前月的额度以及数据 全部转换成人民币为计算单位"
def FeeLimit_count(Department,AccountTitle):
    #get all budget of current month data and calculate moneny
    # calendar.monthrange(int(now.year),int(now.month))     indicate for this month has how many days ...
    now = datetime.now()
    startTime = '%s-%s-01 00:00:00'%(now.year,now.month)
    endTime = '%s-%s-%s 23:59:59'%(now.year,now.month,calendar.monthrange(int(now.year),int(now.month))[1])
    ''' get all Status is Approve data and data info include 
    (类别："PurchaseType",单价："UnitPrice",数量："Quantity",单位："Unit",币种："Currency",部门："Department" )
     
     '''









