#views.py
from django.shortcuts import render
from app.login.models import PartItemResult,PartItem
from app.DBexcel.mod_excel import Excel_operation
from AEMSLite.settings import BASE_DIR
from django.db import connection
from app.NGrate.views import check_NGRate
from app.maintain_monitor.views import Check_monitor_equipment
import time,datetime
import os

#监控定时任务的执行动作
def crontab_test():
    # insert_many_to_partItemRestul()
    # update_for_partItem()
    # check_NGRate()
    # Check_monitor_equipment()
    print(datetime.datetime.now())

def insert_many_to_partItemRestul():
    """批量插入PartItemResult表的数据"""
    pathname = os.path.join(BASE_DIR, 'app/DBexcel')
    excel_operation = Excel_operation(pathname)
    file_paths = excel_operation.get_xlsx_list()
    for file_path in file_paths:
        datas = excel_operation.read_by_row(file_path,0)
        insert_list = []
        for data in datas[1:]:
            timearry = datetime.datetime.strptime(data[12], "%m/%d/%Y %I:%M:%S %p")
            data[12] = timearry.strftime('%Y-%m-%d %I:%M:%S')
            case = PartItemResult(
                USN=data[0],SN=data[1],OSN=data[2],Asset=data[3],
                PN=data[4],PartName=data[5],Spec=data[6],
                UsedTimes=data[7],Stage=data[8],FixtureId=data[9],
                Result=data[10],ErrorCode=data[11],TrnDate=data[12],
            )
            insert_list.append(case)
        PartItemResult.objects.bulk_create(insert_list)
        excel_operation.solved_backup(file_path)
    print('insert_to_PartItemResult at %s successfully' %datetime.datetime.now())

def update_for_partItem():
    """批量插入到PartItem 表的数据 """
    sql = 'select max("USN"),"SN",max("OSN"),max("PN"),max("PartName"),max("Spec"),max("UsedTimes") as "UsedTimes",' \
          'count(case when "Result"=\'FAIL\' then "Result" else null end) as "ErrorCounts",max("TrnDate") as TrnDate ' \
          'from "PartItemResult" group by "SN";'
    insert_list = []

    with connection.cursor() as cursor:
        start_time = time.time()
        cursor.execute(sql)
        datas = cursor.fetchall()

    for data in datas:
        SN_foo = PartItem.objects.filter(SN=data[1])
        NG_rate = round(data[7] / data[6], 2) if data[6] > 0 else 0
        if SN_foo:
            SN_foo[0].UsedTimes = data[6]
            SN_foo[0].ErrorCounts = data[7]
            SN_foo[0].TruDate = data[8]
            SN_foo[0].NGRate = NG_rate
            SN_foo[0].save()
        else:
            case = PartItem(
                SN=data[1], OSN=data[2],PN=data[3],
                PartName=data[4], Spec=data[5],
                UsedTimes=data[6], NextCheckDate=None,
                ErrorCounts=data[7],TrnDate=data[8],
                NGRate=NG_rate,
            )
            insert_list.append(case)

    if insert_list:
        PartItem.objects.bulk_create(insert_list)
    print('update_for_PartItem at %s successfully'%datetime.datetime.now())


#mod_excel.py
from openpyxl import load_workbook,Workbook
import datetime
import shutil               #shutil则就是对os中文件操作的补充。--移动 复制  打包 压缩 解压，
import os
class Excel_operation():
    def __init__(self,pathname):
        self.pathname = pathname

    #从指定目录中搜索并返回xlsx文件列表
    def get_xlsx_list(self):
        res = []
        pending_path = os.path.join(self.pathname,'pending')
        try:
            for dirpath, dirname, filenames in os.walk(pending_path):
                for filename in filenames:
                    if filename.endswith(".xlsx") and not filename.startswith('~$'):
                        res.append(os.path.join(dirpath, filename))
        except Exception as e:
            print(e)
        return res

    def solved_backup(self,path_file):
        path_solved = os.path.join(self.pathname,'solved')
        try:
            if not os.path.exists(path_solved):
                os.mkdir(path_solved)
            shutil.move(path_file,path_solved)
        except Exception as e:
            print(e)

    #按行读取内容
    def read_by_row(self,filepath,sheet_num,data_only=True):
        results = []
        try:
            wb = load_workbook(filepath,data_only=data_only)
            sheets = wb.sheetnames
            sheet = wb[sheets[sheet_num]]
            rows = sheet.rows
        except Exception as e:
            print(e)
            return 0

        #生成字段列表
        fields = [cell.value.replace(' ', '') for cell in next(rows)]
        results.append(fields + ['UpdatedTime'])

        #读取数据
        for row in rows:
            result = []
            for cell in row:
                result.append(cell.value)
            result.append(datetime.date.today())
            results.append(result)
        return results












#mod_log.py
import logging.handlers
class Logger(logging.Logger):
    def __init__(self,filename=None):
        super(Logger,self).__init__(self)

        #日志文件名
        if filename is None:
            filename = 'DBexcel/log.log'
        self.filename = filename

        #创建handler(每天生成一个,保留30天的日志)
        fh = logging.handlers.TimedRotatingFileHandler(self.filename,'D',1,30)
        fh.suffix = "%Y%m%d-%H%m.log"
        fh.setLevel(logging.DEBUG)

        #这个handler用于输出控制台
        ch = logging.StreamHandler()
        ch.setLevel(logging.DEBUG)

        #定义handler的输出格式
        formatter = logging.Formatter('[%(asctime)s] - %(filename)s - [Line:%(lineno)d] - [%(levelname)s] - %(message)s')
        fh.setFormatter(formatter)
        ch.setFormatter(formatter)

        #给Logger添加handler
        self.addHandler(fh)
        self.addHandler(ch)
logger = Logger()






































from django.http import HttpResponseRedirect,HttpResponse
from app.login.models import User
def _auth(args):#args 是传入的，需要验证的权限
    def __auth(func):
        def _login(request):
            user_Id = request.session.get('user_Id','')
            user_Role = request.session.get('Role','')
            if len(user_Id)>0 and len(user_Role)>0: #判断是否登录

                try:
                    user = User.objects.get(Id=user_Role)
                    if not args:#如果args没定义的话，直接验证默认权限
                        if user.Role == user_Role:
                          return func(request) #权限验证通过，继续执行视图
                        else:
                            return HttpResponseRedirect("/login/")#否则执行禁止视图
                    # else: #如果定义了就将用户的权限跟预定义的进行匹配
                    #     if user.auth_group in args or user.auth_group == 'admin':
                    #       return func(request)#权限验证通过，继续执行视图
                    #     else: #否则执行禁止视图
                    #       return denied(request)
                except:
                    return HttpResponseRedirect("/login/")
            else:#如果没登录就跳转到登录界面
              return HttpResponseRedirect("/login/")
        return _login
    return __auth