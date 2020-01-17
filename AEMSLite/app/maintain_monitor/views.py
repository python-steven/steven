from django.shortcuts import render
from django.shortcuts import render, redirect
# from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from app.login.models import User,Department,Customer,BudgetCodeForm,PartItem,PartItemResult,MaintenanceLog,Configuration,LocationLog
from app.login.views import Update_User_IsActivated
from django.views.generic.base import View
from django.db import connection
from django.http import HttpResponseRedirect,HttpResponse
from app import restful,mail
from app.access_control import access_control
from datetime import datetime,timedelta
from django.conf import settings
import random
import string
import os
import pytz
import time
from openpyxl import load_workbook,Workbook
import json

class maintain_monitor_info(View):
    @csrf_exempt
    def get(self,request):
        try:
            page = int(request.GET.get('page'))
            number = request.GET.get('num')
            start = datetime.now()
            start_select = (list(PartItem.objects.order_by("-TrnDate").filter(TrnDate__lte=start,UseStatus='normal').values("TrnDate")))[0]['TrnDate']
            end = start_select-timedelta(days=7)
            dict_data={}
            data_count = PartItem.objects.order_by("Id").filter(TrnDate__gte=end,UseStatus='normal').count()
            if number == "All":
                data = list(PartItem.objects.order_by("Id").filter(TrnDate__gte=end,UseStatus='normal').values())
                mt_count = list(Configuration.objects.filter(Type="mt_count").values("Max"))
                mt_date = list(Configuration.objects.filter(Type="mt_date").values("Max"))
                start_time = datetime.strptime(str(start).split(' ')[0], "%Y-%m-%d") #获取当前的日期
                #计算保养次数和保养日期的水准并加入到后面两组数据当中去做判断
                for i in range(0,len(data)):
                    if data[i]['LocationId'] != None:
                        data[i]['location'] = (list(LocationLog.objects.filter(Id=data[i]['LocationId']).values("Location")))[0]['Location']
                    else:
                        data[i]['location'] = ""
                    if data[i]['NextCheckDate'] == None:
                        data[i]['stand_date'] = "null"
                    else:
                        time_end = datetime.strptime(str(data[i]['NextCheckDate']).split(' ')[0],
                                                     "%Y-%m-%d")  # 获取数据表里面的日期数
                        days = time_end - start_time
                        data[i]['stand_date'] = days.days
                    if data[i]['NextCheckCount'] == 0:
                        data[i]['stand_count'] = 'null'
                    else:
                        data[i]['stand_count'] = data[i]['NextCheckCount'] - data[i]['UsedTimes']
                    # 判断是那种状态
                    if data[i]['WarningBeforeDays'] == None and data[i]['WarningBeforeTimes'] == None:
                        if data[i]['stand_date'] != "null" and data[i]['stand_count'] != "null":
                            if data[i]['stand_date'] > mt_date[0]['Max'] * data[i]['CheckCycle'] and data[i][
                                'stand_count'] > mt_count[0]['Max'] * data[i]['CheckCycleCount']:
                                data[i]['stand'] = 'normal'
                            if 0 <= data[i]['stand_date'] <= mt_date[0]['Max'] * data[i]['CheckCycle'] and 0 <= data[i][
                                'stand_count']:
                                data[i]['stand'] = 'warning'
                            if 0 <= data[i]['stand_count'] <= mt_count[0]['Max'] * data[i]['CheckCycleCount'] and 0 <= \
                                    data[i]['stand_date']:
                                data[i]['stand'] = 'warning'
                            if data[i]['stand_count'] < 0 or data[i]['stand_date'] < 0:
                                data[i]['stand'] = 'danger'
                        if data[i]['stand_date'] == "null" and data[i]['stand_count'] != "null":
                            if data[i]['stand_count'] > mt_count[0]['Max'] * data[i]['CheckCycleCount']:
                                data[i]['stand'] = 'normal'
                            if 0 <= data[i]['stand_count'] <= mt_count[0]['Max'] * data[i]['CheckCycleCount']:
                                data[i]['stand'] = 'warning'
                            if data[i]['stand_count'] < 0:
                                data[i]['stand'] = 'danger'
                        if data[i]['stand_date'] != "null" and data[i]['stand_count'] == "null":
                            if data[i]['stand_date'] > mt_date[0]['Max'] * data[i]['CheckCycle']:
                                data[i]['stand'] = 'normal'
                            if 0 <= data[i]['stand_date'] <= mt_date[0]['Max'] * data[i]['CheckCycle']:
                                data[i]['stand'] = 'warning'
                            if data[i]['stand_date'] < 0:
                                data[i]['stand'] = 'danger'
                        if data[i]['stand_date'] == "null" and data[i]['stand_count'] == "null":
                            data[i]['stand'] = 'none'
                    if data[i]['WarningBeforeDays'] != None and data[i]['WarningBeforeTimes'] == None:
                        if data[i]['stand_date'] != "null" and data[i]['stand_count'] != "null":
                            if data[i]['stand_date'] > data[i]['WarningBeforeDays'] and data[i][
                                'stand_count'] > mt_count[0]['Max'] * data[i]['CheckCycleCount']:
                                data[i]['stand'] = 'normal'
                            if 0 <= data[i]['stand_date'] <= data[i]['WarningBeforeDays'] and 0 <= data[i][
                                'stand_count']:
                                data[i]['stand'] = 'warning'
                            if 0 <= data[i]['stand_count'] <= mt_count[0]['Max'] * data[i]['CheckCycleCount'] and 0 <= \
                                    data[i]['stand_date']:
                                data[i]['stand'] = 'warning'
                            if data[i]['stand_count'] < 0 or data[i]['stand_date'] < 0:
                                data[i]['stand'] = 'danger'
                        if data[i]['stand_date'] == "null" and data[i]['stand_count'] != "null":
                            if data[i]['stand_count'] > mt_count[0]['Max'] * data[i]['CheckCycleCount']:
                                data[i]['stand'] = 'normal'
                            if 0 <= data[i]['stand_count'] <= mt_count[0]['Max'] * data[i]['CheckCycleCount']:
                                data[i]['stand'] = 'warning'
                            if data[i]['stand_count'] < 0:
                                data[i]['stand'] = 'danger'
                        if data[i]['stand_date'] != "null" and data[i]['stand_count'] == "null":
                            if data[i]['stand_date'] > data[i]['WarningBeforeDays']:
                                data[i]['stand'] = 'normal'
                            if 0 <= data[i]['stand_date'] <= data[i]['WarningBeforeDays']:
                                data[i]['stand'] = 'warning'
                            if data[i]['stand_date'] < 0:
                                data[i]['stand'] = 'danger'
                        if data[i]['stand_date'] == "null" and data[i]['stand_count'] == "null":
                            data[i]['stand'] = 'none'
                    if data[i]['WarningBeforeDays'] == None and data[i]['WarningBeforeTimes'] != None:
                        if data[i]['stand_date'] != "null" and data[i]['stand_count'] != "null":
                            if data[i]['stand_date'] > mt_date[0]['Max'] * data[i]['CheckCycle'] and data[i][
                                'stand_count'] > data[i]['WarningBeforeTimes']:
                                data[i]['stand'] = 'normal'
                            if 0 <= data[i]['stand_date'] <= mt_date[0]['Max'] * data[i]['CheckCycle'] and 0 <= data[i][
                                'stand_count']:
                                data[i]['stand'] = 'warning'
                            if 0 <= data[i]['stand_count'] <= data[i]['WarningBeforeTimes'] and 0 <= \
                                    data[i]['stand_date']:
                                data[i]['stand'] = 'warning'
                            if data[i]['stand_count'] < 0 or data[i]['stand_date'] < 0:
                                data[i]['stand'] = 'danger'
                        if data[i]['stand_date'] == "null" and data[i]['stand_count'] != "null":
                            if data[i]['stand_count'] > data[i]['WarningBeforeTimes']:
                                data[i]['stand'] = 'normal'
                            if 0 <= data[i]['stand_count'] <= data[i]['WarningBeforeTimes']:
                                data[i]['stand'] = 'warning'
                            if data[i]['stand_count'] < 0:
                                data[i]['stand'] = 'danger'
                        if data[i]['stand_date'] != "null" and data[i]['stand_count'] == "null":
                            if data[i]['stand_date'] > mt_date[0]['Max'] * data[i]['CheckCycle']:
                                data[i]['stand'] = 'normal'
                            if 0 <= data[i]['stand_date'] <= mt_date[0]['Max'] * data[i]['CheckCycle']:
                                data[i]['stand'] = 'warning'
                            if data[i]['stand_date'] < 0:
                                data[i]['stand'] = 'danger'
                        if data[i]['stand_date'] == "null" and data[i]['stand_count'] == "null":
                            data[i]['stand'] = 'none'
                    if data[i]['WarningBeforeDays'] != None and data[i]['WarningBeforeTimes'] != None:
                        if data[i]['stand_date'] != "null" and data[i]['stand_count'] != "null":
                            if data[i]['stand_date'] > data[i]['WarningBeforeDays'] and data[i][
                                'stand_count'] > data[i]['WarningBeforeTimes']:
                                data[i]['stand'] = 'normal'
                            if 0 <= data[i]['stand_date'] <= data[i]['WarningBeforeDays'] and 0 <= data[i][
                                'stand_count']:
                                data[i]['stand'] = 'warning'
                            if 0 <= data[i]['stand_count'] <= data[i]['WarningBeforeTimes'] and 0 <= \
                                    data[i]['stand_date']:
                                data[i]['stand'] = 'warning'
                            if data[i]['stand_count'] < 0 or data[i]['stand_date'] < 0:
                                data[i]['stand'] = 'danger'
                        if data[i]['stand_date'] == "null" and data[i]['stand_count'] != "null":
                            if data[i]['stand_count'] > data[i]['WarningBeforeTimes']:
                                data[i]['stand'] = 'normal'
                            if 0 <= data[i]['stand_count'] <= data[i]['WarningBeforeTimes']:
                                data[i]['stand'] = 'warning'
                            if data[i]['stand_count'] < 0:
                                data[i]['stand'] = 'danger'
                        if data[i]['stand_date'] != "null" and data[i]['stand_count'] == "null":
                            if data[i]['stand_date'] > data[i]['WarningBeforeDays']:
                                data[i]['stand'] = 'normal'
                            if 0 <= data[i]['stand_date'] <= data[i]['WarningBeforeDays']:
                                data[i]['stand'] = 'warning'
                            if data[i]['stand_date'] < 0:
                                data[i]['stand'] = 'danger'
                        if data[i]['stand_date'] == "null" and data[i]['stand_count'] == "null":
                            data[i]['stand'] = 'none'
                dict_data['data'] = data
                dict_data['page_count'] = data_count
            if number != "All":
                number = int(number)
                page_num = data_count // number  # 总共多少页
                if data_count % number > 0:
                    page_num = page_num + 1
                if page_num >= page:
                    data = list(PartItem.objects.order_by("Id").filter(TrnDate__gte=end,UseStatus='normal').values()[(page - 1) * number:number * page])
                    mt_date = list(Configuration.objects.filter(Type="mt_date").values("Max"))
                    mt_count = list(Configuration.objects.filter(Type="mt_count").values("Max"))
                    start_time = datetime.strptime(str(start).split(' ')[0], "%Y-%m-%d")  # 获取当前的日期
                    # 计算保养次数和保养日期的水准并加入到后面两组数据当中去做判断
                    for i in range(0, len(data)):
                        if data[i]['LocationId'] != None:
                            data[i]['location'] =(list(LocationLog.objects.filter(Id=data[i]['LocationId']).values("Location")))[0]['Location']
                        else:
                            data[i]['location'] = ""
                        if data[i]['NextCheckDate'] == None:
                            data[i]['stand_date'] = "null"
                        else:
                            time_end = datetime.strptime(str(data[i]['NextCheckDate']).split(' ')[0],
                                                         "%Y-%m-%d")  # 获取数据表里面的日期数
                            days = time_end - start_time
                            data[i]['stand_date'] = days.days
                        if data[i]['NextCheckCount'] == 0:
                            data[i]['stand_count'] = 'null'
                        else:
                            data[i]['stand_count'] = data[i]['NextCheckCount'] - data[i]['UsedTimes']
                        # 判断是那种状态
                        if data[i]['WarningBeforeDays'] == None and data[i]['WarningBeforeTimes'] == None:
                            if data[i]['stand_date'] != "null" and data[i]['stand_count'] != "null":
                                if data[i]['stand_date'] > mt_date[0]['Max'] * data[i]['CheckCycle'] and data[i][
                                    'stand_count'] > mt_count[0]['Max'] * data[i]['CheckCycleCount']:
                                    data[i]['stand'] = 'normal'
                                if 0 <= data[i]['stand_date'] <= mt_date[0]['Max'] * data[i]['CheckCycle'] and 0 <= \
                                        data[i][
                                            'stand_count']:
                                    data[i]['stand'] = 'warning'
                                if 0 <= data[i]['stand_count'] <= mt_count[0]['Max'] * data[i][
                                    'CheckCycleCount'] and 0 <= \
                                        data[i]['stand_date']:
                                    data[i]['stand'] = 'warning'
                                if data[i]['stand_count'] < 0 or data[i]['stand_date'] < 0:
                                    data[i]['stand'] = 'danger'
                            if data[i]['stand_date'] == "null" and data[i]['stand_count'] != "null":
                                if data[i]['stand_count'] > mt_count[0]['Max'] * data[i]['CheckCycleCount']:
                                    data[i]['stand'] = 'normal'
                                if 0 <= data[i]['stand_count'] <= mt_count[0]['Max'] * data[i]['CheckCycleCount']:
                                    data[i]['stand'] = 'warning'
                                if data[i]['stand_count'] < 0:
                                    data[i]['stand'] = 'danger'
                            if data[i]['stand_date'] != "null" and data[i]['stand_count'] == "null":
                                if data[i]['stand_date'] > mt_date[0]['Max'] * data[i]['CheckCycle']:
                                    data[i]['stand'] = 'normal'
                                if 0 <= data[i]['stand_date'] <= mt_date[0]['Max'] * data[i]['CheckCycle']:
                                    data[i]['stand'] = 'warning'
                                if data[i]['stand_date'] < 0:
                                    data[i]['stand'] = 'danger'
                            if data[i]['stand_date'] == "null" and data[i]['stand_count'] == "null":
                                data[i]['stand'] = 'none'
                        if data[i]['WarningBeforeDays'] != None and data[i]['WarningBeforeTimes'] == None:
                            if data[i]['stand_date'] != "null" and data[i]['stand_count'] != "null":
                                if data[i]['stand_date'] > data[i]['WarningBeforeDays'] and data[i][
                                    'stand_count'] > mt_count[0]['Max'] * data[i]['CheckCycleCount']:
                                    data[i]['stand'] = 'normal'
                                if 0 <= data[i]['stand_date'] <= data[i]['WarningBeforeDays'] and 0 <= data[i][
                                    'stand_count']:
                                    data[i]['stand'] = 'warning'
                                if 0 <= data[i]['stand_count'] <= mt_count[0]['Max'] * data[i][
                                    'CheckCycleCount'] and 0 <= \
                                        data[i]['stand_date']:
                                    data[i]['stand'] = 'warning'
                                if data[i]['stand_count'] < 0 or data[i]['stand_date'] < 0:
                                    data[i]['stand'] = 'danger'
                            if data[i]['stand_date'] == "null" and data[i]['stand_count'] != "null":
                                if data[i]['stand_count'] > mt_count[0]['Max'] * data[i]['CheckCycleCount']:
                                    data[i]['stand'] = 'normal'
                                if 0 <= data[i]['stand_count'] <= mt_count[0]['Max'] * data[i]['CheckCycleCount']:
                                    data[i]['stand'] = 'warning'
                                if data[i]['stand_count'] < 0:
                                    data[i]['stand'] = 'danger'
                            if data[i]['stand_date'] != "null" and data[i]['stand_count'] == "null":
                                if data[i]['stand_date'] > data[i]['WarningBeforeDays']:
                                    data[i]['stand'] = 'normal'
                                if 0 <= data[i]['stand_date'] <= data[i]['WarningBeforeDays']:
                                    data[i]['stand'] = 'warning'
                                if data[i]['stand_date'] < 0:
                                    data[i]['stand'] = 'danger'
                            if data[i]['stand_date'] == "null" and data[i]['stand_count'] == "null":
                                data[i]['stand'] = 'none'
                        if data[i]['WarningBeforeDays'] == None and data[i]['WarningBeforeTimes'] != None:
                            if data[i]['stand_date'] != "null" and data[i]['stand_count'] != "null":
                                if data[i]['stand_date'] > mt_date[0]['Max'] * data[i]['CheckCycle'] and data[i][
                                    'stand_count'] > data[i]['WarningBeforeTimes']:
                                    data[i]['stand'] = 'normal'
                                if 0 <= data[i]['stand_date'] <= mt_date[0]['Max'] * data[i]['CheckCycle'] and 0 <= \
                                        data[i][
                                            'stand_count']:
                                    data[i]['stand'] = 'warning'
                                if 0 <= data[i]['stand_count'] <= data[i]['WarningBeforeTimes'] and 0 <= \
                                        data[i]['stand_date']:
                                    data[i]['stand'] = 'warning'
                                if data[i]['stand_count'] < 0 or data[i]['stand_date'] < 0:
                                    data[i]['stand'] = 'danger'
                            if data[i]['stand_date'] == "null" and data[i]['stand_count'] != "null":
                                if data[i]['stand_count'] > data[i]['WarningBeforeTimes']:
                                    data[i]['stand'] = 'normal'
                                if 0 <= data[i]['stand_count'] <= data[i]['WarningBeforeTimes']:
                                    data[i]['stand'] = 'warning'
                                if data[i]['stand_count'] < 0:
                                    data[i]['stand'] = 'danger'
                            if data[i]['stand_date'] != "null" and data[i]['stand_count'] == "null":
                                if data[i]['stand_date'] > mt_date[0]['Max'] * data[i]['CheckCycle']:
                                    data[i]['stand'] = 'normal'
                                if 0 <= data[i]['stand_date'] <= mt_date[0]['Max'] * data[i]['CheckCycle']:
                                    data[i]['stand'] = 'warning'
                                if data[i]['stand_date'] < 0:
                                    data[i]['stand'] = 'danger'
                            if data[i]['stand_date'] == "null" and data[i]['stand_count'] == "null":
                                data[i]['stand'] = 'none'
                        if data[i]['WarningBeforeDays'] != None and data[i]['WarningBeforeTimes'] != None:
                            if data[i]['stand_date'] != "null" and data[i]['stand_count'] != "null":
                                if data[i]['stand_date'] > data[i]['WarningBeforeDays'] and data[i][
                                    'stand_count'] > data[i]['WarningBeforeTimes']:
                                    data[i]['stand'] = 'normal'
                                if 0 <= data[i]['stand_date'] <= data[i]['WarningBeforeDays'] and 0 <= data[i][
                                    'stand_count']:
                                    data[i]['stand'] = 'warning'
                                if 0 <= data[i]['stand_count'] <= data[i]['WarningBeforeTimes'] and 0 <= \
                                        data[i]['stand_date']:
                                    data[i]['stand'] = 'warning'
                                if data[i]['stand_count'] < 0 or data[i]['stand_date'] < 0:
                                    data[i]['stand'] = 'danger'
                            if data[i]['stand_date'] == "null" and data[i]['stand_count'] != "null":
                                if data[i]['stand_count'] > data[i]['WarningBeforeTimes']:
                                    data[i]['stand'] = 'normal'
                                if 0 <= data[i]['stand_count'] <= data[i]['WarningBeforeTimes']:
                                    data[i]['stand'] = 'warning'
                                if data[i]['stand_count'] < 0:
                                    data[i]['stand'] = 'danger'
                            if data[i]['stand_date'] != "null" and data[i]['stand_count'] == "null":
                                if data[i]['stand_date'] > data[i]['WarningBeforeDays']:
                                    data[i]['stand'] = 'normal'
                                if 0 <= data[i]['stand_date'] <= data[i]['WarningBeforeDays']:
                                    data[i]['stand'] = 'warning'
                                if data[i]['stand_date'] < 0:
                                    data[i]['stand'] = 'danger'
                            if data[i]['stand_date'] == "null" and data[i]['stand_count'] == "null":
                                data[i]['stand'] = 'none'
                    dict_data['data'] = data
                    dict_data['page_count'] = page_num

            count =list(Configuration.objects.filter(Type="mt_count").values("Max"))
            date = list(Configuration.objects.filter(Type="mt_date").values("Max"))
            # 饼状图和表格需要的判断依据数据
            sql = 'SELECT "PartName", COUNT("PartName") FROM "PartItem" where to_char("TrnDate",\'yyyy-MM-dd HH24:MI:SS\') >=\' '+ end.strftime("%Y-%m-%d %H:%M:%S")+'\' and "UseStatus"=\'normal\''
            # 柱状图需要的数据的原生语句
            tab_sql = 'SELECT "Maintainer", COUNT("PartName") FROM "PartItem" where to_char("TrnDate",\'yyyy-MM-dd HH24:MI:SS\') >=\' '+ end.strftime("%Y-%m-%d %H:%M:%S")+'\' and "UseStatus"=\'normal\''

            #饼状图需要获取的数值的sql
            sql1 =sql_select(sql,"正常",count[0]["Max"],date[0]["Max"])+' GROUP BY "PartName"'
            sql2 =sql_select(sql,"预警",count[0]["Max"],date[0]["Max"])+' GROUP BY "PartName"'
            sql3 =sql_select(sql,"超标",count[0]["Max"],date[0]["Max"])+' GROUP BY "PartName"'
            sql_no_nextcheckdate ='SELECT COUNT("PartName") FROM "PartItem" where to_char("TrnDate",\'yyyy-MM-dd HH24:MI:SS\') >=\' '+ end.strftime("%Y-%m-%d %H:%M:%S")+'\' and "UseStatus"=\'normal\' and "NextCheckDate" IS NULL AND "NextCheckCount" =0;'

            #  柱状图 正常
            tab_normal = sql_select(tab_sql,"正常",count[0]["Max"],date[0]["Max"])+' GROUP BY "Maintainer"'
            tab_warning = sql_select(tab_sql,"预警",count[0]["Max"],date[0]["Max"])+' GROUP BY "Maintainer"'
            tab_danger = sql_select(tab_sql,"超标",count[0]["Max"],date[0]["Max"])+' GROUP BY "Maintainer"'
            tab_no_nextcheckdate = sql_no_nextcheckdate

            cur = connection.cursor()
            cur.execute(sql1)
            normal = cur.fetchall()
            cur.execute(sql2)
            warning = cur.fetchall()
            cur.execute(sql3)
            danger = cur.fetchall()
            cur.execute(sql_no_nextcheckdate)
            info = cur.fetchall()
            info = [('None',info[0][0])]

            cur = connection.cursor()
            cur.execute(tab_normal)
            tab_normal = cur.fetchall()
            cur.execute(tab_warning)
            tab_warning = cur.fetchall()
            cur.execute(tab_danger)
            tab_danger = cur.fetchall()
            cur.execute(tab_no_nextcheckdate)
            tab_info = cur.fetchall()
            tab_info =[('None',tab_info[0][0])]

            tab_data =tab_query_way(tab_normal,tab_warning,tab_danger,tab_info)
            dict_data['normal'] = normal
            dict_data['warning'] = warning
            dict_data['danger'] = danger
            dict_data['None'] = info
            dict_data['tab_data'] = tab_data
            dict_data['select_start'] = start_select
            dict_data['select_end'] = end
            return restful.ok(data=dict_data)
        except Exception as e:
            return restful.params_error(message=repr(e))
    # 设置预警次数和预警天数
    @csrf_exempt
    def post(self,request):
        try:
            user_id = request.session.get('user_Id', '')
            admin_user = list(User.objects.filter(Id=user_id).values("Role"))
            if admin_user[0]["Role"] == "admin":
                maintain_count = request.POST['maintain_count']
                maintain_date = request.POST['maintain_date']
                maintain_receiver = request.POST.getlist('maintain_receiver[]')
                maintain_receiver = list(maintain_receiver)
                parameter_count = Configuration.objects.filter(Type="mt_count")
                parameter_date = Configuration.objects.filter(Type="mt_date")
                mail_receiver_count = ",".join(maintain_receiver)
                if parameter_count and parameter_date:
                    Configuration.objects.filter(Type="mt_count").update(Max=maintain_count,Min=0,Reminders=mail_receiver_count)
                    Configuration.objects.filter(Type="mt_date").update(Max=maintain_date,Min=0,Reminders=mail_receiver_count)
                    return restful.ok(message="setup parameter success")
                else:
                    Configuration.objects.create(Type="mt_count", Max=maintain_count, Min=0, Reminders=mail_receiver_count)
                    Configuration.objects.create(Type="mt_date", Max=maintain_date, Min=0, Reminders=mail_receiver_count)
                    return restful.ok(message="setup parameter create success")
            else:
                return restful.params_error(message="Your authority is not admin")
        except Exception as e:
            return restful.params_error(message=repr(e))
#设定之前 先查询 是否有设定。获取之前的设定参数并显示到页面上去
@access_control
def setup_range_before(request):
    if request.method == "GET":
        try:
            parameter_count = Configuration.objects.filter(Type="mt_count").count()
            parameter_date = Configuration.objects.filter(Type="mt_date").count()
            data_user = list(User.objects.filter(Role="admin").values("Email")) + list(User.objects.filter(Role="equipment_room").values("Email"))
            if parameter_count !=0 and parameter_date !=0:
                data1 = list(Configuration.objects.filter(Type="mt_count").values('Type','Max'))
                data2 = list(Configuration.objects.filter(Type="mt_date").values('Type','Max',"Reminders"))
                data = data1+data2+data_user
                return restful.ok(data=data)
            else:
                data = [{'Type':"mt_count",'Max':10}]+[{'Type':"mt_date",'Max':10,'Reminders':'Steven_X_Xu'}]+data_user
                return restful.ok(data=data)
        except Exception as e:
            return restful.params_error(message=repr(e))

def dictdata(sqldict):
    cur = connection.cursor()
    cur.execute(sqldict, None)
    desc = cur.description
    return [dict(zip([col[0] for col in desc], row)) for row in cur.fetchall()]
#对保养得数据进行查询功能的实现
@access_control
def maintain_query(request):
    if request.method == "POST":
        try:
            start = datetime.now()
            page = int(request.POST.get('page',''))
            number = request.POST.get('num','')
            sn = str(request.POST.get('sn',''))
            part_name = request.POST.get('partname','')
            status = request.POST.get('status','')
            s_time = request.POST.get('s_time','')
            e_time = request.POST.get('e_time','')
            user = str(request.POST.get('user',''))
            locationId = str(request.POST.get('location',''))
            dict_data = {}
            normal = []
            warning = []
            danger = []
            info = []
            sql_count = 'select count(*) FROM "PartItem" where "UseStatus"=\'normal\' '      #计算 查询数据的总的多少条数据
            #查询条件的数据
            sql = 'select "Id","SN","PartName","CheckCycleCount","UsedTimes","CheckCycle","NextCheckDate","Maintainer","TrnDate","NextCheckCount","WarningBeforeDays","WarningBeforeTimes","LocationId" FROM "PartItem" WHERE  "UseStatus"=\'normal\' '
            visual_sql = 'SELECT "PartName", COUNT("PartName") FROM "PartItem" where  "UseStatus"=\'normal\' ' # 饼图数据按正常和预警和超标分类计算数量
            tab_sql = 'SELECT "Maintainer", COUNT("PartName") FROM "PartItem" where  "UseStatus"=\'normal\' '  # 柱状图的数据 按条件查询出来之后再 按负责人分类
            mt_count = list(Configuration.objects.filter(Type="mt_count").values("Max"))
            mt_date  = list(Configuration.objects.filter(Type="mt_date").values("Max"))
            # c_date = (start + timedelta(days=int(date.Max))).strftime("%Y-%m-%d")  # 拿到当前的时间+预警天数得到日期的范围是（当前时间,c_date）
            start_select = (list(PartItem.objects.order_by("-TrnDate").filter(TrnDate__lte=start,UseStatus="normal").values("TrnDate")))[0]['TrnDate']  # 预设区间的筛选，回滚点击视图的函数的加载
            end = start_select - timedelta(days=7)
            callback_sql =' AND to_char("TrnDate",\'yyyy-MM-dd HH24:MI:SS\') >= \'' + end.strftime("%Y-%m-%d %H:%M:%S") + '\''
            if sn == "" and part_name == "" and s_time == "" and e_time == "" and user == "" and status =="":
                sql =sql + callback_sql
                sql_count =sql_count + callback_sql
                visual_sql =visual_sql + callback_sql
                tab_sql =tab_sql + callback_sql
            if status != "":
                if sn != "":
                    sql_count = sql_count + 'AND "SN" = \'' + sn + '\''
                    sql = sql + 'AND "SN" = \'' + sn + '\''
                    visual_sql = visual_sql + 'AND "SN" = \'' + sn + '\''
                    tab_sql = tab_sql + 'AND "SN" = \'' + sn + '\''
                if part_name != "":
                    sql_count = sql_count + 'AND "PartName" ilike \'%{0}%\''.format(part_name)
                    sql = sql + 'AND "PartName" ilike \'%{0}%\''.format(part_name)
                    visual_sql = visual_sql + 'AND "PartName" ilike \'%{0}%\''.format(part_name)
                    tab_sql = tab_sql + 'AND "PartName" ilike \'%{0}%\''.format(part_name)
                if s_time != "":
                    sql_count = sql_count + 'AND "TrnDate" >= \'{0}\''.format(s_time)
                    sql = sql + 'AND "TrnDate" >= \'{0}\''.format(s_time)
                    visual_sql = visual_sql + 'AND "TrnDate" >= \'{0}\''.format(s_time)
                    tab_sql = tab_sql + 'AND "TrnDate" >= \'{0}\''.format(s_time)
                if e_time != "":
                    sql_count = sql_count + 'AND "TrnDate" <= \'%{0}%\''.format(e_time)
                    sql = sql + 'AND "TrnDate" <= \'%{0}%\''.format(e_time)
                    visual_sql = visual_sql + 'AND "TrnDate" <= \'%{0}%\''.format(e_time)
                    tab_sql = tab_sql + 'AND "TrnDate" <= \'%{0}%\''.format(e_time)
                if user != "":
                    sql_count = sql_count + 'AND ("Maintainer" = \'{0}\''.format(user)+ ' OR "SubMaintainers" ilike \'%'+ str(user) +'%\')'
                    sql = sql + 'AND ("Maintainer" = \'{0}\''.format(user)+ ' OR "SubMaintainers" ilike \'%'+ str(user) +'%\')'
                    visual_sql = visual_sql + 'AND ("Maintainer" = \'{0}\''.format(user)+ ' OR "SubMaintainers" ilike \'%'+ str(user) +'%\')'
                    tab_sql = tab_sql + 'AND ("Maintainer" = \'{0}\''.format(user)+ ' OR "SubMaintainers" ilike \'%'+ str(user) +'%\')'
                if locationId != "":
                    sql_count = sql_count + 'AND "LocationId" = \'{0}\''.format(locationId)
                    sql = sql + 'AND "LocationId" = \'{0}\''.format(locationId)
                    visual_sql = visual_sql + 'AND "LocationId" = \'{0}\''.format(locationId)
                    tab_sql = tab_sql + 'AND "LocationId" = \'{0}\''.format(locationId)
                if status == "正常":
                    sql_count = sql_select(sql_count,"正常",mt_count[0]['Max'],mt_date[0]['Max'])
                    sql       = sql_select(sql,"正常",mt_count[0]['Max'],mt_date[0]['Max'])
                    visual_sql_normal = sql_select(visual_sql,"正常",mt_count[0]['Max'],mt_date[0]['Max'])+' GROUP BY "PartName";'
                    tab_normal = sql_select(tab_sql,"正常",mt_count[0]['Max'],mt_date[0]['Max'])+' GROUP BY "Maintainer";'
                    cur = connection.cursor()
                    cur.execute(visual_sql_normal)
                    normal = cur.fetchall()

                    cur = connection.cursor()
                    cur.execute(tab_normal)
                    tab_normal =cur.fetchall()
                    tab_normal = tab_query_way(tab_normal,[],[],[])
                    dict_data['tab_data'] = tab_normal
                if status == "预警":
                    sql_count = sql_select(sql_count, "预警", mt_count[0]['Max'],mt_date[0]['Max'])
                    sql = sql_select(sql, "预警", mt_count[0]['Max'],mt_date[0]['Max'])
                    visual_sql_waring = sql_select(visual_sql, "预警", mt_count[0]['Max'],mt_date[0]['Max']) + ' GROUP BY "PartName";'
                    tab_warning = sql_select(tab_sql, "预警", mt_count[0]['Max'],mt_date[0]['Max']) + '  GROUP BY "Maintainer";'
                    cur = connection.cursor()
                    cur.execute(visual_sql_waring)
                    warning = cur.fetchall()

                    cur.execute(tab_warning)
                    tab_warning = cur.fetchall()
                    tab_warning = tab_query_way([],tab_warning,[],[])
                    dict_data['tab_data'] = tab_warning
                if status == "超标":
                    sql_count = sql_select(sql_count, "超标", mt_count[0]['Max'],mt_date[0]['Max'])
                    sql = sql_select(sql, "超标", mt_count[0]['Max'],mt_date[0]['Max'])
                    visual_sql_danger = sql_select(visual_sql, "超标", mt_count[0]['Max'],mt_date[0]['Max']) + ' GROUP BY "PartName";'
                    tab_danger = sql_select(tab_sql, "超标", mt_count[0]['Max'],mt_date[0]['Max']) + '  GROUP BY "Maintainer";'
                    cur = connection.cursor()
                    cur.execute(visual_sql_danger)
                    danger = cur.fetchall()

                    cur.execute(tab_danger)
                    tab_danger = cur.fetchall()
                    tab_danger = tab_query_way([], [], tab_danger,[])
                    dict_data['tab_data'] = tab_danger
                if status == "未设定":
                    sql_count = sql_select(sql_count,"未设定",mt_count[0]['Max'],mt_date[0]['Max'])
                    sql = sql_select(sql,"未设定",mt_count[0]['Max'],mt_date[0]['Max'])
                    visual_sql_no_checkdate = 'SELECT COUNT("PartName") FROM "PartItem" where to_char("TrnDate",\'yyyy-MM-dd HH24:MI:SS\') >=\' '+ end.strftime("%Y-%m-%d %H:%M:%S")+'\' and "UseStatus"=\'normal\' and "NextCheckDate" IS NULL AND "NextCheckCount" =0;'

                    cur = connection.cursor()
                    cur.execute(visual_sql_no_checkdate)
                    visual_info = cur.fetchall()
                    info = [('None',visual_info[0][0])]
                    dict_data['None'] = info
                    tab_info = tab_query_way([], [], [],info)
                    dict_data['tab_data'] = tab_info
                dict_data['None'] = info
                dict_data['normal'] = normal
                dict_data['warning'] = warning
                dict_data['danger'] = danger
                cur = connection.cursor()
                cur.execute(sql_count)
                count = cur.fetchall()  # 数量的总数
                if number == "All":
                    # cur = connection.cursor()
                    # cur.execute(sql)
                    data = dictdata(sql)
                    # data = tidy_dict(data)
                    # 针对data 进行判断和算出次数水平和周期天数
                    # start_time = datetime.strptime(str(start).split(' ')[0], "%Y-%m-%d")  # 获取当前的日期
                    # for i in range(0, len(data)):
                    #     data[i] = list(data[i])
                    #     if data[i][6] == None:
                    #         data[i].append("null")
                    #     else:
                    #         time_end = datetime.strptime(str(data[i][6]).split(' ')[0],"%Y-%m-%d")  # 获取数据表里面的日期数
                    #         days = time_end - start_time
                    #         data[i].append(days.days)
                    #     if data[i][9] == 0:
                    #         data[i].append("null")
                    #     else:
                    #         data[i].append( data[i][9] - data[i][4])
                    #     # 判断是那种状态
                    #     if data[i][10] != "null" and data[i][11] != "null":
                    #         if data[i][10] > mt_date[0]['Max'] * data[i][5] and data[i][11] > mt_count[0]['Max'] * \
                    #                 data[i][3]:
                    #             data[i].append('normal')
                    #         if 0 <= data[i][10] <= mt_date[0]['Max'] * data[i][5] and 0 <= data[i][11]:
                    #             data[i].append('warning')
                    #         if 0 <= data[i][11] <= mt_count[0]['Max'] * data[i][3] and 0 <= data[i][10]:
                    #             data[i].append('warning')
                    #         if data[i][10] < 0 or data[i][11] < 0:
                    #             data[i].append('danger')
                    #     if data[i][10] == "null" and data[i][11] != "null":
                    #         if data[i][11] > mt_count[0]['Max'] * data[i][3]:
                    #             data[i].append('normal')
                    #         if 0 <= data[i][11] <= mt_count[0]['Max'] * data[i][3]:
                    #             data[i].append('warning')
                    #         if data[i][11] < 0:
                    #             data[i].append('danger')
                    #     if data[i][10] != "null" and data[i][11] == "null":
                    #         if data[i][10] > mt_date[0]['Max'] * data[i][5]:
                    #             data[i].append('normal')
                    #         if 0 <= data[i][10] <= mt_date[0]['Max'] * data[i][5]:
                    #             data[i].append('warning')
                    #         if data[i][10] < 0:
                    #             data[i].append('danger')
                    #     if data[i][10] == "null" and data[i][11] == "null":
                    #         data[i].append('none')
                    dict_data['data'] = data
                    dict_data['page_count'] = count[0][0]
                if number != "All":
                    number = int(number)
                    count_page = count[0][0] // int(number)  # 总数除以一页显示多少条，得到总的页数
                    if count[0][0] % number > 0:
                        count_page += 1
                    if page <= count_page:
                        sql = sql + ' order by "Id" limit ' + str(number) + ' offset ' + str((page - 1)*number)
                        data = dictdata(sql)
                        data = tidy_dict(data)
                        # cur = connection.cursor()
                        # cur.execute(sql)
                        # data = cur.fetchall()
                        # data = tidy(data)
                        # 针对data 进行判断和算出次数水平和周期天数
                        # start_time = datetime.strptime(str(start).split(' ')[0], "%Y-%m-%d")  # 获取当前的日期
                        # for i in range(0, len(data)):
                        #     data[i] = list(data[i])
                        #     data[i][3] =int(data[i][3])
                        #     data[i][5] =int(data[i][5])
                        #     if data[i][6] == None:
                        #         data[i].append("null")
                        #     else:
                        #         time_end = datetime.strptime(str(data[i][6]).split(' ')[0], "%Y-%m-%d")  # 获取数据表里面的日期数
                        #         days = time_end - start_time
                        #         data[i].append(days.days)
                        #     if data[i][9] == 0:
                        #         data[i].append("null")
                        #     else:
                        #         data[i].append(data[i][9] - data[i][4])
                        #     # 判断是那种状态
                        #     if data[i][10] != "null" and data[i][11] != "null":
                        #         if data[i][10]>mt_date[0]['Max']*data[i][5] and data[i][11]>mt_count[0]['Max']*data[i][3]:
                        #             data[i].append('normal')
                        #         if 0 <= data[i][10] <= mt_date[0]['Max'] * data[i][5] and 0 <= data[i][11]:
                        #             data[i].append('warning')
                        #         if 0 <= data[i][11] <= mt_count[0]['Max'] * data[i][3] and 0 <= data[i][10]:
                        #             data[i].append('warning')
                        #         if data[i][10] < 0 or data[i][11] < 0:
                        #             data[i].append('danger')
                        #     if data[i][10] == "null" and data[i][11] != "null":
                        #         if data[i][11] > mt_count[0]['Max'] * data[i][3]:
                        #             data[i].append('normal')
                        #         if 0 <= data[i][11] <= mt_count[0]['Max'] * data[i][3]:
                        #             data[i].append('warning')
                        #         if data[i][11] < 0:
                        #             data[i].append('danger')
                        #     if data[i][10] != "null" and data[i][11] == "null":
                        #         if data[i][10] > mt_date[0]['Max'] * data[i][5]:
                        #             data[i].append('normal')
                        #         if 0 <= data[i][10] <= mt_date[0]['Max'] * data[i][5]:
                        #             data[i].append('warning')
                        #         if data[i][10] < 0:
                        #             data[i].append('danger')
                        #     if data[i][10] == "null" and data[i][11] == "null":
                        #         data[i].append('none')
                        dict_data['data'] = data
                        dict_data['page_count'] = count_page
                    if page > count_page:
                        sql = sql + ' order by "Id" limit ' + str(number) + ' offset ' + str(0)
                        data = dictdata(sql)
                        data = tidy_dict(data)
                        # cur = connection.cursor()
                        # cur.execute(sql)
                        # data = cur.fetchall()
                        # data = tidy(data)
                        # 针对data 进行判断和算出次数水平和周期天数
                        # start_time = datetime.strptime(str(start).split(' ')[0], "%Y-%m-%d")  # 获取当前的日期
                        # for i in range(0, len(data)):
                        #     data[i] = list(data[i])
                        #     if data[i][6] == None:
                        #         data[i].append("null")
                        #     else:
                        #         time_end = datetime.strptime(str(data[i][6]).split(' ')[0], "%Y-%m-%d")  # 获取数据表里面的日期数
                        #         days = time_end - start_time
                        #         data[i].append(days.days)
                        #     if data[i][9] == 0:
                        #         data[i].append("null")
                        #     else:
                        #         data[i].append(data[i][9] - data[i][4])
                        #     # 判断是那种状态
                        #     if data[i][10] != "null" and data[i][11] != "null":
                        #         if data[i][10] > mt_date[0]['Max'] * data[i][5] and data[i][11] > mt_count[0]['Max'] * \
                        #                 data[i][3]:
                        #             data[i].append('normal')
                        #         if 0 <= data[i][10] <= mt_date[0]['Max'] * data[i][5] and 0 <= data[i][11]:
                        #             data[i].append('warning')
                        #         if 0 <= data[i][11] <= mt_count[0]['Max'] * data[i][3] and 0 <= data[i][10]:
                        #             data[i].append('warning')
                        #         if data[i][10] < 0 or data[i][11] < 0:
                        #             data[i].append('danger')
                        #     if data[i][10] == "null" and data[i][11] != "null":
                        #         if data[i][11] > mt_count[0]['Max'] * data[i][3]:
                        #             data[i].append('normal')
                        #         if 0 <= data[i][11] <= mt_count[0]['Max'] * data[i][3]:
                        #             data[i].append('warning')
                        #         if data[i][11] < 0:
                        #             data[i].append('danger')
                        #     if data[i][10] != "null" and data[i][11] == "null":
                        #         if data[i][10] > mt_date[0]['Max'] * data[i][5]:
                        #             data[i].append('normal')
                        #         if 0 <= data[i][10] <= mt_date[0]['Max'] * data[i][5]:
                        #             data[i].append('warning')
                        #         if data[i][10] < 0:
                        #             data[i].append('danger')
                        #     if data[i][10] == "null" and data[i][11] == "null":
                        #         data[i].append('none')
                        dict_data['data'] = data
                        dict_data['page_count'] = count_page
                return restful.ok(data=dict_data)
                #
                # dict_data['normal'] = normal
                # dict_data['warning'] = warning
                # dict_data['danger'] = danger
                # dict_data['None'] = info
            else:
                if sn != "":
                    sql_count = sql_count + 'AND "SN" = \'' + sn + '\''
                    sql = sql + 'AND "SN" = \'' + sn + '\''
                    visual_sql = visual_sql + 'AND "SN" = \'' + sn + '\''
                    tab_sql = tab_sql + 'AND "SN" = \'' + sn + '\''
                if part_name != "":
                    sql_count = sql_count + 'AND "PartName" ilike \'%{0}%\''.format(part_name)
                    sql = sql + 'AND "PartName" ilike \'%{0}%\''.format(part_name)
                    visual_sql = visual_sql + 'AND "PartName" ilike \'%{0}%\''.format(part_name)
                    tab_sql = tab_sql + 'AND "PartName" ilike \'%{0}%\''.format(part_name)
                if s_time != "":
                    sql_count = sql_count + 'AND "TrnDate" >= \'{0}\''.format(s_time)
                    sql = sql + 'AND "TrnDate" >= \'{0}\''.format(s_time)
                    visual_sql = visual_sql + 'AND "TrnDate" >= \'{0}\''.format(s_time)
                    tab_sql = tab_sql + 'AND "TrnDate" >= \'{0}\''.format(s_time)
                if e_time != "":
                    sql_count = sql_count + 'AND "TrnDate" <= \'{0}\''.format(e_time)
                    sql = sql + 'AND "TrnDate" <= \'{0}\''.format(e_time)
                    visual_sql = visual_sql + 'AND "TrnDate" <= \'{0}\''.format(e_time)
                    tab_sql = tab_sql + 'AND "TrnDate" <= \'{0}\''.format(e_time)
                if user != "":
                    sql_count = sql_count + ' AND ("Maintainer" = \'{0}\' '.format(user)+ ' OR "SubMaintainers" ilike \'%'+ str(user) +'%\')'
                    sql = sql + ' AND ("Maintainer" = \'{0}\''.format(user)+ ' OR "SubMaintainers" ilike \'%'+ str(user) +'%\')'
                    visual_sql = visual_sql + ' AND ("Maintainer" = \'{0}\''.format(user)+ ' OR "SubMaintainers" ilike \'%'+ str(user) +'%\')'
                    tab_sql = tab_sql + ' AND ("Maintainer" = \'{0}\''.format(user)+ ' OR "SubMaintainers" ilike \'%'+ str(user) +'%\')'
                if locationId != "":
                    sql_count = sql_count + 'AND "LocationId" = \'{0}\''.format(locationId)
                    sql = sql + 'AND "LocationId" = \'{0}\''.format(locationId)
                    visual_sql = visual_sql + 'AND "LocationId" = \'{0}\''.format(locationId)
                    tab_sql = tab_sql + 'AND "LocationId" = \'{0}\''.format(locationId)
                #normal
                visual_sql_normal = sql_select(visual_sql,"正常",mt_count[0]['Max'],mt_date[0]['Max'])+' GROUP BY "PartName"'
                tab_normal        = sql_select(tab_sql,"正常",mt_count[0]['Max'],mt_date[0]['Max'])+' GROUP BY "Maintainer"'
                #warning
                visual_sql_waring = sql_select(visual_sql,"预警",mt_count[0]['Max'],mt_date[0]['Max'])+' GROUP BY "PartName"'
                tab_warning       = sql_select(tab_sql,"预警",mt_count[0]['Max'],mt_date[0]['Max'])+' GROUP BY "Maintainer"'
                #danger
                visual_sql_danger = sql_select(visual_sql, "超标", mt_count[0]['Max'],mt_date[0]['Max'])+' GROUP BY "PartName"'
                tab_danger        = sql_select(tab_sql, "超标", mt_count[0]['Max'], mt_date[0]['Max'])+' GROUP BY "Maintainer"'
                #None
                visual_sql_no_checkdate = sql_count+'  and "UseStatus"=\'normal\' AND "NextCheckDate" IS NULL AND "NextCheckCount" =0;'
                cur = connection.cursor()
                cur.execute(visual_sql_normal)
                normal = cur.fetchall()
                cur.execute(visual_sql_waring)
                warning = cur.fetchall()
                cur.execute(visual_sql_danger)
                danger = cur.fetchall()
                cur.execute(visual_sql_no_checkdate)
                info = cur.fetchall()
                info =[('None',info[0][0])]

                cur.execute(tab_normal)
                tab_normal = cur.fetchall()
                cur.execute(tab_warning)
                tab_warning = cur.fetchall()
                cur.execute(tab_danger)
                tab_danger = cur.fetchall()
                tab_info = info
                tab_data = tab_query_way(tab_normal,tab_warning, tab_danger,tab_info )
                dict_data['tab_data'] = tab_data

                dict_data['normal'] = normal
                dict_data['warning'] = warning
                dict_data['danger'] = danger
                dict_data['None'] = info
                cur = connection.cursor()
                cur.execute(sql_count)
                count = cur.fetchall()  # 数量的总数
                if number == "All":
                    data = dictdata(sql)
                    data = tidy_dict(data)
                    # cur = connection.cursor()
                    # cur.execute(sql)
                    # data = cur.fetchall()
                    # data = tidy(data)
                    # 针对data 进行判断和算出次数水平和周期天数
                    # start_time = datetime.strptime(str(start).split(' ')[0], "%Y-%m-%d")  # 获取当前的日期
                    # for i in range(0, len(data)):
                    #     data[i] = list(data[i])
                    #     if data[i][6] == None:
                    #         data[i].append("null")
                    #     else:
                    #         time_end = datetime.strptime(str(data[i][6]).split(' ')[0], "%Y-%m-%d")  # 获取数据表里面的日期数
                    #         days = time_end - start_time
                    #         data[i]['stand_date'].append(days.days)
                    #     if data[i][9] == 0:
                    #         data[i].append("null")
                    #     else:
                    #         data[i].append(data[i][9] - data[i][4])
                    #     # 判断是那种状态
                    #     if data[i][10] != "null" and data[i][11] != "null":
                    #         if data[i][10] > mt_date[0]['Max'] * data[i][5] and data[i][11] > count[0]['Max'] * \
                    #                 data[i][3]:
                    #             data[i].append('normal')
                    #         if 0 <= data[i][10] <= mt_date[0]['Max'] * data[i][5] and 0 <= data[i][11]:
                    #             data[i].append('warning')
                    #         if 0 <= data[i][11] <= mt_count[0]['Max'] * data[i][3] and 0 <= data[i][10]:
                    #             data[i].append('warning')
                    #         if data[i][10] < 0 or data[i][11] < 0:
                    #             data[i].append('danger')
                    #     if data[i][10] == "null" and data[i][11] != "null":
                    #         if data[i][11] > mt_count[0]['Max'] * data[i][3]:
                    #             data[i].append('normal')
                    #         if 0 <= data[i][11] <= mt_count[0]['Max'] * data[i][3]:
                    #             data[i].append('warning')
                    #         if data[i][11] < 0:
                    #             data[i].append('danger')
                    #     if data[i][10] != "null" and data[i][11] == "null":
                    #         if data[i][10] > mt_date[0]['Max'] * data[i][5]:
                    #             data[i].append('normal')
                    #         if 0 <= data[i][10] <= mt_date[0]['Max'] * data[i][5]:
                    #             data[i].append('warning')
                    #         if data[i][10] < 0:
                    #             data[i].append('danger')
                    #     if data[i][10] == "null" and data[i][11] == "null":
                    #         data[i].append('none')
                    dict_data['data'] = data
                    dict_data['page_count'] = count[0][0]
                if number != "All":
                    number = int(number)
                    count_page = count[0][0] // int(number)  # 总数除以一页显示多少条，得到总的页数
                    if count[0][0] % number > 0:
                        count_page += 1
                    if page <= count_page:
                        sql = sql + ' order by "Id" limit ' + str(number) + ' offset ' + str((page - 1)*number)
                        data = dictdata(sql)
                        data = tidy_dict(data)
                        # cur = connection.cursor()
                        # cur.execute(sql)
                        # data = cur.fetchall()
                        # data = tidy(data)
                        # 针对data 进行判断和算出次数水平和周期天数
                        # start_time = datetime.strptime(str(start).split(' ')[0], "%Y-%m-%d")  # 获取当前的日期
                        # for i in range(0, len(data)):
                        #     data[i] = list(data[i])
                        #     if data[i][6] == None:
                        #         data[i].append("null")
                        #     else:
                        #         time_end = datetime.strptime(str(data[i][6]).split(' ')[0], "%Y-%m-%d")  # 获取数据表里面的日期数
                        #         days = time_end - start_time
                        #         data[i].append(days.days)
                        #     if data[i][9] == 0:
                        #         data[i].append("null")
                        #     else:
                        #         data[i].append(data[i][9] - data[i][4])
                        #     # 判断是那种状态
                        #     if data[i][10] != "null" and data[i][11] != "null":
                        #         if data[i][10] > mt_date[0]['Max'] * data[i][5] and data[i][11] > mt_count[0]['Max'] * \
                        #                 data[i][3]:
                        #             data[i].append('normal')
                        #         if 0 <= data[i][10] <= mt_date[0]['Max'] * data[i][5] and 0 <= data[i][11]:
                        #             data[i].append('warning')
                        #         if 0 <= data[i][11] <= mt_count[0]['Max'] * data[i][3] and 0 <= data[i][10]:
                        #             data[i].append('warning')
                        #         if data[i][10] < 0 or data[i][11] < 0:
                        #             data[i].append('danger')
                        #     if data[i][10] == "null" and data[i][11] != "null":
                        #         if data[i][11] > mt_count[0]['Max'] * data[i][3]:
                        #             data[i].append('normal')
                        #         if 0 <= data[i][11] <= mt_count[0]['Max'] * data[i][3]:
                        #             data[i].append('warning')
                        #         if data[i][11] < 0:
                        #             data[i].append('danger')
                        #     if data[i][10] != "null" and data[i][11] == "null":
                        #         if data[i][10] > mt_date[0]['Max'] * data[i][5]:
                        #             data[i].append('normal')
                        #         if 0 <= data[i][10] <= mt_date[0]['Max'] * data[i][5]:
                        #             data[i].append('warning')
                        #         if data[i][10] < 0:
                        #             data[i].append('danger')
                        #     if data[i][10] == "null" and data[i][11] == "null":
                        #         data[i].append('none')
                        dict_data['data'] = data
                        dict_data['page_count'] = count_page
                    if page > count_page:
                        sql = sql + ' order by "Id" limit ' + str(number) + ' offset ' + str(0)
                        data = dictdata(sql)
                        data = tidy_dict(data)
                        # cur = connection.cursor()
                        # cur.execute(sql)
                        # data = cur.fetchall()
                        # data = tidy(data)
                        # 针对data 进行判断和算出次数水平和周期天数
                        # start_time = datetime.strptime(str(start).split(' ')[0], "%Y-%m-%d")  # 获取当前的日期
                        # for i in range(0, len(data)):
                        #     data[i] = list(data[i])
                        #     if data[i][6] == None:
                        #         data[i].append("null")
                        #     else:
                        #         time_end = datetime.strptime(str(data[i][6]).split(' ')[0], "%Y-%m-%d")  # 获取数据表里面的日期数
                        #         days = time_end - start_time
                        #         data[i].append(days.days)
                        #     # if data[i][9] == 0:
                        #     #     data[i].append(0)
                        #     # else:
                        #     data[i].append(data[i][9] - data[i][4])
                        dict_data['data'] = data
                        dict_data['page_count'] = count_page
                return restful.ok(data=dict_data)

                # dict_data['normal'] = normal
                # dict_data['warning'] = warning
                # dict_data['danger'] = danger
                # dict_data['None'] = info
        except Exception as e:
            return restful.params_error(message=repr(e))
# @access_control
# def maintain_query(request):
#     if request.method == "POST":
#         try:
#             start = datetime.now()
#             page = int(request.POST.get('page',''))
#             number = request.POST.get('num','')
#             sn = str(request.POST.get('sn',''))
#             part_name = request.POST.get('partname','')
#             status = request.POST.get('status','')
#             s_time = request.POST.get('s_time','')
#             e_time = request.POST.get('e_time','')
#             user = str(request.POST.get('user',''))
#             dict_data = {}
#             normal = []
#             warning = []
#             danger = []
#             info = []
#             #计算 查询数据的总的多少条数据
#             sql_count = 'select count(*) FROM "PartItem" where 1=1 '
#             #查询条件的数据
#             sql = 'select "Id","SN","PartName","CheckCycleCount","UsedTimes","CheckCycle","NextCheckDate","Maintainer","TrnDate","NextCheckCount" FROM "PartItem" WHERE 1=1 '
#             # 饼图数据按正常和预警和超标分类计算数量
#             visual_sql = 'SELECT "PartName", COUNT("PartName") FROM "PartItem" where 1=1 '
#             # 柱状图的数据 按条件查询出来之后再 按负责人分类
#             tab_sql = 'SELECT "Maintainer", COUNT("PartName") FROM "PartItem" where 1=1 '
#
#             count_count = Configuration.objects.get(Type="mt_count")
#             date = Configuration.objects.get(Type="mt_date")
#             dict_data['limit_value1'] = int(count_count.Max)
#             dict_data['limit_value2'] = date.Max
#             c_count = str(int(count_count.Max))
#             c_date = (start + timedelta(days=int(date.Max))).strftime("%Y-%m-%d")  # 拿到当前的时间+预警天数得到日期的范围是（当前时间,c_date）
#             # 预设区间的筛选，回滚点击视图的函数的加载
#             time_stand = (list(PartItem.objects.order_by("-TrnDate").filter(TrnDate__lte=start).values("TrnDate")))[0]['TrnDate']
#             start_select = time_stand
#             delta = timedelta(days=7)
#             end = start_select - delta
#             callback_sql =' AND to_char("TrnDate",\'yyyy-MM-dd\') >= \'' + end.strftime("%Y-%m-%d") + '\' and to_char("TrnDate",\'yyyy-MM-dd\') <= \'' + start_select.strftime("%Y-%m-%d") + '\''
#             if sn == "" and part_name == "" and s_time == "" and e_time == "" and user == "" and status =="":
#                 sql =sql + callback_sql
#                 sql_count =sql_count + callback_sql
#                 visual_sql =visual_sql + callback_sql
#                 tab_sql =tab_sql + callback_sql
#             if status != "":
#                 if sn != "":
#                     sql_count = sql_count + 'AND "SN" = \'' + sn + '\''
#                     sql = sql + 'AND "SN" = \'' + sn + '\''
#                     visual_sql = visual_sql + 'AND "SN" = \'' + sn + '\''
#                     tab_sql = tab_sql + 'AND "SN" = \'' + sn + '\''
#                 if part_name != "":
#                     sql_count = sql_count + 'AND "PartName" ilike \'%{0}%\''.format(part_name)
#                     sql = sql + 'AND "PartName" ilike \'%{0}%\''.format(part_name)
#                     visual_sql = visual_sql + 'AND "PartName" ilike \'%{0}%\''.format(part_name)
#                     tab_sql = tab_sql + 'AND "PartName" ilike \'%{0}%\''.format(part_name)
#                 if s_time != "":
#                     sql_count = sql_count + 'AND "TrnDate" >= \'{0}\''.format(s_time)
#                     sql = sql + 'AND "TrnDate" >= \'{0}\''.format(s_time)
#                     visual_sql = visual_sql + 'AND "TrnDate" >= \'{0}\''.format(s_time)
#                     tab_sql = tab_sql + 'AND "TrnDate" >= \'{0}\''.format(s_time)
#                 if e_time != "":
#                     sql_count = sql_count + 'AND "TrnDate" <= \'%{0}%\''.format(e_time)
#                     sql = sql + 'AND "TrnDate" <= \'%{0}%\''.format(e_time)
#                     visual_sql = visual_sql + 'AND "TrnDate" <= \'%{0}%\''.format(e_time)
#                     tab_sql = tab_sql + 'AND "TrnDate" <= \'%{0}%\''.format(e_time)
#                 if user != "":
#                     sql_count = sql_count + 'AND "Maintainer" = \'{0}\''.format(user)
#                     sql = sql + 'AND "Maintainer" = \'{0}\''.format(user)
#                     visual_sql = visual_sql + 'AND "Maintainer" = \'{0}\''.format(user)
#                     tab_sql = tab_sql + 'AND "Maintainer" = \'{0}\''.format(user)
#                 # 正常的条件设置
#                 sql_n_t = 'OR ( to_char("NextCheckDate",\'yyyy-MM-dd\') !=\'None\' AND "NextCheckCount" =0 AND to_char("NextCheckDate",\'yyyy-MM-dd\') >\''+c_date+'\')'
#                 sql_n_c = 'OR ("NextCheckDate" IS NULL AND "NextCheckCount" !=0 AND "NextCheckCount"-"UsedTimes" > ' + c_count + ')'
#                 # 预警的条件设置
#                 sql_w_t = 'OR (to_char("NextCheckDate",\'yyyy-MM-dd\') !=\'None\' AND "NextCheckCount" =0 AND to_char("NextCheckDate",\'yyyy-MM-dd\') >=\'' + start.strftime(
#                     "%Y-%m-%d") + '\' AND to_char("NextCheckDate",\'yyyy-MM-dd\') <= \'' + c_date + '\')'
#                 sql_w_c = 'OR ( "NextCheckDate" IS NULL AND "NextCheckCount" !=0 AND "NextCheckCount"-"UsedTimes" <= ' + c_count + ' AND "NextCheckCount"-"UsedTimes" >= 0)'
#                 # 超标的条件设置
#                 sql_t = 'OR ( to_char("NextCheckDate",\'yyyy-MM-dd\') !=\'None\' AND "NextCheckCount" =0 AND to_char("NextCheckDate",\'yyyy-MM-dd\') <\'' + start.strftime("%Y-%m-%d") + '\')'
#                 sql_c = 'OR ("NextCheckDate" IS NULL AND "NextCheckCount" !=0 AND "NextCheckCount"-"UsedTimes" < 0 ) '
#
#                 if status == "正常":
#                     sql_count = sql_count+'AND ( (to_char("NextCheckDate",\'yyyy-MM-dd\') !=\'None\' AND "NextCheckCount" !=0 AND "NextCheckCount"-"UsedTimes" > '+c_count+' AND to_char("NextCheckDate",\'yyyy-MM-dd\') >\''+c_date+'\') '
#                     sql_count =sql_count+sql_n_t+sql_n_c+');'
#
#                     sql = sql+'AND ( (to_char("NextCheckDate",\'yyyy-MM-dd\') !=\'None\' AND "NextCheckCount" !=0 AND "NextCheckCount"-"UsedTimes" > '+c_count+' AND to_char("NextCheckDate",\'yyyy-MM-dd\') >\''+c_date+'\' )'
#                     sql = sql+sql_n_t+sql_n_c+')'
#
#                     visual_sql_normal = visual_sql+'AND ( (to_char("NextCheckDate",\'yyyy-MM-dd\') !=\'None\' AND "NextCheckCount" !=0 AND "NextCheckCount"-"UsedTimes" > '+c_count+' AND to_char("NextCheckDate",\'yyyy-MM-dd\') >\''+c_date+'\' )'
#                     visual_sql_normal =visual_sql_normal+sql_n_t+sql_n_c+' ) GROUP BY "PartName";'
#
#                     tab_normal = tab_sql+'AND ( (to_char("NextCheckDate",\'yyyy-MM-dd\') !=\'None\' AND "NextCheckCount" !=0 AND "NextCheckCount"-"UsedTimes" > '+c_count+' AND to_char("NextCheckDate",\'yyyy-MM-dd\') >\''+c_date+'\' )'
#                     tab_normal = tab_normal+sql_n_t+sql_n_c+' ) GROUP BY "Maintainer";'
#
#                     cur = connection.cursor()
#                     cur.execute(visual_sql_normal)
#                     normal = cur.fetchall()
#
#                     cur = connection.cursor()
#                     cur.execute(tab_normal)
#                     tab_normal =cur.fetchall()
#                     tab_normal = tab_query_way(tab_normal,[],[],[])
#                     dict_data['tab_data'] = tab_normal
#                 if status == "预警":
#                     sql_count = sql_count+'and ( to_char("NextCheckDate",\'yyyy-MM-dd\') !=\'None\' AND "NextCheckCount" !=0 AND (("NextCheckCount"-"UsedTimes" <= ' + c_count + ' AND "NextCheckCount"-"UsedTimes" >= 0 AND to_char("NextCheckDate",\'yyyy-MM-dd\') >=\'' + start.strftime("%Y-%m-%d") + '\')'
#                     sql_count = sql_count + ' OR  (to_char("NextCheckDate",\'yyyy-MM-dd\')>=\'' + start.strftime("%Y-%m-%d") + '\' AND to_char("NextCheckDate",\'yyyy-MM-dd\') <= \'' + c_date + '\' AND "NextCheckCount"-"UsedTimes">=0))'
#                     sql_count = sql_count+sql_w_t+sql_w_c+')'
#
#                     sql = sql+'and ( to_char("NextCheckDate",\'yyyy-MM-dd\') !=\'None\' AND "NextCheckCount" !=0 AND (("NextCheckCount"-"UsedTimes" <= ' + c_count + ' AND "NextCheckCount"-"UsedTimes" >= 0 AND to_char("NextCheckDate",\'yyyy-MM-dd\') >=\'' + start.strftime("%Y-%m-%d") + '\')'
#                     sql = sql + ' OR  (to_char("NextCheckDate",\'yyyy-MM-dd\')>=\'' + start.strftime("%Y-%m-%d") + '\' AND to_char("NextCheckDate",\'yyyy-MM-dd\') <= \'' + c_date + '\' AND "NextCheckCount"-"UsedTimes">=0))'
#                     sql = sql+sql_w_t+sql_w_c+')'
#
#                     visual_sql = visual_sql+'and ( to_char("NextCheckDate",\'yyyy-MM-dd\') !=\'None\' AND "NextCheckCount" !=0 AND (("NextCheckCount"-"UsedTimes" <= ' + c_count + ' AND "NextCheckCount"-"UsedTimes" >= 0 AND to_char("NextCheckDate",\'yyyy-MM-dd\') >=\'' + start.strftime("%Y-%m-%d") + '\')'
#                     visual_sql_waring = visual_sql + ' OR  (to_char("NextCheckDate",\'yyyy-MM-dd\')>=\'' + start.strftime("%Y-%m-%d") + '\' AND to_char("NextCheckDate",\'yyyy-MM-dd\') <= \'' + c_date + '\' AND "NextCheckCount"-"UsedTimes">=0)) '
#                     visual_sql_waring = visual_sql_waring+sql_w_t+sql_w_c+') GROUP BY "PartName"'
#
#                     tab_warning = tab_sql+'and ( to_char("NextCheckDate",\'yyyy-MM-dd\') !=\'None\' AND "NextCheckCount" !=0 AND (("NextCheckCount"-"UsedTimes" <= ' + c_count + ' AND "NextCheckCount"-"UsedTimes" >= 0 AND to_char("NextCheckDate",\'yyyy-MM-dd\') >=\'' + start.strftime("%Y-%m-%d") + '\')'
#                     tab_warning = tab_warning + ' OR  (to_char("NextCheckDate",\'yyyy-MM-dd\')>=\'' + start.strftime("%Y-%m-%d") + '\' AND to_char("NextCheckDate",\'yyyy-MM-dd\') <= \'' + c_date + '\' AND "NextCheckCount"-"UsedTimes">=0)) '
#                     tab_warning = tab_warning+sql_w_t+sql_w_c+' ) GROUP BY "Maintainer"'
#
#                     cur = connection.cursor()
#                     cur.execute(visual_sql_waring)
#                     warning = cur.fetchall()
#
#                     cur.execute(tab_warning)
#                     tab_warning = cur.fetchall()
#                     tab_warning = tab_query_way([],tab_warning,[],[])
#                     dict_data['tab_data'] = tab_warning
#                 if status == "超标":
#                     sql = sql+'and ( (to_char("NextCheckDate",\'yyyy-MM-dd\') !=\'None\' AND "NextCheckCount" !=0 AND ("NextCheckCount"-"UsedTimes" < 0 OR "NextCheckDate" < \''+start.strftime("%Y-%m-%d")+'\'))'
#                     sql = sql+sql_t+sql_c+')'
#
#                     sql_count = sql_count+'and ( (to_char("NextCheckDate",\'yyyy-MM-dd\') !=\'None\' AND "NextCheckCount" !=0 AND ("NextCheckCount"-"UsedTimes" < 0 OR "NextCheckDate" < \''+start.strftime("%Y-%m-%d")+'\'))'
#                     sql_count = sql_count+sql_t+sql_c+')'
#
#                     visual_sql_danger = visual_sql+'and ( (to_char("NextCheckDate",\'yyyy-MM-dd\') !=\'None\' AND "NextCheckCount" !=0 AND ("NextCheckCount"-"UsedTimes" < 0 OR "NextCheckDate" < \''+start.strftime("%Y-%m-%d")+'\')) '
#                     visual_sql_danger = visual_sql_danger+sql_t+sql_c+' ) GROUP BY "PartName"'
#
#                     tab_danger = tab_sql+'and ( (to_char("NextCheckDate",\'yyyy-MM-dd\') !=\'None\' AND "NextCheckCount" !=0 AND ("NextCheckCount"-"UsedTimes" < 0 OR "NextCheckDate" < \''+start.strftime("%Y-%m-%d")+'\'))'
#                     tab_danger =tab_danger+sql_t+sql_c+' ) GROUP BY "Maintainer"'
#                     cur = connection.cursor()
#                     cur.execute(visual_sql_danger)
#                     danger = cur.fetchall()
#                     cur.execute(tab_danger)
#                     tab_danger = cur.fetchall()
#                     tab_danger = tab_query_way([], [], tab_danger,[])
#                     dict_data['tab_data'] = tab_danger
#                 if status == "未设定":
#                     sql = sql +'AND "NextCheckDate" IS NULL AND "NextCheckCount" =0'
#
#                     visual_sql_no_checkdate = sql_count + 'AND "NextCheckDate" IS NULL AND "NextCheckCount" =0;'
#                     sql_count = sql_count + 'AND "NextCheckDate" IS NULL AND "NextCheckCount" =0;'
#
#                     cur = connection.cursor()
#                     cur.execute(visual_sql_no_checkdate)
#                     visual_sql_no_checkdate = cur.fetchall()
#                     info = [('None',visual_sql_no_checkdate[0][0])]
#                     info = tab_query_way([], [], [],info)
#                     dict_data['tab_data'] = info
#                 dict_data['normal'] = normal
#                 dict_data['warning'] = warning
#                 dict_data['danger'] = danger
#                 dict_data['None'] = info
#                 cur = connection.cursor()
#                 cur.execute(sql_count)
#                 count = cur.fetchall()  # 数量的总数
#                 if number == "All":
#                     cur = connection.cursor()
#                     cur.execute(sql)
#                     data = cur.fetchall()
#                     # 针对data 进行判断和算出次数水平和周期天数
#                     start_time = datetime.strptime(str(start).split(' ')[0], "%Y-%m-%d")  # 获取当前的日期
#                     for i in range(0, len(data)):
#                         data[i] = list(data[i])
#                         if data[i][6] == None:
#                             data[i].append("null")
#                         else:
#                             time_end = datetime.strptime(str(data[i][6]).split(' ')[0],"%Y-%m-%d")  # 获取数据表里面的日期数
#                             days = time_end - start_time
#                             data[i].append(days.days)
#                         if data[i][9] == 0:
#                             data[i].append("null")
#                         else:
#                             data[i].append( data[i][9] - data[i][4])
#                     dict_data['data'] = data
#                     dict_data['page_count'] = count[0][0]
#                 if number != "All":
#                     number = int(number)
#                     count_page = count[0][0] // int(number)  # 总数除以一页显示多少条，得到总的页数
#                     if count[0][0] % number > 0:
#                         count_page += 1
#                     if page <= count_page:
#                         sql = sql + ' order by "Id" limit ' + str(number) + ' offset ' + str((page - 1)*number)
#                         cur = connection.cursor()
#                         cur.execute(sql)
#                         data = cur.fetchall()
#                         # 针对data 进行判断和算出次数水平和周期天数
#                         start_time = datetime.strptime(str(start).split(' ')[0], "%Y-%m-%d")  # 获取当前的日期
#                         for i in range(0, len(data)):
#                             data[i] = list(data[i])
#                             if data[i][6] == None:
#                                 data[i].append("null")
#                             else:
#                                 time_end = datetime.strptime(str(data[i][6]).split(' ')[0], "%Y-%m-%d")  # 获取数据表里面的日期数
#                                 days = time_end - start_time
#                                 data[i].append(days.days)
#                             if data[i][9] == 0:
#                                 data[i].append("null")
#                             else:
#                                 data[i].append(data[i][9] - data[i][4])
#                         dict_data['data'] = data
#                         dict_data['page_count'] = count_page
#                     if page > count_page:
#                         sql = sql + ' order by "Id" limit ' + str(number) + ' offset ' + str(0)
#                         cur = connection.cursor()
#                         cur.execute(sql)
#                         data = cur.fetchall()
#                         # 针对data 进行判断和算出次数水平和周期天数
#                         start_time = datetime.strptime(str(start).split(' ')[0], "%Y-%m-%d")  # 获取当前的日期
#                         for i in range(0, len(data)):
#                             data[i] = list(data[i])
#                             if data[i][6] == None:
#                                 data[i].append("null")
#                             else:
#                                 time_end = datetime.strptime(str(data[i][6]).split(' ')[0], "%Y-%m-%d")  # 获取数据表里面的日期数
#                                 days = time_end - start_time
#                                 data[i].append(days.days)
#                             if data[i][9] == 0:
#                                 data[i].append("null")
#                             else:
#                                 data[i].append(data[i][9] - data[i][4])
#                         dict_data['data'] = data
#                         dict_data['page_count'] = count_page
#
#                 dict_data['normal'] = normal
#                 dict_data['warning'] = warning
#                 dict_data['danger'] = danger
#                 dict_data['None'] = info
#                 return restful.ok(data=dict_data)
#             else:
#                 if sn != "":
#                     sql_count = sql_count + 'AND "SN" = \'' + sn + '\''
#                     sql = sql + 'AND "SN" = \'' + sn + '\''
#                     visual_sql = visual_sql + 'AND "SN" = \'' + sn + '\''
#                     tab_sql = tab_sql + 'AND "SN" = \'' + sn + '\''
#                 if part_name != "":
#                     sql_count = sql_count + 'AND "PartName" ilike \'%{0}%\''.format(part_name)
#                     sql = sql + 'AND "PartName" ilike \'%{0}%\''.format(part_name)
#                     visual_sql = visual_sql + 'AND "PartName" ilike \'%{0}%\''.format(part_name)
#                     tab_sql = tab_sql + 'AND "PartName" ilike \'%{0}%\''.format(part_name)
#                 if s_time != "":
#                     sql_count = sql_count + 'AND "TrnDate" >= \'{0}\''.format(s_time)
#                     sql = sql + 'AND "TrnDate" >= \'{0}\''.format(s_time)
#                     visual_sql = visual_sql + 'AND "TrnDate" >= \'{0}\''.format(s_time)
#                     tab_sql = tab_sql + 'AND "TrnDate" >= \'{0}\''.format(s_time)
#                 if e_time != "":
#                     sql_count = sql_count + 'AND "TrnDate" <= \'{0}\''.format(e_time)
#                     sql = sql + 'AND "TrnDate" <= \'{0}\''.format(e_time)
#                     visual_sql = visual_sql + 'AND "TrnDate" <= \'{0}\''.format(e_time)
#                     tab_sql = tab_sql + 'AND "TrnDate" <= \'{0}\''.format(e_time)
#                 if user != "":
#                     sql_count = sql_count + 'AND "Maintainer" = \'{0}\''.format(user)
#                     sql = sql + 'AND "Maintainer" = \'{0}\''.format(user)
#                     visual_sql = visual_sql + 'AND "Maintainer" = \'{0}\''.format(user)
#                     tab_sql = tab_sql + 'AND "Maintainer" = \'{0}\''.format(user)
#                 # 正常的条件设置
#                 sql_n_t = 'OR ( to_char("NextCheckDate",\'yyyy-MM-dd\') !=\'None\' AND "NextCheckCount" =0 AND to_char("NextCheckDate",\'yyyy-MM-dd\') >\'' + c_date + '\')'
#                 sql_n_c = 'OR ("NextCheckDate" IS NULL AND "NextCheckCount" !=0 AND "NextCheckCount"-"UsedTimes" > ' + c_count + ')'
#                 # 预警的条件设置
#                 sql_w_t = 'OR (to_char("NextCheckDate",\'yyyy-MM-dd\') !=\'None\' AND "NextCheckCount" =0 AND to_char("NextCheckDate",\'yyyy-MM-dd\') >=\'' + start.strftime(
#                     "%Y-%m-%d") + '\' AND to_char("NextCheckDate",\'yyyy-MM-dd\') <= \'' + c_date + '\')'
#                 sql_w_c = 'OR ( "NextCheckDate" IS NULL AND "NextCheckCount" !=0 AND "NextCheckCount"-"UsedTimes" <= ' + c_count + ' AND "NextCheckCount"-"UsedTimes" >= 0)'
#                 # 超标的条件设置
#                 sql_t = 'OR ( to_char("NextCheckDate",\'yyyy-MM-dd\') !=\'None\' AND "NextCheckCount" =0 AND to_char("NextCheckDate",\'yyyy-MM-dd\') <\'' + start.strftime(
#                     "%Y-%m-%d") + '\')'
#                 sql_c = 'OR ("NextCheckDate" IS NULL AND "NextCheckCount" !=0 AND "NextCheckCount"-"UsedTimes" < 0 ) '
#                 #normal
#                 visual_sql_normal = visual_sql+'and ( to_char("NextCheckDate",\'yyyy-MM-dd\') !=\'None\' AND "NextCheckCount" !=0 AND "NextCheckCount"-"UsedTimes" > ' + c_count + ' AND to_char("NextCheckDate",\'yyyy-MM-dd\') >\'' + c_date + '\''
#                 visual_sql_normal = visual_sql_normal+sql_n_t+sql_n_c+')  GROUP BY "PartName"'
#                 tab_normal = tab_sql+'and ( to_char("NextCheckDate",\'yyyy-MM-dd\') !=\'None\' AND "NextCheckCount" !=0 AND "NextCheckCount"-"UsedTimes" > ' + c_count + ' AND to_char("NextCheckDate",\'yyyy-MM-dd\') >\'' + c_date + '\' '
#                 tab_normal = tab_normal+sql_n_t+sql_n_c+') GROUP BY "Maintainer"'
#                 #warning
#                 visual_sql_waring = visual_sql+'and ( to_char("NextCheckDate",\'yyyy-MM-dd\') !=\'None\' AND "NextCheckCount" !=0 AND (("NextCheckCount"-"UsedTimes" <= ' + c_count + ' AND "NextCheckCount"-"UsedTimes" >= 0 AND to_char("NextCheckDate",\'yyyy-MM-dd\') >=\'' + start.strftime("%Y-%m-%d") + '\')'
#                 visual_sql_waring = visual_sql_waring + ' OR  (to_char("NextCheckDate",\'yyyy-MM-dd\')>=\'' + start.strftime("%Y-%m-%d") + '\' AND to_char("NextCheckDate",\'yyyy-MM-dd\') <= \'' + c_date + '\' AND "NextCheckCount"-"UsedTimes">=0)) '
#                 visual_sql_waring = visual_sql_waring+sql_w_t+sql_w_c+')  GROUP BY "PartName"'
#                 tab_warning = tab_sql+'and ( to_char("NextCheckDate",\'yyyy-MM-dd\') !=\'None\' AND "NextCheckCount" !=0 AND (("NextCheckCount"-"UsedTimes" <= ' + c_count + ' AND "NextCheckCount"-"UsedTimes" >= 0 AND to_char("NextCheckDate",\'yyyy-MM-dd\') >=\'' + start.strftime("%Y-%m-%d") + '\')'
#                 tab_warning = tab_warning + ' OR  (to_char("NextCheckDate",\'yyyy-MM-dd\')>=\'' + start.strftime("%Y-%m-%d") + '\' AND to_char("NextCheckDate",\'yyyy-MM-dd\') <= \'' + c_date + '\' AND "NextCheckCount"-"UsedTimes">=0)) '
#                 tab_warning = tab_warning+sql_w_t+sql_w_c+')  GROUP BY "Maintainer"'
#
#                 #danger
#                 visual_sql_danger = visual_sql+'and (( to_char("NextCheckDate",\'yyyy-MM-dd\') !=\'None\'AND "NextCheckCount" !=0 AND ("NextCheckCount"-"UsedTimes" < 0 OR "NextCheckDate" < \'' + start.strftime("%Y-%m-%d") + '\'))'
#                 visual_sql_danger = visual_sql_danger+sql_t+sql_c+')  GROUP BY "PartName"'
#                 tab_danger = tab_sql+'and (( to_char("NextCheckDate",\'yyyy-MM-dd\') !=\'None\'AND "NextCheckCount" !=0 AND ("NextCheckCount"-"UsedTimes" < 0 OR "NextCheckDate" < \'' + start.strftime("%Y-%m-%d") + '\') )'
#                 tab_danger =tab_danger+sql_t+sql_c+') GROUP BY "Maintainer"'
#
#                 #None
#                 visual_sql_no_checkdate = sql_count + 'AND "NextCheckDate" IS NULL AND "NextCheckCount" =0;'
#                 cur = connection.cursor()
#                 cur.execute(visual_sql_normal)
#                 normal = cur.fetchall()
#                 cur.execute(visual_sql_waring)
#                 warning = cur.fetchall()
#                 cur.execute(visual_sql_danger)
#                 danger = cur.fetchall()
#                 cur.execute(visual_sql_no_checkdate)
#                 info = cur.fetchall()
#                 info =[('None',info[0][0])]
#
#                 cur.execute(tab_normal)
#                 tab_normal = cur.fetchall()
#                 cur.execute(tab_warning)
#                 tab_warning = cur.fetchall()
#                 cur.execute(tab_danger)
#                 tab_danger = cur.fetchall()
#                 tab_info = info
#                 tab_data = tab_query_way(tab_normal,tab_warning, tab_danger,tab_info )
#                 dict_data['tab_data'] = tab_data
#
#                 dict_data['normal'] = normal
#                 dict_data['warning'] = warning
#                 dict_data['danger'] = danger
#                 dict_data['None'] = info
#                 cur = connection.cursor()
#                 cur.execute(sql_count)
#                 count = cur.fetchall()  # 数量的总数
#                 if number == "All":
#                     cur = connection.cursor()
#                     cur.execute(sql)
#                     data = cur.fetchall()
#                     # 针对data 进行判断和算出次数水平和周期天数
#                     start_time = datetime.strptime(str(start).split(' ')[0], "%Y-%m-%d")  # 获取当前的日期
#                     for i in range(0, len(data)):
#                         data[i] = list(data[i])
#                         if data[i][6] == None:
#                             data[i].append("null")
#                         else:
#                             time_end = datetime.strptime(str(data[i][6]).split(' ')[0], "%Y-%m-%d")  # 获取数据表里面的日期数
#                             days = time_end - start_time
#                             data[i]['stand_date'].append(days.days)
#                         if data[i][9] == 0:
#                             data[i].append("null")
#                         else:
#                             data[i].append(data[i][9] - data[i][4])
#                     dict_data['data'] = data
#                     dict_data['page_count'] = count[0][0]
#                 if number != "All":
#                     number = int(number)
#                     count_page = count[0][0] // int(number)  # 总数除以一页显示多少条，得到总的页数
#                     if count[0][0] % number > 0:
#                         count_page += 1
#                     if page <= count_page:
#                         sql = sql + ' order by "Id" limit ' + str(number) + ' offset ' + str((page - 1)*number)
#                         cur = connection.cursor()
#                         cur.execute(sql)
#                         data = cur.fetchall()
#                         # 针对data 进行判断和算出次数水平和周期天数
#                         start_time = datetime.strptime(str(start).split(' ')[0], "%Y-%m-%d")  # 获取当前的日期
#                         for i in range(0, len(data)):
#                             data[i] = list(data[i])
#                             if data[i][6] == None:
#                                 data[i].append("null")
#                             else:
#                                 time_end = datetime.strptime(str(data[i][6]).split(' ')[0], "%Y-%m-%d")  # 获取数据表里面的日期数
#                                 days = time_end - start_time
#                                 data[i].append(days.days)
#                             if data[i][9] == 0:
#                                 data[i].append("null")
#                             else:
#                                 data[i].append(data[i][9] - data[i][4])
#                         dict_data['data'] = data
#                         dict_data['page_count'] = count_page
#                     if page > count_page:
#                         sql = sql + ' order by "Id" limit ' + str(number) + ' offset ' + str(0)
#                         cur = connection.cursor()
#                         cur.execute(sql)
#                         data = cur.fetchall()
#                         # 针对data 进行判断和算出次数水平和周期天数
#                         start_time = datetime.strptime(str(start).split(' ')[0], "%Y-%m-%d")  # 获取当前的日期
#                         for i in range(0, len(data)):
#                             data[i] = list(data[i])
#                             if data[i][6] == None:
#                                 data[i].append("null")
#                             else:
#                                 time_end = datetime.strptime(str(data[i][6]).split(' ')[0], "%Y-%m-%d")  # 获取数据表里面的日期数
#                                 days = time_end - start_time
#                                 data[i].append(days.days)
#                             # if data[i][9] == 0:
#                             #     data[i].append(0)
#                             # else:
#                             data[i].append(data[i][9] - data[i][4])
#                         dict_data['data'] = data
#                         dict_data['page_count'] = count_page
#
#                 dict_data['normal'] = normal
#                 dict_data['warning'] = warning
#                 dict_data['danger'] = danger
#                 dict_data['None'] = info
#                 return restful.ok(data=dict_data)
#         except Exception as e:
#             return restful.params_error(message=repr(e))



#对视图的点击事件的修改功能的实现
@access_control
def maintain_monitor_visual(request):
    if request.method == "POST":
        try:
            start = datetime.now()
            page = int(request.POST.get('page',''))
            number = request.POST.get('num','')
            sn = str(request.POST.get('sn',''))
            part_name = request.POST.get('partname','')
            status = request.POST.get('status','')
            s_time = request.POST.get('s_time','')
            e_time = request.POST.get('e_time','')
            user = str(request.POST.get('user',''))
            locationId = str(request.POST.get('location', ''))
            dict_data = {}
            normal = []
            warning = []
            danger = []
            info = []
            #计算 查询数据的总的多少条数据
            sql_count = 'select count(*) FROM "PartItem" where "UseStatus"=\'normal\' '
            #查询条件的数据
            sql = 'select "Id","SN","PartName","CheckCycleCount","UsedTimes","CheckCycle","NextCheckDate","Maintainer","TrnDate","NextCheckCount","WarningBeforeDays","WarningBeforeTimes","LocationId" FROM "PartItem" WHERE "UseStatus"=\'normal\' '
            # 饼图数据按正常和预警和超标分类计算数量
            visual_sql = 'SELECT "PartName", COUNT("PartName") FROM "PartItem" where "UseStatus"=\'normal\' '
            # 柱状图的数据 按条件查询出来之后再 按负责人分类
            tab_sql = 'SELECT "Maintainer", COUNT("PartName") FROM "PartItem" where  "UseStatus"=\'normal\' '

            mt_count = list(Configuration.objects.filter(Type="mt_count").values("Max"))
            mt_date = list(Configuration.objects.filter(Type="mt_date").values("Max"))
            # c_date = (start + timedelta(days=int(date.Max))).strftime("%Y-%m-%d")  # 拿到当前的时间+预警天数得到日期的范围是（当前时间,c_date）
            # 预设区间的筛选，回滚点击视图的函数的加载
            start_select = (list(PartItem.objects.order_by("-TrnDate").filter(TrnDate__lte=start,UseStatus="normal").values("TrnDate")))[0]['TrnDate']
            end = start_select - timedelta(days=7)
            callback_sql =' AND to_char("TrnDate",\'yyyy-MM-dd  HH24:MI:SS\') >= \'' + end.strftime("%Y-%m-%d %H:%M:%S") +'\''
            if sn == "" and part_name == "" and s_time == "" and e_time == "" and user == "" :
                sql =sql + callback_sql
                sql_count =sql_count + callback_sql
                visual_sql =visual_sql + callback_sql
                tab_sql =tab_sql + callback_sql
            if sn != "":
                sql_count = sql_count + 'AND "SN" = \'' + sn + '\''
                sql = sql + 'AND "SN" = \'' + sn + '\''
                visual_sql = visual_sql + 'AND "SN" = \'' + sn + '\''
                tab_sql = tab_sql + 'AND "SN" = \'' + sn + '\''
            if part_name != "":
                sql_count = sql_count + 'AND "PartName" ilike \'%{0}%\''.format(part_name)
                sql = sql + 'AND "PartName" ilike \'%{0}%\''.format(part_name)
                visual_sql = visual_sql + 'AND "PartName" ilike \'%{0}%\''.format(part_name)
                tab_sql = tab_sql + 'AND "PartName" ilike \'%{0}%\''.format(part_name)
            if s_time != "":
                sql_count = sql_count + 'AND "TrnDate" >= \'{0}\''.format(s_time)
                sql = sql + 'AND "TrnDate" >= \'{0}\''.format(s_time)
                visual_sql = visual_sql + 'AND "TrnDate" >= \'{0}\''.format(s_time)
                tab_sql = tab_sql + 'AND "TrnDate" >= \'{0}\''.format(s_time)
            if e_time != "":
                sql_count = sql_count + 'AND "TrnDate" <= \'%{0}%\''.format(e_time)
                sql = sql + 'AND "TrnDate" <= \'%{0}%\''.format(e_time)
                visual_sql = visual_sql + 'AND "TrnDate" <= \'%{0}%\''.format(e_time)
                tab_sql = tab_sql + 'AND "TrnDate" <= \'%{0}%\''.format(e_time)
            if user != "":
                sql_count = sql_count + 'AND ("Maintainer" = \'{0}\''.format(user)+ ' OR "SubMaintainers" ilike \'%'+ str(user) +'%\')'
                sql = sql + 'AND ("Maintainer" = \'{0}\''.format(user)+ ' OR "SubMaintainers" ilike \'%'+ str(user) +'%\')'
                visual_sql = visual_sql + 'AND ("Maintainer" = \'{0}\''.format(user)+ ' OR "SubMaintainers" ilike \'%'+ str(user) +'%\')'
                tab_sql = tab_sql + 'AND ("Maintainer" = \'{0}\''.format(user)+ ' OR "SubMaintainers" ilike \'%'+ str(user) +'%\')'
            if locationId != "":
                sql_count = sql_count + 'AND "LocationId" = \'{0}\''.format(locationId)
                sql = sql + 'AND "LocationId" = \'{0}\''.format(locationId)
                visual_sql = visual_sql + 'AND "LocationId" = \'{0}\''.format(locationId)
                tab_sql = tab_sql + 'AND "LocationId" = \'{0}\''.format(locationId)
            if status == "#28a745":
                sql_count = sql_select(sql_count,"正常",mt_count[0]['Max'],mt_date[0]['Max'])
                sql       = sql_select(sql, "正常", mt_count[0]['Max'], mt_date[0]['Max'])
                visual_sql_normal= sql_select(visual_sql,"正常",mt_count[0]['Max'],mt_date[0]['Max'])+' GROUP BY "PartName";'
                tab_normal = sql_select(tab_sql,"正常",mt_count[0]['Max'],mt_date[0]['Max'])+' GROUP BY "Maintainer";'
                cur = connection.cursor()
                cur.execute(visual_sql_normal)
                normal = cur.fetchall()
                cur = connection.cursor()
                cur.execute(tab_normal)
                tab_normal =cur.fetchall()
                tab_normal = tab_query_way(tab_normal,[],[],[])
                dict_data['tab_data'] = tab_normal
            if status == "#ffc107":
                sql_count = sql_select(sql_count, "预警", mt_count[0]['Max'], mt_date[0]['Max'])
                sql = sql_select(sql, "预警", mt_count[0]['Max'], mt_date[0]['Max'])
                visual_sql_waring = sql_select(visual_sql, "预警", mt_count[0]['Max'],mt_date[0]['Max']) + ' GROUP BY "PartName";'
                tab_warning = sql_select(tab_sql, "预警", mt_count[0]['Max'],mt_date[0]['Max']) + '  GROUP BY "Maintainer";'
                cur = connection.cursor()
                cur.execute(visual_sql_waring)
                warning = cur.fetchall()
                cur.execute(tab_warning)
                tab_warning = cur.fetchall()
                tab_warning = tab_query_way([],tab_warning,[],[])
                dict_data['tab_data'] = tab_warning
            if status == "#dc3545":
                sql_count = sql_select(sql_count, "超标", mt_count[0]['Max'], mt_date[0]['Max'])
                sql = sql_select(sql, "超标", mt_count[0]['Max'], mt_date[0]['Max'])
                visual_sql_danger = sql_select(visual_sql, "超标", mt_count[0]['Max'],mt_date[0]['Max']) + ' GROUP BY "PartName";'
                tab_danger = sql_select(tab_sql, "超标", mt_count[0]['Max'],mt_date[0]['Max']) + '  GROUP BY "Maintainer";'
                cur = connection.cursor()
                cur.execute(visual_sql_danger)
                danger = cur.fetchall()
                cur.execute(tab_danger)
                tab_danger = cur.fetchall()
                tab_danger = tab_query_way([], [], tab_danger,[])
                dict_data['tab_data'] = tab_danger
            if status == "#17a2b8":
                sql = sql_select(sql, "未设定", mt_count[0]['Max'], mt_date[0]['Max'])
                sql_count = sql_select(sql_count, "未设定", mt_count[0]['Max'], mt_date[0]['Max'])
                # visual_sql_no_checkdate ='SELECT COUNT("PartName") FROM "PartItem" where to_char("TrnDate",\'yyyy-MM-dd HH24:MI:SS\') >=\' '+ end.strftime("%Y-%m-%d %H:%M:%S")+'\'  and "UseStatus"=\'normal\' and "NextCheckDate" IS NULL AND "NextCheckCount" =0;'
                visual_sql_no_checkdate =sql_count

                cur = connection.cursor()
                cur.execute(visual_sql_no_checkdate)
                visual_sql_no_checkdate = cur.fetchall()
                info = [('None',visual_sql_no_checkdate[0][0])]
                dict_data['None'] = info
                info = tab_query_way([], [], [],info)
                dict_data['tab_data'] = info
            dict_data['normal'] = normal
            dict_data['warning'] = warning
            dict_data['danger'] = danger
            dict_data['None'] = info
            cur = connection.cursor()
            cur.execute(sql_count)
            count = cur.fetchall()  # 数量的总数
            if number == "All":
                data = dictdata(sql)
                data = tidy_dict(data)
                # cur = connection.cursor()
                # cur.execute(sql)
                # data = cur.fetchall()
                # data=tidy(data)
                # 针对data 进行判断和算出次数水平和周期天数
                # start_time = datetime.strptime(str(start).split(' ')[0], "%Y-%m-%d")  # 获取当前的日期
                # for i in range(0, len(data)):
                #     data[i] = list(data[i])
                #     if data[i][6] == None:
                #         data[i].append("null")
                #     else:
                #         time_end = datetime.strptime(str(data[i][6]).split(' ')[0],"%Y-%m-%d")  # 获取数据表里面的日期数
                #         days = time_end - start_time
                #         data[i].append(days.days)
                #     if data[i][9] == 0:
                #         data[i].append("null")
                #     else:
                #         data[i].append( data[i][9] - data[i][4])
                #     # 判断是那种状态
                #     if data[i][10] != "null" and data[i][11] != "null":
                #         if data[i][10] > mt_date[0]['Max'] * data[i][5] and data[i][11] > mt_count[0]['Max'] * \
                #                 data[i][3]:
                #             data[i].append('normal')
                #         if 0 <= data[i][10] <= mt_date[0]['Max'] * data[i][5] and 0 <= data[i][11]:
                #             data[i].append('warning')
                #         if 0 <= data[i][11] <= mt_count[0]['Max'] * data[i][3] and 0 <= data[i][10]:
                #             data[i].append('warning')
                #         if data[i][10] < 0 or data[i][11] < 0:
                #             data[i].append('danger')
                #     if data[i][10] == "null" and data[i][11] != "null":
                #         if data[i][11] > mt_count[0]['Max'] * data[i][3]:
                #             data[i].append('normal')
                #         if 0 <= data[i][11] <= mt_count[0]['Max'] * data[i][3]:
                #             data[i].append('warning')
                #         if data[i][11] < 0:
                #             data[i].append('danger')
                #     if data[i][10] != "null" and data[i][11] == "null":
                #         if data[i][10] > mt_date[0]['Max'] * data[i][5]:
                #             data[i].append('normal')
                #         if 0 <= data[i][10] <= mt_date[0]['Max'] * data[i][5]:
                #             data[i].append('warning')
                #         if data[i][10] < 0:
                #             data[i].append('danger')
                #     if data[i][10] == "null" and data[i][11] == "null":
                #         data[i].append('none')
                dict_data['data'] = data
                dict_data['page_count'] = count[0][0]
            if number != "All":
                number = int(number)
                count_page = count[0][0] // int(number)  # 总数除以一页显示多少条，得到总的页数
                if count[0][0] % number > 0:
                    count_page += 1
                if page <= count_page:
                    sql = sql + ' order by "Id" limit ' + str(number) + ' offset ' + str((page - 1)*number)
                    data = dictdata(sql)
                    data = tidy_dict(data)
                    # cur = connection.cursor()
                    # cur.execute(sql)
                    # data = cur.fetchall()
                    # data=tidy(data)
                    # 针对data 进行判断和算出次数水平和周期天数
                    # start_time = datetime.strptime(str(start).split(' ')[0], "%Y-%m-%d")  # 获取当前的日期
                    # for i in range(0, len(data)):
                    #     data[i] = list(data[i])
                    #     if data[i][6] == None:
                    #         data[i].append("null")
                    #     else:
                    #         time_end = datetime.strptime(str(data[i][6]).split(' ')[0], "%Y-%m-%d")  # 获取数据表里面的日期数
                    #         days = time_end - start_time
                    #         data[i].append(days.days)
                    #     if data[i][9] == 0:
                    #         data[i].append("null")
                    #     else:
                    #         data[i].append(data[i][9] - data[i][4])
                    #     # 判断是那种状态
                    #     if data[i][10] != "null" and data[i][11] != "null":
                    #         if data[i][10] > mt_date[0]['Max'] * data[i][5] and data[i][11] > mt_count[0]['Max'] * \
                    #                 data[i][3]:
                    #             data[i].append('normal')
                    #         if 0 <= data[i][10] <= mt_date[0]['Max'] * data[i][5] and 0 <= data[i][11]:
                    #             data[i].append('warning')
                    #         if 0 <= data[i][11] <= mt_count[0]['Max'] * data[i][3] and 0 <= data[i][10]:
                    #             data[i].append('warning')
                    #         if data[i][10] < 0 or data[i][11] < 0:
                    #             data[i].append('danger')
                    #     if data[i][10] == "null" and data[i][11] != "null":
                    #         if data[i][11] > mt_count[0]['Max'] * data[i][3]:
                    #             data[i].append('normal')
                    #         if 0 <= data[i][11] <= mt_count[0]['Max'] * data[i][3]:
                    #             data[i].append('warning')
                    #         if data[i][11] < 0:
                    #             data[i].append('danger')
                    #     if data[i][10] != "null" and data[i][11] == "null":
                    #         if data[i][10] > mt_date[0]['Max'] * data[i][5]:
                    #             data[i].append('normal')
                    #         if 0 <= data[i][10] <= mt_date[0]['Max'] * data[i][5]:
                    #             data[i].append('warning')
                    #         if data[i][10] < 0:
                    #             data[i].append('danger')
                    #     if data[i][10] == "null" and data[i][11] == "null":
                    #         data[i].append('none')
                    dict_data['data'] = data
                    dict_data['page_count'] = count_page
                if page > count_page:
                    sql = sql + ' order by "Id" limit ' + str(number) + ' offset ' + str(0)
                    data = dictdata(sql)
                    data = tidy_dict(data)
                    # cur = connection.cursor()
                    # cur.execute(sql)
                    # data = cur.fetchall()
                    # data=tidy(data)
                    # 针对data 进行判断和算出次数水平和周期天数
                    # start_time = datetime.strptime(str(start).split(' ')[0], "%Y-%m-%d")  # 获取当前的日期
                    # for i in range(0, len(data)):
                    #     data[i] = list(data[i])
                    #     if data[i][6] == None:
                    #         data[i].append("null")
                    #     else:
                    #         time_end = datetime.strptime(str(data[i][6]).split(' ')[0], "%Y-%m-%d")  # 获取数据表里面的日期数
                    #         days = time_end - start_time
                    #         data[i].append(days.days)
                    #     if data[i][9] == 0:
                    #         data[i].append("null")
                    #     else:
                    #         data[i].append(data[i][9] - data[i][4])
                    #     # 判断是那种状态
                    #     if data[i][10] != "null" and data[i][11] != "null":
                    #         if data[i][10] > mt_date[0]['Max'] * data[i][5] and data[i][11] > mt_count[0]['Max'] * \
                    #                 data[i][3]:
                    #             data[i].append('normal')
                    #         if 0 <= data[i][10] <= mt_date[0]['Max'] * data[i][5] and 0 <= data[i][11]:
                    #             data[i].append('warning')
                    #         if 0 <= data[i][11] <= mt_count[0]['Max'] * data[i][3] and 0 <= data[i][10]:
                    #             data[i].append('warning')
                    #         if data[i][10] < 0 or data[i][11] < 0:
                    #             data[i].append('danger')
                    #     if data[i][10] == "null" and data[i][11] != "null":
                    #         if data[i][11] > mt_count[0]['Max'] * data[i][3]:
                    #             data[i].append('normal')
                    #         if 0 <= data[i][11] <= mt_count[0]['Max'] * data[i][3]:
                    #             data[i].append('warning')
                    #         if data[i][11] < 0:
                    #             data[i].append('danger')
                    #     if data[i][10] != "null" and data[i][11] == "null":
                    #         if data[i][10] > mt_date[0]['Max'] * data[i][5]:
                    #             data[i].append('normal')
                    #         if 0 <= data[i][10] <= mt_date[0]['Max'] * data[i][5]:
                    #             data[i].append('warning')
                    #         if data[i][10] < 0:
                    #             data[i].append('danger')
                    #     if data[i][10] == "null" and data[i][11] == "null":
                    #         data[i].append('none')
                    dict_data['data'] = data
                    dict_data['page_count'] = count_page
            return restful.ok(data=dict_data)

            # dict_data['normal'] = normal
            # dict_data['warning'] = warning
            # dict_data['danger'] = danger
            # dict_data['None'] = info

        except Exception as e:
            return restful.params_error(message=repr(e))


#生成报表的数据
def maintain_record(request):
    if request.method == "POST":
        try:
            sn = str(request.POST.get('sn', ''))
            part_name = request.POST.get('partname', '')
            status = request.POST.get('status', '')
            s_time = request.POST.get('s_time', '')
            e_time = request.POST.get('e_time', '')
            user = str(request.POST.get('user', ''))
            locationId = str(request.POST.get('location', ''))
            # maintain_id = request.POST.getlist('maintain_id[]')
            tb1=['SN',"品名","保养周期(次数)","已使用次数","保养/检测周期(时间)","下次保养/检测时间","下次保养/检测次数","保养/检测人","保养/检测提醒提前天数","保养/检测提醒提前次数","位置","保养/检测次数预警比例","保养/检测时间预警比例","状态"]    #表头信息
            sql = 'select "SN","Spec","CheckCycleCount","UsedTimes","CheckCycle",to_char("NextCheckDate",\'yyyy-MM-dd\'),"NextCheckCount","Maintainer","WarningBeforeDays","WarningBeforeTimes","LocationId" from "PartItem" where "UseStatus"=\'normal\' '
            # sql = 'select "SN","PartName","CheckCycleCount","UsedTimes","CheckCycle","NextCheckDate","Maintainer","TrnDate","NextCheckCount","WarningBeforeDays","WarningBeforeTimes","LocationId" FROM "PartItem" WHERE "UseStatus"=\'normal\' '

            config = 'select "Type","Max" from "Configuration" where "Type" = \''+'mt_count'+'\' or "Type" = \''+'mt_date'+'\''
            mt_count = list(Configuration.objects.filter(Type="mt_count").values("Max"))
            mt_date = list(Configuration.objects.filter(Type="mt_date").values("Max"))
            if sn != "":
                sql = sql + 'AND "SN" = \'' + sn + '\''
            if part_name != "":
                sql = sql + 'AND "PartName" ilike \'%{0}%\''.format(part_name)
            if s_time != "":
                sql = sql + 'AND "TrnDate" >= \'{0}\''.format(s_time)
            if e_time != "":
                sql = sql + 'AND "TrnDate" <= \'{0}\''.format(e_time)
            if user != "":
                sql = sql + 'AND "Maintainer" = \'{0}\''.format(user)
            if locationId != "":
                sql = sql + 'AND "LocationId" = \'{0}\''.format(locationId)
            if status == "正常":
                sql = sql_select(sql, "正常", mt_count[0]['Max'], mt_date[0]['Max'])
            if status == "预警":
                sql = sql_select(sql, "预警", mt_count[0]['Max'], mt_date[0]['Max'])
            if status == "超标":
                sql = sql_select(sql, "超标", mt_count[0]['Max'], mt_date[0]['Max'])
            if status == "未设定":
                sql = sql_select(sql, "未设定", mt_count[0]['Max'], mt_date[0]['Max'])
            # if len(maintain_id) == 1:
            #     maintain_id = maintain_id[0]
            #     sql = sql + '"PartItem"."Id"=' + maintain_id
            # else:
            #     maintain_id = tuple(maintain_id)
            #     maintain_id = str(maintain_id)
            #     sql = sql + '"PartItem"."Id" in ' + maintain_id
            cur = connection.cursor()
            cur.execute(sql)
            data = cur.fetchall()
            cur.execute(config)
            stand = cur.fetchall()
            cur.close()
            stand_l = [attr[1] for attr in list(stand)] #[9,18]
            for j in range(0,len(data)):
                data[j]=list(data[j])
                data[j].extend(stand_l)
                if data[j][10] != None:
                    data[j][10] = (list(LocationLog.objects.filter(Id=data[j][10]).values("Location")))[0]["Location"]
                else:
                    data[j][10] = ""
            data.insert(0,tb1)
            sheet_name = "设备保养监控报表"
            filename = 'Equipment_Maintain' + str(int(time.time())) + '.xlsx'           #表格的名字
            file_root = settings.MEDIA_MAINTAIN_MONITOR_ROOT                          #表格的存入目录下
            file_url = settings.MEDIA_MAINTAIN_MONITOR_URL                            #表格的url
            data = statement_excle(request, data, sheet_name, file_root, file_url, filename)
            return restful.ok(data=data)
        except:
            return restful.params_error(message="download fail")

#生成报表并存入服务器的函数
def statement_excle(request,data,sheet_name,file_root,file_url,filename):
    current_time = datetime.strptime(str(datetime.now()).split(' ')[0], "%Y-%m-%d")  # 获取当前的日期
    wb = Workbook()                                                                 #创建对象
    wb.create_sheet(sheet_name, index=0)                                            #创建一个实例对象 第一张表格的表名字,和从第几个开始
    sheet = wb[sheet_name]
    for row in data:
        if len(row)==13:
            #row[5]-->下次保养时间  row[6] 下次保养次数  row[8] 提前提醒的天数 row[9] 提前提醒的次数
            #计算公式
            date_right  = row[4]*row[12]
            count_right = row[11]*row[2]
            #normal 正常：
            if row[8] == None and row[9] == None and row[5] != None and row[6] != 0:
                if row[6]-row[3]>count_right and (datetime.strptime(row[5], "%Y-%m-%d")-current_time).days> date_right:
                    row.append('正常')
            if row[8] == None and row[9] == None and row[5] == None and row[6] != 0:
                if row[6]-row[3]>count_right:
                    row.append('正常')
            if row[8] == None and row[9] == None and row[5] != None and row[6] == 0:
                if (datetime.strptime(row[5], "%Y-%m-%d")-current_time).days > date_right:
                    row.append('正常')
            #警告warning :
            if row[8] == None and row[9] == None and row[5] != None and row[6] != 0:
                if 0<=row[6]-row[3]<=count_right and (datetime.strptime(row[5], "%Y-%m-%d")-current_time).days >=0 :
                    row.append('预警')
                if 0 <=(datetime.strptime(row[5], "%Y-%m-%d")-current_time).days <= date_right and row[6]-row[3] >=0:
                    row.append('预警')
            if row[8] == None and row[9] == None and row[5] == None and row[6] != 0:
                if 0<=row[6]-row[3]<=count_right :
                    row.append('预警')
            if row[8] == None and row[9] == None and row[5] != None and row[6] == 0:
                if 0 <=(datetime.strptime(row[5], "%Y-%m-%d")-current_time).days <= date_right:
                    row.append('预警')
            #超标 danger:
            if row[8] == None and row[9] == None and row[5] != None and row[6] != 0:
                if row[6]-row[3]<0 or (datetime.strptime(row[5], "%Y-%m-%d")-current_time).days<0:
                    row.append('超标')
            if row[8] == None and row[9] == None and row[5] == None and row[6] != 0:
                if row[6]-row[3]<0:
                    row.append('超标')
            if row[8] == None and row[9] == None and row[5] != None and row[6] == 0:
                if (datetime.strptime(row[5], "%Y-%m-%d")-current_time).days<0:
                    row.append('超标')
            # 未设定的：
            if row[8] == None and row[9] == None and row[5] == None and row[6] == 0:
                row.append('未监控')

            # if row[8] != None and row[9] == None:
            #normal 正常：
            if row[8] != None and row[9] == None and row[5] != None and row[6] !=0:
                if row[6]-row[3]>count_right and (datetime.strptime(row[5], "%Y-%m-%d")-current_time).days> row[8]:
                    row.append('正常')
            if row[8] != None and row[9] == None and row[5] == None and row[6] != 0:
                if row[6]-row[3]>count_right:
                    row.append('正常')
            if row[8] != None and row[9] == None and row[5] != None and row[6] == 0:
                if (datetime.strptime(row[5], "%Y-%m-%d")-current_time).days > row[8]:
                    row.append('正常')
            #警告warning :
            if row[8] != None and row[9] == None and row[5] != None and row[6] != 0:
                if 0<=row[6]-row[3]<=count_right and (datetime.strptime(row[5], "%Y-%m-%d")-current_time).days >=0 :
                    row.append('预警')
                if 0 <=(datetime.strptime(row[5], "%Y-%m-%d")-current_time).days <= row[8] and row[6]-row[3] >=0:
                    row.append('预警')
            if row[8] != None and row[9] == None and row[5] == None and row[6] != 0:
                if 0<=row[6]-row[3]<=count_right :
                    row.append('预警')
            if row[8] != None and row[9] == None and row[5] != None and row[6] == 0:
                if 0 <=(datetime.strptime(row[5], "%Y-%m-%d")-current_time).days <= row[8]:
                    row.append('预警')
            #超标 danger:
            if row[8] != None and row[9] == None and row[5] != None and row[6] != 0:
                if row[6]-row[3]<0 or (datetime.strptime(row[5], "%Y-%m-%d")-current_time).days<0:
                    row.append('超标')
            if row[8] != None and row[9] == None and row[5] == None and row[6] != 0:
                if row[6]-row[3]<0:
                    row.append('超标')
            if row[8] != None and row[9] == None and row[5] != None and row[6] == 0:
                if (datetime.strptime(row[5], "%Y-%m-%d")-current_time).days<0:
                    row.append('超标')
            # 未设定的：
            if row[8] != None and row[9] == None and row[5] == None and row[6] == 0:
                row.append('未监控')

            # if row[8] == None and row[9] != None:
            #normal 正常：
            if row[8] == None and row[9] != None and row[5] != None and row[6] !=0:
                if row[6]-row[3]>row[9] and (datetime.strptime(row[5], "%Y-%m-%d")-current_time).days> date_right:
                    row.append('正常')
            if row[8] == None and row[9] != None and row[5] == None and row[6] != 0:
                if row[6]-row[3]>row[9]:
                    row.append('正常')
            if row[8] == None and row[9] != None and row[5] != None and row[6] == 0:
                if (datetime.strptime(row[5], "%Y-%m-%d")-current_time).days > date_right:
                    row.append('正常')
            #警告warning :
            if row[8] == None and row[9] != None and row[5] != None and row[6] != 0:
                if 0<=row[6]-row[3]<=row[9] and (datetime.strptime(row[5], "%Y-%m-%d")-current_time).days >=0 :
                    row.append('预警')
                if 0 <=(datetime.strptime(row[5], "%Y-%m-%d")-current_time).days <= date_right and row[6]-row[3] >=0:
                    row.append('预警')
            if row[8] == None and row[9] != None and row[5] == None and row[6] != 0:
                if 0<=row[6]-row[3]<=row[9] :
                    row.append('预警')
            if row[8] == None and row[9] != None and row[5] != None and row[6] == 0:
                if 0 <=(datetime.strptime(row[5], "%Y-%m-%d")-current_time).days <= date_right:
                    row.append('预警')
            #超标 danger:
            if row[8] == None and row[9] != None and row[5] != None and row[6] != 0:
                if row[6]-row[3]<0 or (datetime.strptime(row[5], "%Y-%m-%d")-current_time).days<0:
                    row.append('超标')
            if row[8] == None and row[9] != None and row[5] == None and row[6] != 0:
                if row[6]-row[3]<0:
                    row.append('超标')
            if row[8] == None and row[9] != None and row[5] != None and row[6] == 0:
                if (datetime.strptime(row[5], "%Y-%m-%d")-current_time).days<0:
                    row.append('超标')
            # 未设定的：
            if row[8] == None and row[9] != None and row[5] == None and row[6] == 0:
                row.append('未监控')

            # if row[8] != None and row[9] != None:
            #normal 正常：
            if row[8] != None and row[9] != None and row[5] != None and row[6] !=0:
                if row[6]-row[3]>row[9] and (datetime.strptime(row[5], "%Y-%m-%d")-current_time).days> row[8]:
                    row.append('正常')
            if row[8] != None and row[9] != None and row[5] == None and row[6] != 0:
                if row[6]-row[3]>row[9]:
                    row.append('正常')
            if row[8] != None and row[9] != None and row[5] != None and row[6] == 0:
                if (datetime.strptime(row[5], "%Y-%m-%d")-current_time).days > row[8]:
                    row.append('正常')
            #警告warning :
            if row[8] != None and row[9] != None and row[5] != None and row[6] != 0:
                if 0<=row[6]-row[3]<=row[9] and (datetime.strptime(row[5], "%Y-%m-%d")-current_time).days >=0 :
                    row.append('预警')
                if 0 <=(datetime.strptime(row[5], "%Y-%m-%d")-current_time).days <= row[8] and row[6]-row[3] >=0:
                    row.append('预警')
            if row[8] != None and row[9] != None and row[5] == None and row[6] != 0:
                if 0<=row[6]-row[3]<=row[9] :
                    row.append('预警')
            if row[8] != None and row[9] != None and row[5] != None and row[6] == 0:
                if 0 <=(datetime.strptime(row[5], "%Y-%m-%d")-current_time).days <= row[8]:
                    row.append('预警')
            #超标 danger:
            if row[8] != None and row[9] != None and row[5] != None and row[6] != 0:
                if row[6]-row[3]<0 or (datetime.strptime(row[5], "%Y-%m-%d")-current_time).days<0:
                    row.append('超标')
            if row[8] != None and row[9] != None and row[5] == None and row[6] != 0:
                if row[6]-row[3]<0:
                    row.append('超标')
            if row[8] != None and row[9] != None and row[5] != None and row[6] == 0:
                if (datetime.strptime(row[5], "%Y-%m-%d")-current_time).days<0:
                    row.append('超标')
            # 未设定的：
            if row[8] != None and row[9] != None and row[5] == None and row[6] == 0:
                row.append('未监控')
        sheet.append(row)

    wb.save(os.path.join(file_root, filename))
    file_url = request.build_absolute_uri(file_url + filename)
    data = [file_url]
    return data

#捞出捞取数据库所有NG率达到或超过预警区间的USN发邮件提醒给收件人 定时的功能在被使用在DBexcle app。views里面的函数crontab_test使用了
def Check_monitor_equipment():
    try:
        Total = 0
        Overdue = 0
        Warning = 0
        mt_count = list(Configuration.objects.filter(Type="mt_count").values('Max'))
        mt_date  = list(Configuration.objects.filter(Type="mt_date").values('Max'))
        count_stand = Configuration.objects.get(Type="mt_count")
        # 有保养人的 达到超标的数量和SN+ 没有第二保养人的数据
        M_danger = 'select "Maintainer",count(*),array_agg("SN") from "PartItem" where "Maintainer" IS NOT NULL and "SubMaintainers" IS NULL and "UseStatus"=\'normal\''
        M_danger = sql_select(M_danger,"超标",mt_count[0]['Max'],mt_date[0]['Max'])+' GROUP BY "Maintainer"'
        # 没有有保养人的 达到超标的数量和SN        + 没有第二保养人的数据
        N_danger = 'select "Maintainer",count(*),array_agg("SN") from "PartItem" where "Maintainer" IS NULL and "SubMaintainers" IS NULL  and "UseStatus"=\'normal\''
        N_danger = sql_select(N_danger,"超标",mt_count[0]['Max'],mt_date[0]['Max'])+' GROUP BY "Maintainer"'

        # 有保养人的 达到预警的数量和SN + 没有第二保养人的数据
        M_warning = 'select "Maintainer",count(*),array_agg("SN") FROM "PartItem" where "Maintainer" IS NOT NULL and "SubMaintainers" IS NULL  and "UseStatus"=\'normal\''
        M_warning = sql_select(M_warning,"预警",mt_count[0]['Max'],mt_date[0]['Max'])+' GROUP BY "Maintainer"'
        # 没有保养人的 达到预警的数量和SN + 没有第二保养人的数据
        N_warning = 'select "Maintainer",count(*),array_agg("SN") FROM "PartItem" where "Maintainer" IS NULL and "SubMaintainers" IS NULL  and "UseStatus"=\'normal\''
        N_warning = sql_select(N_warning, "预警", mt_count[0]['Max'], mt_date[0]['Max']) + ' GROUP BY "Maintainer"'

        # """ 第二保养人邮件提醒功能的实现 """
        # 有第二保养人的数据 超标数据
        SM_danger2 = 'select "SN","Maintainer","SubMaintainers" from "PartItem" where "SubMaintainers" IS NOT NULL and "UseStatus"=\'normal\''
        SM_danger2 = sql_select(SM_danger2, "超标",mt_count[0]['Max'],mt_date[0]['Max'])
        # 有第二保养人的数据预警数据
        SM_warning2 = 'select "SN","Maintainer","SubMaintainers" FROM "PartItem" where "SubMaintainers" IS NOT NULL and "UseStatus"=\'normal\''
        SM_warning2 = sql_select(SM_warning2, "预警",mt_count[0]['Max'],mt_date[0]['Max'])

        cur = connection.cursor()
        cur.execute(M_danger)
        danger = cur.fetchall()
        cur.execute(N_danger)
        danger2 = cur.fetchall()
        cur.execute(M_warning)
        warning = cur.fetchall()
        cur.execute(N_warning)
        warning2 = cur.fetchall()

        cur = connection.cursor()
        cur.execute(SM_warning2)
        SM_warning2=cur.fetchall()
        cur = connection.cursor()
        cur.execute(SM_danger2)
        SM_danger2 = cur.fetchall()
        #存在第二保养人的数据的邮件发送函数
        data_mail3 = mail_data_list(SM_danger2,SM_warning2)
        #有保养人   的超标和预警数据 没有第二保养人的data
        data_mail = mail_data_fun(danger, warning)

        # 没有保养人 的超标和预警数据 没有第二保养人的data
        data_mail2 = mail_data_fun(danger2, warning2)
        #第二保养人的邮件发送
        data_mail = mail_maintainer_sub(data_mail,data_mail3)

        #有保养人的 超标的和预警的数据的提醒的发送mail给保养人
        if len(data_mail) > 0:
            for i in range(0,len(data_mail)):
                email_maintainer = []
                user_mail_data = list(User.objects.filter(Name=data_mail[i][0]).values("Email"))
                if len(user_mail_data) != 0:
                    email_maintainer.append(str(user_mail_data[0]['Email']))
                    Total = data_mail[i][1] + data_mail[i][3]
                    Overdue = data_mail[i][1]
                    Warning = data_mail[i][3]
                    subject = "AEMS Lite System Notification for Equipment Maintenance"
                    content = """
<pre>
Dears,
    AEMS Lite system found that there is equipment to be maintained.
    Please take it to do maintenance in time...
    Total: """ + str(Total) + """
    Overdue: """ + str(Overdue) +"""
    Warning: """ + str(Warning) +"""


    THIS EMAIL WAS SENT BY PTS SERVER AUTOMATICALLY. PLEASE DON'T DIRECTLY REPLY!!!
    AEMS Lite System http://10.41.95.89:90/index/
</pre>
"""
                    mail.sendmail(email_maintainer, content, subject)
        #没有保养人的 超标的和预警的数据的提醒的发送给设定的邮件接收人
                time.sleep(13)
        if len(data_mail2) >0:
            email_1 = []
            receiver_list = str((count_stand.Reminders)).split(',')

            mail_user_data_info = list(User.objects.filter(Name__in=receiver_list).values("Email"))
            for i in range(0,len(mail_user_data_info)):
                email_1.append(mail_user_data_info[i]['Email'])
            # for i in range(len(receiver_list)):
            # email_1.append(receiver_list[i] + '@wistron.com')
            Total = int(data_mail2[0][1]) + int(data_mail2[0][3])
            Overdue = data_mail2[0][1]
            Warning = data_mail2[0][3]
            subject = "AEMS Lite System Notification for Equipment Maintenance"
            content = """
<pre>
Dears,
    AEMS Lite system found that there is equipment to be maintained.
    Please take it to do maintenance in time...
    Total: """+str(Total)+"""
    Overdue: """+str(Overdue)+"""
    Warning: """+str(Warning)+"""


    THIS EMAIL WAS SENT BY PTS SERVER AUTOMATICALLY. PLEASE DON'T DIRECTLY REPLY!!!
    AEMS Lite System http://10.41.95.89:90/index/
</pre>
"""
            mail.sendmail(email_1, content, subject)
    except Exception as e:
        return repr(e)
#数据重组的方法的小工具
def tab_query_way(tab_normal,tab_warning,tab_danger,tab_info):
    name_list=[]
    name_data=[]
    #获取名字的列表
    for i in range(0, len(tab_normal)):
        n_name = tab_normal[i][0]
        name_list.append(n_name)
    for j in range(0, len(tab_warning)):
        w_name = tab_warning[j][0]
        if w_name not in name_list:
            name_list.append(w_name)
    for k in range(0, len(tab_danger)):
        d_name = tab_danger[k][0]
        if d_name not in name_list:
            name_list.append(d_name)
    for l in range(0, len(tab_info)):
        i_name = tab_info[l][0]
        if i_name not in name_list:
            name_list.append(i_name)
    # 把数据变成[["name",0,0,0]] 初始化数据
    for m in range(0, len(name_list)):
        name_data.append([name_list[m], 0, 0, 0, 0])
    if len(tab_normal) > 0:
        for a in range(0, len(name_data)):
            for b in range(0, len(tab_normal)):
                if tab_normal[b][0] == name_data[a][0]:
                    name_data[a][1] = tab_normal[b][1]
    # 添加warning 数据
    if len(tab_warning) > 0:
        for a in range(0, len(name_data)):
            for b in range(0, len(tab_warning)):
                if tab_warning[b][0] == name_data[a][0]:
                    name_data[a][2] = tab_warning[b][1]
    # 添加danger 数据
    if len(tab_danger) > 0:
        for a in range(0, len(name_data)):
            for b in range(0, len(tab_danger)):
                if tab_danger[b][0] == name_data[a][0]:
                    name_data[a][3] = tab_danger[b][1]
    # 添加info 数据
    if len(tab_info) > 0:
        for a in range(0, len(name_data)):
            for b in range(0, len(tab_info)):
                if tab_info[b][0] == name_data[a][0]:
                    name_data[a][4] = tab_info[b][1]
    return name_data

#邮件接收人数据重组
def mail_data_fun(danger,warning):
    mail_list = []
    mail_data=[]
    for i in range(0,len(danger)):
        mail_list.append(danger[i][0])
    for j in range(0,len(warning)):
        warning_name = warning[j][0]
        if warning_name not in mail_list:
            mail_list.append(warning_name)
    for k in range(0,len(mail_list)):
        mail_data.append([mail_list[k],0,[],0,[]])
    if len(danger) > 0:
        for a in range(0, len(mail_data)):
            for b in range(0, len(danger)):
                if danger[b][0] == mail_data[a][0]:
                    mail_data[a][1] = danger[b][1]
                    mail_data[a][2].extend(danger[b][2])
    if len(warning)>0:
        for c in range(0,len(mail_data)):
            for d in range(0,len(warning)):
                if warning[d][0] == mail_data[c][0]:
                    mail_data[c][3] = warning[d][1]
                    mail_data[c][4].extend(warning[d][2])
    return mail_data
#第二保养人存在邮件接收者的数据重组
def mail_data_list(danger,warning):
    mingzi = []
    sns = []
    data_type = [] #[名字,超标数量,[SN,SN,SN,SN],预警,[SN,SN,SN,] ]
    for i in range(0, len(danger)):
        danger[i] = list(danger[i])
        s = danger[i][2].split(',')
        danger[i].remove(danger[i][2])
        danger[i].extend(s)
    for i in range(0, len(warning)):
        warning[i] = list(warning[i])
        s = warning[i][2].split(',')
        warning[i].remove(warning[i][2])
        warning[i].extend(s)

    for j in range(0, len(danger)):
        sns.append(danger[j][0])
        for k in range(1, len(danger[j])):
            if danger[j][k] not in mingzi:
                mingzi.append(danger[j][k])
    for j in range(0, len(warning)):
        if warning[j][0] not in sns:
            sns.append(warning[j][0])
        for k in range(1, len(warning[j])):
            if warning[j][k] not in mingzi:
                mingzi.append(warning[j][k])

    for i in range(0, len(mingzi)):
        if mingzi[i] != None:
            data_type.append([mingzi[i], 0, [],0,[]])

    for m in range(0, len(danger)):
        for n in range(0, len(danger[m])):
            for p in range(0, len(data_type)):
                if danger[m][n] == data_type[p][0]:
                    data_type[p][2].append(danger[m][0])
    for m in range(0, len(warning)):
        for n in range(0, len(warning[m])):
            for p in range(0, len(data_type)):
                if warning[m][n] == data_type[p][0]:
                    data_type[p][4].append(warning[m][0])

    for i in range(0, len(data_type)):
        data_type[i][1] = len(data_type[i][2])
        data_type[i][3] = len(data_type[i][4])

    return data_type
#第二保养人和有保养人的数据的整合在一起
def mail_maintainer_sub(warning,danger):
    data_list = []
    ming = []
    if len(warning) > 0 and len(danger) > 0:
        for i in range(0, len(warning)):
            for j in range(0, len(danger)):
                if warning[i][0] == danger[j][0]:
                    dan_sns, war_sns, data = [], [], []
                    data.append(warning[i][0])
                    data.append(danger[j][1] + warning[i][1])
                    dan_sns.extend(danger[j][2])
                    dan_sns.extend(warning[i][2])
                    data.append(dan_sns)
                    data.append(danger[j][3] + warning[i][3])
                    war_sns.extend(danger[j][4])
                    war_sns.extend(warning[i][4])
                    data.append(war_sns)
                    data_list.append(data)
        for i in range(0, len(data_list)):
            ming.append(data_list[i][0])
        for j in range(0, len(warning)):
            if warning[j][0] not in ming:
                data_list.append(warning[j])
        for j in range(0, len(danger)):
            if danger[j][0] not in ming:
                data_list.append(danger[j])
    if len(warning) > 0 and len(danger) == 0:
        data_list = warning
    if len(warning) == 0 and len(danger) > 0:
        data_list = danger
    #去重：
    for i in range(0,len(data_list)):
        data_list[i][2]=list(set(data_list[i][2]))
        data_list[i][4]=list(set(data_list[i][4]))
        data_list[i][1]=len(data_list[i][2])
        data_list[i][3]=len(data_list[i][4])

    return data_list

#第二版
def sql_select(sql,Status,count,date):
    # 针对SN的 WarningBeforeDays，WarningBeforeTimes有设定的分析
    user_sn_1 = ' ("WarningBeforeDays" IS NOT NULL AND "WarningBeforeTimes" IS NOT NULL) AND '
    user_sn_2 = ' ("WarningBeforeDays" IS NOT NULL AND "WarningBeforeTimes" IS NULL) AND '
    user_sn_3 = ' ("WarningBeforeDays" IS NULL AND "WarningBeforeTimes" IS NOT NULL) AND '
    user_sn_4 = ' ("WarningBeforeDays" IS NULL AND "WarningBeforeTimes" IS NULL) AND '
    # 正常的条件设置01
    n1 = '('+user_sn_4+'"NextCheckDate" IS NOT NULL AND "NextCheckCount" !=0 AND "NextCheckCount"-"UsedTimes">'+str(count)+'*"CheckCycleCount" AND extract(day from("NextCheckDate"-current_date))>'+str(date)+'*"CheckCycle")'
    n2 = '('+user_sn_4+'"NextCheckDate" IS NOT NULL AND "NextCheckCount" =0 AND extract(day from("NextCheckDate"-current_date))>'+str(date)+'*"CheckCycle")'
    n3 = '('+user_sn_4+'"NextCheckDate" IS NULL AND "NextCheckCount" !=0 AND "NextCheckCount"-"UsedTimes">'+str(count)+'*"CheckCycleCount")'
    # 正常的条件设置02
    n11 = '('+user_sn_3+'"NextCheckDate" IS NOT NULL AND "NextCheckCount" !=0 AND "NextCheckCount"-"UsedTimes">"WarningBeforeTimes" AND extract(day from("NextCheckDate"-current_date))>'+str(date)+'*"CheckCycle")'
    n21 = '('+user_sn_3+'"NextCheckDate" IS NOT NULL AND "NextCheckCount" =0 AND extract(day from("NextCheckDate"-current_date))>'+str(date)+'*"CheckCycle")'
    n31 = '('+user_sn_3+'"NextCheckDate" IS NULL AND "NextCheckCount" !=0 AND "NextCheckCount"-"UsedTimes">"WarningBeforeTimes")'
    # 正常的条件设置03
    n12 = '('+user_sn_2+'"NextCheckDate" IS NOT NULL AND "NextCheckCount" !=0 AND "NextCheckCount"-"UsedTimes">'+str(count)+'*"CheckCycleCount" AND extract(day from("NextCheckDate"-current_date))>"WarningBeforeDays")'
    n22 = '('+user_sn_2+'"NextCheckDate" IS NOT NULL AND "NextCheckCount" =0 AND extract(day from("NextCheckDate"-current_date))>"WarningBeforeDays")'
    n32 = '('+user_sn_2+'"NextCheckDate" IS NULL AND "NextCheckCount" !=0 AND "NextCheckCount"-"UsedTimes">'+str(count)+'*"CheckCycleCount")'
    # 正常的条件设置04
    n13 = '('+user_sn_1+'"NextCheckDate" IS NOT NULL AND "NextCheckCount" !=0 AND "NextCheckCount"-"UsedTimes">"WarningBeforeTimes" AND extract(day from("NextCheckDate"-current_date))>"WarningBeforeDays")'
    n23 = '('+user_sn_1+'"NextCheckDate" IS NOT NULL AND "NextCheckCount" =0 AND extract(day from("NextCheckDate"-current_date))>"WarningBeforeDays")'
    n33 = '('+user_sn_1+'"NextCheckDate" IS NULL AND "NextCheckCount" !=0 AND "NextCheckCount"-"UsedTimes">"WarningBeforeTimes")'


    # 预警的条件设置01
    w1 = '('+user_sn_4+'"NextCheckDate" IS NOT NULL AND "NextCheckCount" !=0 AND (("NextCheckCount"-"UsedTimes"<='+str(count)+'*"CheckCycleCount" AND "NextCheckCount"-"UsedTimes">= 0 AND extract(day from("NextCheckDate"-current_date))>=0)'
    w1 = w1+'OR ("NextCheckCount"-"UsedTimes">= 0 AND extract(day from("NextCheckDate"-current_date))>=0 AND extract(day from("NextCheckDate"-current_date))<='+str(date)+'*"CheckCycle")))'
    w2 = '('+user_sn_4+'"NextCheckDate" IS NOT NULL AND "NextCheckCount" =0 AND extract(day from("NextCheckDate"-current_date))<='+str(date)+'*"CheckCycle" AND extract(day from("NextCheckDate"-current_date))>=0)'
    w3 = '('+user_sn_4+'"NextCheckDate" IS NULL AND "NextCheckCount" !=0 AND "NextCheckCount"-"UsedTimes"<='+str(count)+'*"CheckCycleCount" AND "NextCheckCount"-"UsedTimes">=0)'
    # 预警的条件设置02
    w11 = '('+user_sn_3+'"NextCheckDate" IS NOT NULL AND "NextCheckCount" !=0 AND (("NextCheckCount"-"UsedTimes"<="WarningBeforeTimes" AND "NextCheckCount"-"UsedTimes">=0 AND extract(day from("NextCheckDate"-current_date))>=0)'
    w11 = w11 + 'OR ("NextCheckCount"-"UsedTimes">= 0 AND extract(day from("NextCheckDate"-current_date))>=0 AND extract(day from("NextCheckDate"-current_date))<=' + str(date) + '*"CheckCycle")))'
    w21 = '('+user_sn_3+'"NextCheckDate" IS NOT NULL AND "NextCheckCount" =0 AND extract(day from("NextCheckDate"-current_date))<=' + str(date) + '*"CheckCycle" AND extract(day from("NextCheckDate"-current_date))>=0)'
    w31 = '('+user_sn_3+'"NextCheckDate" IS NULL AND "NextCheckCount" !=0 AND "NextCheckCount"-"UsedTimes"<="WarningBeforeTimes" AND "NextCheckCount"-"UsedTimes">=0)'
    # 预警的条件设置03
    w12 = '('+user_sn_2+'"NextCheckDate" IS NOT NULL AND "NextCheckCount" !=0 AND (("NextCheckCount"-"UsedTimes"<=' + str(count) + '*"CheckCycleCount" AND "NextCheckCount"-"UsedTimes">= 0 AND extract(day from("NextCheckDate"-current_date))>=0)'
    w12 = w12 + 'OR ("NextCheckCount"-"UsedTimes">= 0 AND extract(day from("NextCheckDate"-current_date))>=0 AND extract(day from("NextCheckDate"-current_date))<="WarningBeforeDays")))'
    w22 = '('+user_sn_2+'"NextCheckDate" IS NOT NULL AND "NextCheckCount" =0 AND extract(day from("NextCheckDate"-current_date))<="WarningBeforeDays" AND extract(day from("NextCheckDate"-current_date))>=0)'
    w32 = '('+user_sn_2+'"NextCheckDate" IS NULL AND "NextCheckCount" !=0 AND "NextCheckCount"-"UsedTimes"<=' + str(count) + '*"CheckCycleCount" AND "NextCheckCount"-"UsedTimes">=0)'
    # 预警的条件设置04
    w13 = '('+user_sn_1+'"NextCheckDate" IS NOT NULL AND "NextCheckCount" !=0 AND (("NextCheckCount"-"UsedTimes"<="WarningBeforeTimes" AND "NextCheckCount"-"UsedTimes">=0 AND extract(day from("NextCheckDate"-current_date))>=0)'
    w13 = w13 + 'OR ("NextCheckCount"-"UsedTimes">=0 AND extract(day from("NextCheckDate"-current_date))>=0 AND extract(day from("NextCheckDate"-current_date))<="WarningBeforeDays")))'
    w23 = '('+user_sn_1+'"NextCheckDate" IS NOT NULL AND "NextCheckCount" =0 AND extract(day from("NextCheckDate"-current_date))<="WarningBeforeDays" AND extract(day from("NextCheckDate"-current_date))>=0)'
    w33 = '('+user_sn_1+'"NextCheckDate" IS NULL AND "NextCheckCount" !=0 AND "NextCheckCount"-"UsedTimes"<="WarningBeforeTimes" AND "NextCheckCount"-"UsedTimes">=0)'

    # 超标的条件设置
    c1 ='("NextCheckDate" IS NOT NULL AND "NextCheckCount" !=0 AND ("NextCheckCount"-"UsedTimes"<0 OR extract(day from("NextCheckDate"-current_date))<0))' #下次日期和下次保养次数不为空
    c2 ='("NextCheckDate" IS NOT NULL AND "NextCheckCount" =0 AND extract(day from("NextCheckDate"-current_date))<0)'                                      #下次日期不为空
    c3 ='("NextCheckDate" IS NULL AND "NextCheckCount" !=0 AND "NextCheckCount"-"UsedTimes"<0)'                                                            #下次保养次数不为空

    noon = '("NextCheckDate" IS NULL AND "NextCheckCount"=0)'
    if Status == "正常":
        sql = sql + ' AND('+n1+'OR'+n2+'OR'+n3+'OR'+n11+'OR'+n21+'OR'+n31+'OR'+n12+'OR'+n22+'OR'+n32+'OR'+n13+'OR'+n23+'OR'+n33+')'
    if Status == "预警":
        sql = sql + ' AND('+w1+'OR'+w2+'OR'+w3+'OR'+w11+'OR'+w21+'OR'+w31+'OR'+w12+'OR'+w22+'OR'+w32+'OR'+w13+'OR'+w23+'OR'+w33+')'
    if Status == "超标":
        sql = sql + ' AND('+c1+'OR'+c2+'OR'+c3+')'
    if Status == "未设定":
        sql = sql +' AND('+noon+')'
    return sql


#保养的查询结果的数据整理和清洗
def tidy(data):
    mt_count = list(Configuration.objects.filter(Type="mt_count").values("Max"))
    mt_date = list(Configuration.objects.filter(Type="mt_date").values("Max"))
    for i in range(len(data)):
        data[i] = list(data[i])
        if data[i][6] == None:
            data[i].append("null")
        else:
            start_time = datetime.strptime(str(datetime.now()).split(' ')[0], "%Y-%m-%d")
            time_end = datetime.strptime(str(data[i][6]).split(' ')[0], "%Y-%m-%d")  # 获取数据表里面的日期数
            days = time_end - start_time
            data[i].append(days.days)
        if data[i][9] == 0:
            data[i].append("null")
        else:
            data[i].append(data[i][9] - data[i][4])
        if data[i][10] ==None and data[i][11] ==None:
            if data[i][12] != "null" and data[i][13] != "null":
                if data[i][12]> mt_date[0]['Max']*data[i][5] and data[i][13]>mt_count[0]['Max']*data[i][3]:
                    data[i].append('normal')
                if 0<=data[i][12]<= mt_date[0]['Max']*data[i][5] and 0<=data[i][13]:
                    data[i].append('warning')
                if 0<=data[i][13]<= mt_count[0]['Max']*data[i][3] and 0<=data[i][12]:
                    data[i].append('warning')
                if data[i][12]<0 or data[i][13]<0:
                    data[i].append('danger')
            if data[i][12] != "null" and data[i][13] == "null":
                if data[i][12] > mt_date[0]['Max']*data[i][5]:
                    data[i].append('normal')
                if 0 <= data[i][12] <= mt_date[0]['Max']*data[i][5]:
                    data[i].append('warning')
                if data[i][12]<0:
                    data[i].append('danger')
            if data[i][12] == "null" and data[i][13] != "null":
                if data[i][13]>mt_count[0]['Max']*data[i][3]:
                    data[i].append('normal')
                if 0<=data[i][13]<= mt_count[0]['Max']*data[i][3]:
                    data[i].append('warning')
                if data[i][13]<0:
                    data[i].append('danger')
            if data[i][12] == "null" and data[i][13] == "null":
                data[i].append('none')
        if data[i][10] !=None and data[i][11] ==None:
            if data[i][12] != "null" and data[i][13] != "null":
                if data[i][12]> data[i][10] and data[i][13]>mt_count[0]['Max']*data[i][3]:
                    data[i].append('normal')
                if 0<=data[i][12]<= data[i][10] and 0<=data[i][13]:
                    data[i].append('warning')
                if 0<=data[i][13]<= mt_count[0]['Max']*data[i][3] and 0<=data[i][12]:
                    data[i].append('warning')
                if data[i][12]<0 or data[i][13]<0:
                    data[i].append('danger')
            if data[i][12] != "null" and data[i][13] == "null":
                if data[i][12] > data[i][10]:
                    data[i].append('normal')
                if 0 <= data[i][12] <= data[i][10]:
                    data[i].append('warning')
                if data[i][12]<0:
                    data[i].append('danger')
            if data[i][12] == "null" and data[i][13] != "null":
                if data[i][13]>mt_count[0]['Max']*data[i][3]:
                    data[i].append('normal')
                if 0<=data[i][13]<= mt_count[0]['Max']*data[i][3]:
                    data[i].append('warning')
                if data[i][13]<0:
                    data[i].append('danger')
            if data[i][12] == "null" and data[i][13] == "null":
                data[i].append('none')
        if data[i][10] ==None and data[i][11] !=None:
            if data[i][12] != "null" and data[i][13] != "null":
                if data[i][12]> mt_date[0]['Max']*data[i][5] and data[i][13]>data[i][11]:
                    data[i].append('normal')
                if 0<=data[i][12]<= mt_date[0]['Max']*data[i][5] and 0<=data[i][13]:
                    data[i].append('warning')
                if 0<=data[i][13]<= data[i][11] and 0<=data[i][12]:
                    data[i].append('warning')
                if data[i][12]<0 or data[i][13]<0:
                    data[i].append('danger')
            if data[i][12] != "null" and data[i][13] == "null":
                if data[i][12] > mt_date[0]['Max']*data[i][5]:
                    data[i].append('normal')
                if 0 <= data[i][12] <= mt_date[0]['Max']*data[i][5]:
                    data[i].append('warning')
                if data[i][12]<0:
                    data[i].append('danger')
            if data[i][12] == "null" and data[i][13] != "null":
                if data[i][13]>data[i][11]:
                    data[i].append('normal')
                if 0<=data[i][13]<= data[i][11]:
                    data[i].append('warning')
                if data[i][13]<0:
                    data[i].append('danger')
            if data[i][12] == "null" and data[i][13] == "null":
                data[i].append('none')
        if data[i][10] !=None and data[i][11] !=None:
            if data[i][12] != "null" and data[i][13] != "null":
                if data[i][12]> data[i][10] and data[i][13]>data[i][11]:
                    data[i].append('normal')
                if 0<=data[i][12]<= data[i][10] and 0<=data[i][13]:
                    data[i].append('warning')
                if 0<=data[i][13]<= data[i][11] and 0<=data[i][12]:
                    data[i].append('warning')
                if data[i][12]<0 or data[i][13]<0:
                    data[i].append('danger')
            if data[i][12] != "null" and data[i][13] == "null":
                if data[i][12] > data[i][10]:
                    data[i].append('normal')
                if 0 <= data[i][12] <= data[i][10]:
                    data[i].append('warning')
                if data[i][12]<0:
                    data[i].append('danger')
            if data[i][12] == "null" and data[i][13] != "null":
                if data[i][13]>data[i][11]:
                    data[i].append('normal')
                if 0<=data[i][13]<= data[i][11]:
                    data[i].append('warning')
                if data[i][13]<0:
                    data[i].append('danger')
            if data[i][12] == "null" and data[i][13] == "null":
                data[i].append('none')
    return data
def tidy_dict(data):
    mt_count = list(Configuration.objects.filter(Type="mt_count").values("Max"))
    mt_date = list(Configuration.objects.filter(Type="mt_date").values("Max"))
    start_time = datetime.strptime(str(datetime.now()).split(' ')[0], "%Y-%m-%d")  # 获取当前的日期
    for i in range(0, len(data)):
        if data[i]['LocationId'] != None:
            data[i]['location'] = (list(LocationLog.objects.filter(Id=data[i]['LocationId']).values("Location")))[0][
                'Location']
        else:
            data[i]['location'] = ""
        if data[i]['NextCheckDate'] == None:
            data[i]['stand_date'] = "null"
        else:
            time_end = datetime.strptime(str(data[i]['NextCheckDate']).split(' ')[0],
                                         "%Y-%m-%d")  # 获取数据表里面的日期数
            days = time_end - start_time
            data[i]['stand_date'] = days.days
        if data[i]['NextCheckCount'] == 0:
            data[i]['stand_count'] = 'null'
        else:
            data[i]['stand_count'] = data[i]['NextCheckCount'] - data[i]['UsedTimes']
        # 判断是那种状态
        if data[i]['WarningBeforeDays'] == None and data[i]['WarningBeforeTimes'] == None:
            if data[i]['stand_date'] != "null" and data[i]['stand_count'] != "null":
                if data[i]['stand_date'] > mt_date[0]['Max'] * data[i]['CheckCycle'] and data[i][
                    'stand_count'] > mt_count[0]['Max'] * data[i]['CheckCycleCount']:
                    data[i]['stand'] = 'normal'
                if 0 <= data[i]['stand_date'] <= mt_date[0]['Max'] * data[i]['CheckCycle'] and 0 <= data[i][
                    'stand_count']:
                    data[i]['stand'] = 'warning'
                if 0 <= data[i]['stand_count'] <= mt_count[0]['Max'] * data[i]['CheckCycleCount'] and 0 <= \
                        data[i]['stand_date']:
                    data[i]['stand'] = 'warning'
                if data[i]['stand_count'] < 0 or data[i]['stand_date'] < 0:
                    data[i]['stand'] = 'danger'
            if data[i]['stand_date'] == "null" and data[i]['stand_count'] != "null":
                if data[i]['stand_count'] > mt_count[0]['Max'] * data[i]['CheckCycleCount']:
                    data[i]['stand'] = 'normal'
                if 0 <= data[i]['stand_count'] <= mt_count[0]['Max'] * data[i]['CheckCycleCount']:
                    data[i]['stand'] = 'warning'
                if data[i]['stand_count'] < 0:
                    data[i]['stand'] = 'danger'
            if data[i]['stand_date'] != "null" and data[i]['stand_count'] == "null":
                if data[i]['stand_date'] > mt_date[0]['Max'] * data[i]['CheckCycle']:
                    data[i]['stand'] = 'normal'
                if 0 <= data[i]['stand_date'] <= mt_date[0]['Max'] * data[i]['CheckCycle']:
                    data[i]['stand'] = 'warning'
                if data[i]['stand_date'] < 0:
                    data[i]['stand'] = 'danger'
            if data[i]['stand_date'] == "null" and data[i]['stand_count'] == "null":
                data[i]['stand'] = 'none'
        if data[i]['WarningBeforeDays'] != None and data[i]['WarningBeforeTimes'] == None:
            if data[i]['stand_date'] != "null" and data[i]['stand_count'] != "null":
                if data[i]['stand_date'] > data[i]['WarningBeforeDays'] and data[i][
                    'stand_count'] > mt_count[0]['Max'] * data[i]['CheckCycleCount']:
                    data[i]['stand'] = 'normal'
                if 0 <= data[i]['stand_date'] <= data[i]['WarningBeforeDays'] and 0 <= data[i][
                    'stand_count']:
                    data[i]['stand'] = 'warning'
                if 0 <= data[i]['stand_count'] <= mt_count[0]['Max'] * data[i]['CheckCycleCount'] and 0 <= \
                        data[i]['stand_date']:
                    data[i]['stand'] = 'warning'
                if data[i]['stand_count'] < 0 or data[i]['stand_date'] < 0:
                    data[i]['stand'] = 'danger'
            if data[i]['stand_date'] == "null" and data[i]['stand_count'] != "null":
                if data[i]['stand_count'] > mt_count[0]['Max'] * data[i]['CheckCycleCount']:
                    data[i]['stand'] = 'normal'
                if 0 <= data[i]['stand_count'] <= mt_count[0]['Max'] * data[i]['CheckCycleCount']:
                    data[i]['stand'] = 'warning'
                if data[i]['stand_count'] < 0:
                    data[i]['stand'] = 'danger'
            if data[i]['stand_date'] != "null" and data[i]['stand_count'] == "null":
                if data[i]['stand_date'] > data[i]['WarningBeforeDays']:
                    data[i]['stand'] = 'normal'
                if 0 <= data[i]['stand_date'] <= data[i]['WarningBeforeDays']:
                    data[i]['stand'] = 'warning'
                if data[i]['stand_date'] < 0:
                    data[i]['stand'] = 'danger'
            if data[i]['stand_date'] == "null" and data[i]['stand_count'] == "null":
                data[i]['stand'] = 'none'
        if data[i]['WarningBeforeDays'] == None and data[i]['WarningBeforeTimes'] != None:
            if data[i]['stand_date'] != "null" and data[i]['stand_count'] != "null":
                if data[i]['stand_date'] > mt_date[0]['Max'] * data[i]['CheckCycle'] and data[i][
                    'stand_count'] > data[i]['WarningBeforeTimes']:
                    data[i]['stand'] = 'normal'
                if 0 <= data[i]['stand_date'] <= mt_date[0]['Max'] * data[i]['CheckCycle'] and 0 <= data[i][
                    'stand_count']:
                    data[i]['stand'] = 'warning'
                if 0 <= data[i]['stand_count'] <= data[i]['WarningBeforeTimes'] and 0 <= \
                        data[i]['stand_date']:
                    data[i]['stand'] = 'warning'
                if data[i]['stand_count'] < 0 or data[i]['stand_date'] < 0:
                    data[i]['stand'] = 'danger'
            if data[i]['stand_date'] == "null" and data[i]['stand_count'] != "null":
                if data[i]['stand_count'] > data[i]['WarningBeforeTimes']:
                    data[i]['stand'] = 'normal'
                if 0 <= data[i]['stand_count'] <= data[i]['WarningBeforeTimes']:
                    data[i]['stand'] = 'warning'
                if data[i]['stand_count'] < 0:
                    data[i]['stand'] = 'danger'
            if data[i]['stand_date'] != "null" and data[i]['stand_count'] == "null":
                if data[i]['stand_date'] > mt_date[0]['Max'] * data[i]['CheckCycle']:
                    data[i]['stand'] = 'normal'
                if 0 <= data[i]['stand_date'] <= mt_date[0]['Max'] * data[i]['CheckCycle']:
                    data[i]['stand'] = 'warning'
                if data[i]['stand_date'] < 0:
                    data[i]['stand'] = 'danger'
            if data[i]['stand_date'] == "null" and data[i]['stand_count'] == "null":
                data[i]['stand'] = 'none'
        if data[i]['WarningBeforeDays'] != None and data[i]['WarningBeforeTimes'] != None:
            if data[i]['stand_date'] != "null" and data[i]['stand_count'] != "null":
                if data[i]['stand_date'] > data[i]['WarningBeforeDays'] and data[i][
                    'stand_count'] > data[i]['WarningBeforeTimes']:
                    data[i]['stand'] = 'normal'
                if 0 <= data[i]['stand_date'] <= data[i]['WarningBeforeDays'] and 0 <= data[i][
                    'stand_count']:
                    data[i]['stand'] = 'warning'
                if 0 <= data[i]['stand_count'] <= data[i]['WarningBeforeTimes'] and 0 <= \
                        data[i]['stand_date']:
                    data[i]['stand'] = 'warning'
                if data[i]['stand_count'] < 0 or data[i]['stand_date'] < 0:
                    data[i]['stand'] = 'danger'
            if data[i]['stand_date'] == "null" and data[i]['stand_count'] != "null":
                if data[i]['stand_count'] > data[i]['WarningBeforeTimes']:
                    data[i]['stand'] = 'normal'
                if 0 <= data[i]['stand_count'] <= data[i]['WarningBeforeTimes']:
                    data[i]['stand'] = 'warning'
                if data[i]['stand_count'] < 0:
                    data[i]['stand'] = 'danger'
            if data[i]['stand_date'] != "null" and data[i]['stand_count'] == "null":
                if data[i]['stand_date'] > data[i]['WarningBeforeDays']:
                    data[i]['stand'] = 'normal'
                if 0 <= data[i]['stand_date'] <= data[i]['WarningBeforeDays']:
                    data[i]['stand'] = 'warning'
                if data[i]['stand_date'] < 0:
                    data[i]['stand'] = 'danger'
            if data[i]['stand_date'] == "null" and data[i]['stand_count'] == "null":
                data[i]['stand'] = 'none'
    return data
##查询函数的sql语句调用
# def sql_select(sql,Status,count,date):
#     # 正常的条件设置
#     n1 = '("NextCheckDate" IS NOT NULL AND "NextCheckCount" !=0 AND "NextCheckCount"-"UsedTimes">'+str(count)+'*"CheckCycleCount" AND extract(day from("NextCheckDate"-current_date))>'+str(date)+'*"CheckCycle")'
#     n2 = '("NextCheckDate" IS NOT NULL AND "NextCheckCount" =0 AND extract(day from("NextCheckDate"-current_date))>'+str(date)+'*"CheckCycle")'
#     n3 = '("NextCheckDate" IS NULL AND "NextCheckCount" !=0 AND "NextCheckCount"-"UsedTimes">'+str(count)+'*"CheckCycleCount")'
#     # 预警的条件设置
#     w1 = '("NextCheckDate" IS NOT NULL AND "NextCheckCount" !=0 AND (("NextCheckCount"-"UsedTimes"<='+str(count)+'*"CheckCycleCount" AND "NextCheckCount"-"UsedTimes">= 0 AND extract(day from("NextCheckDate"-current_date))>=0)'
#     w1 = w1+'OR ("NextCheckCount"-"UsedTimes">= 0 AND extract(day from("NextCheckDate"-current_date))>=0 AND extract(day from("NextCheckDate"-current_date))<='+str(date)+'*"CheckCycle")))'
#     w2 = '("NextCheckDate" IS NOT NULL AND "NextCheckCount" =0 AND extract(day from("NextCheckDate"-current_date))<='+str(date)+'*"CheckCycle" AND extract(day from("NextCheckDate"-current_date))>=0)'
#     w3 = '("NextCheckDate" IS NULL AND "NextCheckCount" !=0 AND "NextCheckCount"-"UsedTimes"<='+str(count)+'*"CheckCycleCount" AND "NextCheckCount"-"UsedTimes">=0)'
#     # 超标的条件设置
#     c1 ='("NextCheckDate" IS NOT NULL AND "NextCheckCount" !=0 AND ("NextCheckCount"-"UsedTimes"<0 OR extract(day from("NextCheckDate"-current_date))<0))'
#     c2 ='("NextCheckDate" IS NOT NULL AND "NextCheckCount" =0 AND extract(day from("NextCheckDate"-current_date))<0)'
#     c3 ='("NextCheckDate" IS NULL AND "NextCheckCount" !=0 AND "NextCheckCount"-"UsedTimes"<0)'
#     # 未设定的条件设置
#     noon = '("NextCheckDate" IS NULL AND "NextCheckCount"=0)'
#     if Status == "正常":
#         sql = sql + ' AND('+n1+'OR'+n2+'OR'+n3+')'
#     if Status == "预警":
#         sql = sql + ' AND('+w1+'OR'+w2+'OR'+w3+')'
#     if Status == "超标":
#         sql = sql + ' AND('+c1+'OR'+c2+'OR'+c3+')'
#     if Status == "未设定":
#         sql = sql +' AND('+noon+')'
#     return sql

#查询函数的sql语句调用
# def sql_select(sql,Status,count,date):
#     # 正常的条件设置
#     n1 = '("NextCheckDate" IS NOT NULL AND "NextCheckCount" !=0 AND "NextCheckCount"-"UsedTimes">' + str(count) + '*"CheckCycleCount" AND extract(day from("NextCheckDate"-current_date))>' + str(date) + '*"CheckCycle")'
#     n2 = '("NextCheckDate" IS NOT NULL AND "NextCheckCount" =0 AND extract(day from("NextCheckDate"-current_date))>' + str(date) + '*"CheckCycle")'
#     n3 = '("NextCheckDate" IS NULL AND "NextCheckCount" !=0 AND "NextCheckCount"-"UsedTimes">' + str(count) + '*"CheckCycleCount")'
#     #针对SN的 WarningBeforeDays，WarningBeforeTimes有设定的分析
#     user_sn_1=' ("WarningBeforeDays" IS NOT NULL AND "WarningBeforeTimes" IS NOT NULL) AND '
#     user_sn_2=' ("WarningBeforeDays" IS NOT NULL AND "WarningBeforeTimes" IS NULL) AND '
#     user_sn_3=' ("WarningBeforeDays" IS NULL AND "WarningBeforeTimes" IS NOT NULL) AND '
#     user_sn_4=' ("WarningBeforeDays" IS NULL AND "WarningBeforeTimes" IS NULL) AND '
#
#
#     # 预警的条件设置01
#     w1 = '('+user_sn_4+'"NextCheckDate" IS NOT NULL AND "NextCheckCount" !=0 AND (("NextCheckCount"-"UsedTimes"<='+str(count)+'*"CheckCycleCount" AND "NextCheckCount"-"UsedTimes">= 0 AND extract(day from("NextCheckDate"-current_date))>=0)'
#     w1 = w1+'OR ("NextCheckCount"-"UsedTimes">= 0 AND extract(day from("NextCheckDate"-current_date))>=0 AND extract(day from("NextCheckDate"-current_date))<='+str(date)+'*"CheckCycle")))'
#     w2 = '('+user_sn_4+'"NextCheckDate" IS NOT NULL AND "NextCheckCount" =0 AND extract(day from("NextCheckDate"-current_date))<='+str(date)+'*"CheckCycle" AND extract(day from("NextCheckDate"-current_date))>=0)'
#     w3 = '('+user_sn_4+'"NextCheckDate" IS NULL AND "NextCheckCount" !=0 AND "NextCheckCount"-"UsedTimes"<='+str(count)+'*"CheckCycleCount" AND "NextCheckCount"-"UsedTimes">=0)'
#     # 预警的条件设置02
#     w11 = '('+user_sn_3+'"NextCheckDate" IS NOT NULL AND "NextCheckCount" !=0 AND (("NextCheckCount"-"UsedTimes"<=' + str(
#         count) + '*"CheckCycleCount"+"WarningBeforeTimes" AND "NextCheckCount"-"UsedTimes">="WarningBeforeTimes" AND extract(day from("NextCheckDate"-current_date))>=0)'
#     w11 = w11 + 'OR ("NextCheckCount"-"UsedTimes">= "WarningBeforeTimes" AND extract(day from("NextCheckDate"-current_date))>=0 AND extract(day from("NextCheckDate"-current_date))<=' + str(
#         date) + '*"CheckCycle")))'
#     w21 = '('+user_sn_3+'"NextCheckDate" IS NOT NULL AND "NextCheckCount" =0 AND extract(day from("NextCheckDate"-current_date))<=' + str(
#         date) + '*"CheckCycle" AND extract(day from("NextCheckDate"-current_date))>=0)'
#     w31 = '('+user_sn_3+'"NextCheckDate" IS NULL AND "NextCheckCount" !=0 AND "NextCheckCount"-"UsedTimes"<=' + str(
#         count) + '*"CheckCycleCount"+"WarningBeforeTimes" AND "NextCheckCount"-"UsedTimes">="WarningBeforeTimes")'
#     # 预警的条件设置03
#     w12 = '('+user_sn_2+'"NextCheckDate" IS NOT NULL AND "NextCheckCount" !=0 AND (("NextCheckCount"-"UsedTimes"<=' + str(count) + '*"CheckCycleCount" AND "NextCheckCount"-"UsedTimes">= 0 AND extract(day from("NextCheckDate"-current_date))>="WarningBeforeDays")'
#     w12 = w12 + 'OR ("NextCheckCount"-"UsedTimes">= 0 AND extract(day from("NextCheckDate"-current_date))>="WarningBeforeDays" AND extract(day from("NextCheckDate"-current_date))<='+str(date) + '*"CheckCycle"+"WarningBeforeDays")))'
#     w22 = '('+user_sn_2+'"NextCheckDate" IS NOT NULL AND "NextCheckCount" =0 AND extract(day from("NextCheckDate"-current_date))<='+str(date) + '*"CheckCycle"+"WarningBeforeDays" AND extract(day from("NextCheckDate"-current_date))>="WarningBeforeDays")'
#     w32 = '('+user_sn_2+'"NextCheckDate" IS NULL AND "NextCheckCount" !=0 AND "NextCheckCount"-"UsedTimes"<=' + str(count) + '*"CheckCycleCount" AND "NextCheckCount"-"UsedTimes">=0)'
#     # 预警的条件设置04
#     w13 = '('+user_sn_1+'"NextCheckDate" IS NOT NULL AND "NextCheckCount" !=0 AND (("NextCheckCount"-"UsedTimes"<='+str(count)+'*"CheckCycleCount"+"WarningBeforeTimes" AND "NextCheckCount"-"UsedTimes">="WarningBeforeTimes" AND extract(day from("NextCheckDate"-current_date))>="WarningBeforeDays")'
#     w13 = w13 + 'OR ("NextCheckCount"-"UsedTimes">= "WarningBeforeTimes" AND extract(day from("NextCheckDate"-current_date))>="WarningBeforeDays" AND extract(day from("NextCheckDate"-current_date))<='+str(date)+'*"CheckCycle"+"WarningBeforeDays")))'
#     w23 = '('+user_sn_1+'"NextCheckDate" IS NOT NULL AND "NextCheckCount" =0 AND extract(day from("NextCheckDate"-current_date))<='+str(date)+'*"CheckCycle"+"WarningBeforeDays" AND extract(day from("NextCheckDate"-current_date))>="WarningBeforeDays")'
#     w33 = '('+user_sn_1+'"NextCheckDate" IS NULL AND "NextCheckCount" !=0 AND "NextCheckCount"-"UsedTimes"<='+str(count)+'*"CheckCycleCount"+"WarningBeforeTimes" AND "NextCheckCount"-"UsedTimes">="WarningBeforeTimes")'
#
#     # 超标的条件设置01中情况
#     c1 ='('+user_sn_4+'"NextCheckDate" IS NOT NULL AND "NextCheckCount" !=0 AND ("NextCheckCount"-"UsedTimes"<0 OR extract(day from("NextCheckDate"-current_date))<0))' #下次日期和下次保养次数不为空
#     c2 ='('+user_sn_4+'"NextCheckDate" IS NOT NULL AND "NextCheckCount" =0 AND extract(day from("NextCheckDate"-current_date))<0)'                                      #下次日期不为空
#     c3 ='('+user_sn_4+'"NextCheckDate" IS NULL AND "NextCheckCount" !=0 AND "NextCheckCount"-"UsedTimes"<0)'                                                            #下次保养次数不为空
#     # 超标的条件设置02中情况
#     c11 = '(' + user_sn_3 + '"NextCheckDate" IS NOT NULL AND "NextCheckCount" !=0 AND ("NextCheckCount"-"UsedTimes"<"WarningBeforeTimes" OR extract(day from("NextCheckDate"-current_date))<0))'  # 下次日期和下次保养次数不为空
#     c21 = '(' + user_sn_3 + '"NextCheckDate" IS NOT NULL AND "NextCheckCount" =0 AND extract(day from("NextCheckDate"-current_date))<0)'  # 下次日期不为空
#     c31 = '(' + user_sn_3 + '"NextCheckDate" IS NULL AND "NextCheckCount" !=0 AND "NextCheckCount"-"UsedTimes"<"WarningBeforeTimes")'  # 下次保养次数不为空
#     # 超标的条件设置03中情况
#     c12 = '(' + user_sn_2 + '"NextCheckDate" IS NOT NULL AND "NextCheckCount" !=0 AND ("NextCheckCount"-"UsedTimes"<0 OR extract(day from("NextCheckDate"-current_date))<"WarningBeforeDays"))'  # 下次日期和下次保养次数不为空
#     c22 = '(' + user_sn_2 + '"NextCheckDate" IS NOT NULL AND "NextCheckCount" =0 AND extract(day from("NextCheckDate"-current_date))<"WarningBeforeDays")'  # 下次日期不为空
#     c32 = '(' + user_sn_2 + '"NextCheckDate" IS NULL AND "NextCheckCount" !=0 AND "NextCheckCount"-"UsedTimes"<0)'  # 下次保养次数不为空
#     # 超标的条件设置04中情况
#     c13 = '(' + user_sn_1 + '"NextCheckDate" IS NOT NULL AND "NextCheckCount" !=0 AND ("NextCheckCount"-"UsedTimes"<"WarningBeforeTimes" OR extract(day from("NextCheckDate"-current_date))<"WarningBeforeDays"))'  # 下次日期和下次保养次数不为空
#     c23 = '(' + user_sn_1 + '"NextCheckDate" IS NOT NULL AND "NextCheckCount" =0 AND extract(day from("NextCheckDate"-current_date))<"WarningBeforeDays")'  # 下次日期不为空
#     c33 = '(' + user_sn_1 + '"NextCheckDate" IS NULL AND "NextCheckCount" !=0 AND "NextCheckCount"-"UsedTimes"<"WarningBeforeTimes")'  # 下次保养次数不为空
#     # 未设定的条件设置
#     noon = '("NextCheckDate" IS NULL AND "NextCheckCount"=0)'
#     if Status == "正常":
#         sql = sql + ' AND('+n1+'OR'+n2+'OR'+n3+')'
#     if Status == "预警":
#         sql = sql + ' AND('+w1+'OR'+w2+'OR'+w3+'OR'+w11+'OR'+w21+'OR'+w31+'OR'+w12+'OR'+w22+'OR'+w32+'OR'+w13+'OR'+w23+'OR'+w33+')'
#     if Status == "超标":
#         sql = sql + ' AND('+c1+'OR'+c2+'OR'+c3+'OR'+c11+'OR'+c21+'OR'+c31+'OR'+c12+'OR'+c22+'OR'+c32+'OR'+c13+'OR'+c23+'OR'+c33+')'
#     if Status == "未设定":
#         sql = sql +' AND('+noon+')'
#     return sql





