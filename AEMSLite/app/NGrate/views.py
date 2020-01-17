from django.shortcuts import render
from django.shortcuts import render, redirect
# from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from app.login.models import User,Department,Customer,BudgetCodeForm,PartItem,PartItemResult,MaintenanceLog,Configuration
from app.login.views import Update_User_IsActivated
from django.views.generic.base import View
from django.db import connection
from django.http import HttpResponseRedirect,HttpResponse
from app import restful,mail
from app.access_control import access_control
from datetime import datetime,timedelta,date
from django.conf import settings
import random
import string
import os
import time
from openpyxl import load_workbook,Workbook
import json
#生成报表并存入服务器的函数
def statement_excle(request,data,sheet_name,file_root,file_url,filename):
    wb = Workbook()
    index = 0
    wb.create_sheet(sheet_name, index=index)
    sheet = wb[sheet_name]
    for row in data:
        row = list(row)
        if len(row)==7:
            #正常
            if row[2] > row[4]:
                row.append("正常")
            #预警
            if row[2] <= row[4] <=row[3]:
                row.append("预警")
            #超标
            if row[3] < row[4]:
                row.append("超标")
        sheet.append(row)
    wb.save(os.path.join(file_root, filename))
    file_url = request.build_absolute_uri(file_url + filename)
    data = [file_url]
    return data

# 设备NG率的监控类以及生成报表
class MonitorEquipment(View):
    def get(self, request):
        try:
            start = datetime.now()  # 这里默认是查询前几周的数据
            # end_time = datetime.now()  # 这里默认是查询前几周的数据
            # start_time = date(2017, 1, 1)  # 这里默认是查询前几周的数据
            start_select = (list(PartItem.objects.order_by("-TrnDate").filter(TrnDate__lte=start,UseStatus='normal').values("TrnDate")))[0]['TrnDate']
            delta_select = timedelta(days=7)
            end_select = start_select - delta_select
            # delta = timedelta(days=21)
            # start_time = end_time-delta
            # 页面html显示需要的数据+分页效果
            page = int(request.GET.get('page'))
            number = request.GET.get('num')
            dict_data = {}
            count = PartItem.objects.order_by("Id").filter(TrnDate__range=(end_select, start_select),UseStatus='normal').count()
            # 饼状图需要的数据
            sql = 'SELECT "PartName", COUNT("PartName") FROM "PartItem" where "UseStatus"=\'normal\' and "TrnDate">= \'{0}\' AND "TrnDate"<= \'{1}\''.format(
                end_select, start_select)
            try:
                parmeter_stands = Configuration.objects.get(Type="NG_range")
                parmeter_stands_min = str(parmeter_stands.Min)
                parmeter_stands_max = str(parmeter_stands.Max)
                sql1 = sql + 'AND "NGRate" < \'' + parmeter_stands_min + '\' group by "PartName"'   # 正常
                sql2 = sql + 'AND "NGRate" >= \'' + parmeter_stands_min + '\'' \
                       + 'AND "NGRate" <= \'' + parmeter_stands_max + '\' group by "PartName"'      # 预警
                sql3 = sql + 'AND "NGRate" > \'' + parmeter_stands_max + '\' group by "PartName"'   # 超标
                cur = connection.cursor()
                cur.execute(sql1)
                normal = cur.fetchall()
                cur.execute(sql2)
                warning = cur.fetchall()
                cur.execute(sql3)
                danger = cur.fetchall()
                cur.close()
                dict_data['normal'] = normal
                dict_data['warning'] = warning
                dict_data['danger'] = danger
            except:
                sql4 = sql
                cur = connection.cursor()
                cur.execute(sql4)
                normal = cur.fetchall()
                dict_data['normal'] = normal
            # 页面显示需要的数据
            try:
                if number == "All":
                    data = PartItem.objects.order_by("Id").filter(TrnDate__range=(end_select, start_select),UseStatus='normal')
                    limit_value = Configuration.objects.filter(Type="NG_range").values("Max", "Min", "Id")
                    limit_value = list(limit_value)
                    data = list(data.values())
                    dict_data['data'] = data
                    dict_data['limit_value'] = limit_value
                    dict_data['page_count'] = count
                    return restful.ok(data=dict_data)
                if number != "All":
                    number = int(number)
                    page_num = count // number  # 总共多少页
                    if count % number > 0:
                        page_num = page_num + 1
                    if page_num >= page:
                        data = PartItem.objects.order_by("Id").filter(TrnDate__range=(end_select, start_select),UseStatus='normal')[
                               (page - 1) * number:number * page]
                        limit_value = Configuration.objects.filter(Type="NG_range").values("Max", "Min", "Id")
                        limit_value = list(limit_value)
                        data = list(data.values())
                        dict_data['data'] = data
                        dict_data['limit_value'] = limit_value
                        dict_data['page_count'] = page_num
                        dict_data['select_end'] = end_select
                        dict_data['select_start'] = start_select
                        return restful.ok(data=dict_data)
                    else:
                        return restful.params_error(message='it had no others page')
            except Exception as e:
                return restful.params_error(message=repr(e))
        except Exception as e:
            return restful.params_error(message=repr(e))
    # 生成报表的形式
    def post(self, request):
        try:
            NG_ids = request.POST.getlist('NG_ids[]')
            sql1 = '''SELECT a.attname as name FROM pg_class as c,pg_attribute as a where c.relname = '%s' 
            and a.attrelid = c.oid and a.attnum>0''' % 'PartItem'
            sql2 = 'SELECT "SN","PartName","Configuration"."Min","Configuration"."Max","NGRate","ErrorCounts"' \
                   ',"UsedTimes" FROM "PartItem" INNER JOIN "Configuration" on "Configuration"."Id"=2 where '
            if len(NG_ids) == 1:
                NG_ids = NG_ids[0]
                sql3 = sql2 + '"PartItem"."Id"=' + NG_ids
            else:
                NG_ids = tuple(NG_ids)
                NG_ids = str(NG_ids)
                sql3 = sql2 + '"PartItem"."Id" in ' + NG_ids
            cur = connection.cursor()
            cur.execute(sql3)
            data = cur.fetchall()
            cur.close()
            cur2 = connection.cursor()
            cur2.execute(sql1)
            statement_data = cur2.fetchall()
            cur2.close()
            statement_data = [attr[0] for attr in statement_data]
            data1 = []
            data1.append(statement_data[statement_data.index("SN")])
            data1.append(statement_data[statement_data.index("PartName")])
            data1.append('Min')
            data1.append('Max')
            data1.append(statement_data[statement_data.index("NGRate")])
            data1.append(statement_data[statement_data.index("ErrorCounts")])
            data1.append(statement_data[statement_data.index("UsedTimes")])
            data1.append("status")
            data.insert(0, data1)
            sheet_name = "NG率监控表单"
            time_num = int(time.time())
            time_num = str(time_num)
            file_root = settings.MEDIA_MONITOR_ROOT
            file_url = settings.MEDIA_MONITOR_URL
            filename = 'Equipment_monitor' + time_num + '.xlsx'
            data = statement_excle(request, data, sheet_name, file_root, file_url, filename)
            return restful.ok(data=data)
        except:
            return restful.params_error(message="download fail")

#设置之前获取数据
@access_control
def setup_before(request):
    if request.method == "GET":
        try:
            data1 = list(Configuration.objects.filter(Type="NG_range").values("Max","Min","Reminders"))
            data_user = list(User.objects.filter(Role="admin").values("Name"))+list(User.objects.filter(Role="equipment_room").values("Name"))
            data = data1+data_user
            return restful.ok(data=data)
        except Exception as e:
            return restful.params_error(repr(e))

# 设置预警区间
@access_control
def setup_parameter(request):
    if request.method == "POST":
        try:
            min = request.POST['min']
            max = request.POST['max']
            mail_receiver = request.POST.getlist('mail_receiver[]')
            mail_receiver = list(mail_receiver)
            NG_monitor_type = "NG_range"
            parmeter_id = Configuration.objects.filter(Type=NG_monitor_type)
            mail_receiver_count = ",".join(mail_receiver)
            if parmeter_id:
                Configuration.objects.filter(Type=NG_monitor_type).update(Max=max, Min=min,
                                                                          Reminders=mail_receiver_count)
                return restful.ok(message="setup parameter success")
            else:
                Configuration.objects.create(Type=NG_monitor_type, Max=max, Min=min, Reminders=mail_receiver_count)
                return restful.ok(message="setup parameter success")
        except:
            return restful.params_error(message="setup parameter fail")

# 查询数据的信息 模糊查询数据
@access_control
def monitor_query_info(request):
    if request.method == "POST":
        try:
            page = int(request.POST['page'])
            number = request.POST['num']
            sn = str(request.POST['sn'])
            part_name = request.POST['part_name']
            status = request.POST['status']
            start_time = request.POST['start_tim']
            end_time = request.POST['end_tim']
            NG_monitor_type = "NG_range"
            #时间区间回滚到初始状态
            start = datetime.now()
            start_select = (list(PartItem.objects.order_by("-TrnDate").filter(TrnDate__lte=start,UseStatus='normal').values("TrnDate")))[0]['TrnDate']
            delta_select = timedelta(days=7)
            end_select = start_select - delta_select
            visual_callback = ' AND "TrnDate">= \'{0}\' AND "TrnDate"<= \'{1}\''.format(end_select, start_select)
            sql_count = 'select count(*) FROM "PartItem" where "UseStatus"=\'normal\' '
            sql = 'select "Id","SN","OSN","PN","PartName","Spec","UsedTimes","CreatedTime","UpdatedTime","CheckCycle"' \
                  ',"CheckCycleCount","NextCheckCount","NextCheckDate","ErrorCounts","TrnDate","NGRate" ' \
                  'FROM "PartItem" WHERE "UseStatus"=\'normal\''
            visual_sql = 'SELECT "PartName", COUNT("PartName") FROM "PartItem" where "UseStatus"=\'normal\' '

            parmeter_stands = Configuration.objects.get(Type=NG_monitor_type)
            limit_value = [parmeter_stands.Min, parmeter_stands.Max]
            parmeter_stands_min = str(parmeter_stands.Min)
            parmeter_stands_max = str(parmeter_stands.Max)
            dict_data = {}
            normal = []
            warning = []
            danger = []
            if start_time == "" and end_time == "" and part_name == "" and sn == "" and status =="":
                sql = sql +visual_callback
                visual_sql = visual_sql +visual_callback
            if status == "":
                if start_time != "":
                    sql = sql + 'AND "TrnDate" >= \'%{0}%\''.format(start_time)
                    sql_count = sql_count + 'AND "TrnDate" >= \'%{0}%\''.format(start_time)
                    visual_sql = visual_sql + 'AND "TrnDate" >= \'%{0}%\''.format(start_time)
                if end_time != "":
                    sql = sql + 'AND "TrnDate" <= \'%{0}%\''.format(end_time)
                    sql_count = sql_count + 'AND "TrnDate" <= \'%{0}%\''.format(end_time)
                    visual_sql = visual_sql + 'AND "TrnDate" <= \'%{0}%\''.format(end_time)
                if part_name != "":
                    sql = sql + 'AND "PartName" like \'%{0}%\''.format(part_name)
                    sql_count = sql_count + 'AND "PartName" like \'%{0}%\''.format(part_name)
                    visual_sql = visual_sql + 'AND "PartName" like \'%{0}%\''.format(part_name)
                if sn != "":
                    sql = sql + 'AND "SN" = \'' + sn + '\''
                    sql_count = sql_count + 'AND "SN" = \'' + sn + '\''
                    visual_sql = visual_sql + 'AND "SN" = \'' + sn + '\''
                # 正常
                visual_sql_normal = visual_sql + 'AND "NGRate" < \'' + parmeter_stands_min + '\' group by "PartName"'
                # 预警
                visual_sql_waring = visual_sql + 'AND "NGRate" >= \'' + parmeter_stands_min + '\'' \
                                    + 'AND "NGRate" <= \'' + parmeter_stands_max + '\' group by "PartName"'
                # 超标
                visual_sql_danger = visual_sql + 'AND "NGRate" > \'' + parmeter_stands_max + '\' group by "PartName"'
                cur = connection.cursor()
                cur.execute(visual_sql_normal)
                normal = cur.fetchall()
                cur = connection.cursor()
                cur.execute(visual_sql_waring)
                warning = cur.fetchall()
                cur = connection.cursor()
                cur.execute(visual_sql_danger)
                danger = cur.fetchall()

                cur = connection.cursor()
                cur.execute(sql_count)
                count = cur.fetchall() #数量的总数
                if number == "All":
                    cur = connection.cursor()
                    cur.execute(sql)
                    data = cur.fetchall()
                    dict_data['data'] = data
                    dict_data['page_count'] = count[0][0]
                if number != "All":
                    number = int(number)
                    count_page = count[0][0] // int(number)  # 总数除以一页显示多少条，得到总的页数
                    if count[0][0] % number > 0:
                        count_page += 1
                    if page <= count_page:
                        sql = sql + ' order by "Id" limit ' + str(number) + ' offset ' + str((page - 1)*number)
                        cur = connection.cursor()
                        cur.execute(sql)
                        data = cur.fetchall()
                        dict_data['data'] = data
                        dict_data['page_count'] = count_page
                    else:
                        return restful.params_error(message="it had no other pages")
                dict_data['limit_value'] = limit_value
                dict_data['normal'] = normal
                dict_data['warning'] = warning
                dict_data['danger'] = danger
                return restful.ok(data=dict_data)
            else:
                if sn != "":
                    sql = sql + 'AND "SN" = \'' + sn + '\''
                    sql_count = sql_count + 'AND "SN" = \'' + sn + '\''
                    visual_sql = visual_sql + 'AND "SN" = \'' + sn + '\''
                if part_name != "":
                    sql = sql + 'AND "PartName" like \'%{0}%\''.format(part_name)
                    sql_count = sql_count + 'AND "PartName" like \'%{0}%\''.format(part_name)
                    visual_sql = visual_sql + 'AND "PartName" like \'%{0}%\''.format(part_name)
                if start_time != "":
                    sql = sql + 'AND "TrnDate" >= \'%{0}%\''.format(start_time)
                    sql_count = sql_count + 'AND "TrnDate" >= \'%{0}%\''.format(start_time)
                    visual_sql = visual_sql + 'AND "TrnDate" >= \'%{0}%\''.format(start_time)
                if end_time != "":
                    sql = sql + 'AND "TrnDate" <= \'%{0}%\''.format(end_time)
                    sql_count = sql_count + 'AND "TrnDate" <= \'%{0}%\''.format(end_time)
                    visual_sql = visual_sql + 'AND "TrnDate" <= \'%{0}%\''.format(end_time)
                if status == "正常":
                    sql = sql + 'AND "NGRate" < \'' + parmeter_stands_min + '\''
                    sql_count = sql_count + 'AND "NGRate" < \'' + parmeter_stands_min + '\''
                    visual_sql_normal = visual_sql + 'AND "NGRate" < \'' + parmeter_stands_min + '\' group by "PartName"'
                    cur = connection.cursor()
                    cur.execute(visual_sql_normal)
                    normal = cur.fetchall()
                if status == "预警":
                    sql = sql + 'AND "NGRate" >= \'' + parmeter_stands_min + '\'' + 'AND "NGRate" <= \'' \
                          + parmeter_stands_max + '\''
                    sql_count = sql_count + 'AND "NGRate" >= \'' + parmeter_stands_min + '\'' + 'AND "NGRate" <= \'' \
                          + parmeter_stands_max + '\''
                    visual_sql_waring = visual_sql + 'AND "NGRate" >= \'' + parmeter_stands_min + '\'' \
                                        + 'AND "NGRate" <= \'' + parmeter_stands_max + '\'group by "PartName"'
                    cur = connection.cursor()
                    cur.execute(visual_sql_waring)
                    warning = cur.fetchall()
                if status == "超标":
                    sql = sql + 'AND "NGRate" > \'' + parmeter_stands_max + '\''
                    sql_count = sql_count + 'AND "NGRate" > \'' + parmeter_stands_max + '\''
                    visual_sql_danger = visual_sql + 'AND "NGRate" > \'' + parmeter_stands_max + '\' group by "PartName"'
                    cur = connection.cursor()
                    cur.execute(visual_sql_danger)
                    danger = cur.fetchall()
                cur = connection.cursor()
                cur.execute(sql_count)
                count = cur.fetchall()  # 数量的总数
                if number == "All":
                    cur = connection.cursor()
                    cur.execute(sql)
                    data = cur.fetchall()
                    dict_data['data'] = data
                    dict_data['page_count'] = count[0][0]
                if number != "All":
                    number = int(number)
                    count_page = count[0][0] // int(number)  # 总数除以一页显示多少条，得到总的页数
                    if count[0][0] % number > 0:
                        count_page += 1
                    if page <= count_page:
                        sql = sql + ' order by "Id" limit ' + str(number) + ' offset ' + str((page - 1)*number)
                        cur = connection.cursor()
                        cur.execute(sql)
                        data = cur.fetchall()
                        dict_data['data'] = data
                        dict_data['page_count'] = count_page
                    else:
                        return restful.params_error(message="it had no other pages")
                dict_data['normal'] = normal
                dict_data['warning'] = warning
                dict_data['danger'] = danger
                dict_data['limit_value'] = limit_value
                return restful.ok(data=dict_data)
        except Exception as e:
            return restful.params_error(message=repr(e))

# html视图中点击事件的查询
# def visual_data(request):
#     if request.method == "POST":
#         try:
#             page = int(request.POST.get('page'))
#             number = request.POST.get('num')
#             status = request.POST.get('status',"")
#             part_name = request.POST.get('part_name',"")
#             sn = request.POST.get('sn',"")
#             start_time = request.POST.get('start_tim',"")
#             end_time = request.POST.get('end_tim',"")
#             NG_monitor_type = "NG_range"
#             dict_data = {}
#             parmeter_stands = Configuration.objects.get(Type=NG_monitor_type)
#             limit_value = [parmeter_stands.Min, parmeter_stands.Max]
#             parmeter_stands_min = str(parmeter_stands.Min)
#             parmeter_stands_max = str(parmeter_stands.Max)
#             #回滚到初始页面的条件
#             start = datetime.now()
#             start_select = (list(PartItem.objects.order_by("-TrnDate").filter(TrnDate__lte=start).values("TrnDate")))[0]['TrnDate']
#             delta_select = timedelta(days=7)
#             end_select = start_select - delta_select
#             sql = 'select "Id","SN","OSN","PN","PartName","Spec","UsedTimes","CreatedTime","UpdatedTime","CheckCycle"' \
#                   ',"CheckCycleCount","NextCheckCount","NextCheckDate","ErrorCounts","TrnDate","NGRate" ' \
#                   'FROM "PartItem" WHERE 1 = 1 '
#             sql_count ='select count(*) from "PartItem" where 1=1'
#             if start_time =="" and end_time == "" and part_name == "" and sn == "":
#                 sql = sql +'AND "TrnDate">= \'{0}\' AND "TrnDate"<= \'{1}\''.format(end_select, start_select)
#                 sql_count = sql_count +'AND "TrnDate">= \'{0}\' AND "TrnDate"<= \'{1}\''.format(end_select, start_select)
#             if status == "#28a745":
#                 sql = sql + 'AND "NGRate" < \'' + parmeter_stands_min + '\''
#                 sql_count = sql_count + 'AND "NGRate" < \'' + parmeter_stands_min + '\''
#             if status == "#ffc107":
#                 sql = sql + 'AND "NGRate" >= \'' + parmeter_stands_min + '\'' \
#                       + 'AND "NGRate" <= \'' + parmeter_stands_max + '\''
#                 sql_count = sql_count + 'AND "NGRate" >= \'' + parmeter_stands_min + '\'' \
#                       + 'AND "NGRate" <= \'' + parmeter_stands_max + '\''
#             if status == "#dc3545":
#                 sql = sql + 'AND "NGRate" > \'' + parmeter_stands_max + '\''
#                 sql_count = sql_count + 'AND "NGRate" > \'' + parmeter_stands_max + '\''
#             if start_time != "":
#                 sql = sql + 'AND "TrnDate" >= \'' + start_time + '\''
#                 sql_count = sql_count + 'AND "TrnDate" >= \'' + start_time + '\''
#             if end_time != "":
#                 sql = sql + 'AND "TrnDate" <= \'' + end_time + '\''
#                 sql_count = sql_count + 'AND "TrnDate" <= \'' + end_time + '\''
#             if part_name != "":
#                 sql = sql + 'AND "PartName" like  \'' + part_name + '\''
#                 sql_count = sql_count + 'AND "PartName" like  \'' + part_name + '\''
#             if sn != "":
#                 sql = sql + 'AND "SN" = \'' + sn + '\''
#                 sql_count = sql_count + 'AND "SN" = \'' + sn + '\''
#             cur = connection.cursor()
#             cur.execute(sql_count)
#             count = cur.fetchall() #获取赛选的条件的总数值
#             if number == "All":
#                 cur = connection.cursor()
#                 cur.execute(sql)
#                 data = cur.fetchall()
#                 dict_data['data'] = data
#                 dict_data['page_count'] = count[0][0]
#             if number != "All":
#                 number = int(number)
#                 count_page = count[0][0] // number  # 总数除以一页显示多少条，得到总的页数
#                 if count[0][0] % number > 0:
#                     count_page += 1
#                 if page <= count_page:
#                     sql = sql + ' order by "UpdatedTime" limit ' + str(number) + ' offset ' + str((page - 1)*number)
#                     cur = connection.cursor()
#                     cur.execute(sql)
#                     data = cur.fetchall()
#                     dict_data['data'] = data
#                     dict_data['page_count'] = count_page
#                 else:
#                     return restful.params_error(message="it had no other pages")
#             dict_data['limit_value'] = limit_value
#             return restful.ok(data=dict_data)
#         except Exception as e:
#             return restful.params_error(message=repr(e))
#视图点击事件的查询问题的修改
@access_control
def visual_data(request):
    if request.method == "POST":
        try:
            page = int(request.POST['page'])
            number = request.POST['num']
            sn = str(request.POST.get('sn',''))
            part_name = request.POST.get('part_name','')
            status = request.POST.get('status','')
            start_time = request.POST.get('start_tim','')
            end_time = request.POST.get('end_tim','')
            NG_monitor_type = "NG_range"
            # 时间区间回滚到初始状态
            start = datetime.now()
            start_select = (list(PartItem.objects.order_by("-TrnDate").filter(TrnDate__lte=start,UseStatus='normal').values("TrnDate")))[0][
                'TrnDate']
            delta_select = timedelta(days=7)
            end_select = start_select - delta_select
            visual_callback = ' AND "TrnDate">= \'{0}\' AND "TrnDate"<= \'{1}\''.format(end_select, start_select)
            sql_count = 'select count(*) FROM "PartItem" where "UseStatus"=\'normal\' '
            sql = 'select "Id","SN","OSN","PN","PartName","Spec","UsedTimes","CreatedTime","UpdatedTime","CheckCycle"' \
                  ',"CheckCycleCount","NextCheckCount","NextCheckDate","ErrorCounts","TrnDate","NGRate" ' \
                  'FROM "PartItem" WHERE "UseStatus"=\'normal\''
            visual_sql = 'SELECT "PartName", COUNT("PartName") FROM "PartItem" where "UseStatus"=\'normal\' '

            parmeter_stands = Configuration.objects.get(Type=NG_monitor_type)
            limit_value = [parmeter_stands.Min, parmeter_stands.Max]
            parmeter_stands_min = str(parmeter_stands.Min)
            parmeter_stands_max = str(parmeter_stands.Max)
            dict_data = {}
            normal = []
            warning = []
            danger = []
            if start_time == "" and end_time == "" and part_name == "" and sn == "":
                sql = sql + visual_callback
                visual_sql = visual_sql + visual_callback

            if sn != "":
                sql = sql + 'AND "SN" = \'' + sn + '\''
                sql_count = sql_count + 'AND "SN" = \'' + sn + '\''
                visual_sql = visual_sql + 'AND "SN" = \'' + sn + '\''
            if part_name != "":
                sql = sql + 'AND "PartName" like \'%{0}%\''.format(part_name)
                sql_count = sql_count + 'AND "PartName" like \'%{0}%\''.format(part_name)
                visual_sql = visual_sql + 'AND "PartName" like \'%{0}%\''.format(part_name)
            if start_time != "":
                sql = sql + 'AND "TrnDate" >= \'%{0}%\''.format(start_time)
                sql_count = sql_count + 'AND "TrnDate" >= \'%{0}%\''.format(start_time)
                visual_sql = visual_sql + 'AND "TrnDate" >= \'%{0}%\''.format(start_time)
            if end_time != "":
                sql = sql + 'AND "TrnDate" <= \'%{0}%\''.format(end_time)
                sql_count = sql_count + 'AND "TrnDate" <= \'%{0}%\''.format(end_time)
                visual_sql = visual_sql + 'AND "TrnDate" <= \'%{0}%\''.format(end_time)
            if status == "#28a745":
                sql = sql + 'AND "NGRate" < \'' + parmeter_stands_min + '\''
                sql_count = sql_count + 'AND "NGRate" < \'' + parmeter_stands_min + '\''
                visual_sql_normal = visual_sql + 'AND "NGRate" < \'' + parmeter_stands_min + '\' group by "PartName"'
                cur = connection.cursor()
                cur.execute(visual_sql_normal)
                normal = cur.fetchall()
            if status == "#ffc107":
                sql = sql + 'AND "NGRate" >= \'' + parmeter_stands_min + '\'' + 'AND "NGRate" <= \'' \
                      + parmeter_stands_max + '\''
                sql_count = sql_count + 'AND "NGRate" >= \'' + parmeter_stands_min + '\'' + 'AND "NGRate" <= \'' \
                            + parmeter_stands_max + '\''
                visual_sql_waring = visual_sql + 'AND "NGRate" >= \'' + parmeter_stands_min + '\'' \
                                    + 'AND "NGRate" <= \'' + parmeter_stands_max + '\'group by "PartName"'
                cur = connection.cursor()
                cur.execute(visual_sql_waring)
                warning = cur.fetchall()
            if status == "#dc3545":
                sql = sql + 'AND "NGRate" > \'' + parmeter_stands_max + '\''
                sql_count = sql_count + 'AND "NGRate" > \'' + parmeter_stands_max + '\''
                visual_sql_danger = visual_sql + 'AND "NGRate" > \'' + parmeter_stands_max + '\' group by "PartName"'
                cur = connection.cursor()
                cur.execute(visual_sql_danger)
                danger = cur.fetchall()
            cur = connection.cursor()
            cur.execute(sql_count)
            count = cur.fetchall()  # 数量的总数
            if number == "All":
                cur = connection.cursor()
                cur.execute(sql)
                data = cur.fetchall()
                dict_data['data'] = data
                dict_data['page_count'] = count[0][0]
            if number != "All":
                number = int(number)
                count_page = count[0][0] // int(number)  # 总数除以一页显示多少条，得到总的页数
                if count[0][0] % number > 0:
                    count_page += 1
                if page <= count_page:
                    sql = sql + ' order by "Id" limit ' + str(number) + ' offset ' + str((page - 1) * number)
                    cur = connection.cursor()
                    cur.execute(sql)
                    data = cur.fetchall()
                    dict_data['data'] = data
                    dict_data['page_count'] = count_page
                else:
                    return restful.params_error(message="it had no other pages")
            dict_data['normal'] = normal
            dict_data['warning'] = warning
            dict_data['danger'] = danger
            dict_data['limit_value'] = limit_value
            return restful.ok(data=dict_data)
        except Exception as e:
            return restful.params_error(message=repr(e))

# 捞出所有NG率达到或超过预警区间的SN发邮件提醒给收件人 定时的功能在被使用在DBexcle app。views里面的函数crontab_test使用了
def check_NGRate():
    try:
        NG_monitor_type = "NG_range"
        parmeter_stands = Configuration.objects.get(Type=NG_monitor_type)
        parmeter_stands_min = str(parmeter_stands.Min)
        parmeter_stands_max = str(parmeter_stands.Max)
        parmeter_stands_receiver = str(parmeter_stands.Reminders)
        sql_warning = 'select count(*) from "PartItem" where "UseStatus"=\'normal\' and "NGRate" >= \'' + parmeter_stands_min + '\' and "NGRate" <= \'' + parmeter_stands_max + '\''
        sql_danger = 'select count(*) from "PartItem" where "UseStatus"=\'normal\' and "NGRate" > \'' + parmeter_stands_max + '\''
        cur = connection.cursor()
        cur.execute(sql_warning)
        data_warning = cur.fetchall()
        cur = connection.cursor()
        cur.execute(sql_danger)
        data_danger = cur.fetchall()
        Total =int(data_warning[0][0])+int(data_danger[0][0])
        Warning =data_warning[0][0]
        danger =data_danger[0][0]
        if int(data_warning[0][0]) > 0 or int(data_danger[0][0]) >0:
            email_1 = []
            receiver_list = parmeter_stands_receiver.split(',')
            mail_user_data_info = list(User.objects.filter(Name__in=receiver_list).values("Email"))
            for i in range(0, len(mail_user_data_info)):
                email_1.append(mail_user_data_info[i]['Email'])
            # for i in range(len(receiver_list)):
            #     email_1.append(receiver_list[i] + '@wistron.com')
            subject = "AEMS Lite System Notification for NG Reported "
            content = """
<pre>
Dears,
    AEMS Lite system found that the equipment NG rate have reached or exceed the standard.
    Total: """+str(Total)+"""
    Exceeding Standard: """+str(danger)+"""
    Warning: """+str(Warning)+"""


THIS EMAIL WAS SENT BY PTS SERVER AUTOMATICALLY. PLEASE DON'T DIRECTLY REPLY!!!
AEMS Lite System http://10.41.95.89:90/index/
</pre>
"""
            mail.sendmail(email_1, content, subject)
            return [Total,Warning,danger]
        else:
            return [0,0,0]
    except Exception as e:
        return repr(e)

#页码转换之后的获取数据
def monitor_equipment_change(request):
    if request.method == "GET":
        try:
            end_time = datetime.now()  # 这里默认是查询前几周的数据
            start_time = date(2017, 1, 1)  # 这里默认是查询前几周的数据
            # delta = timedelta(days=21)
            # start_time = end_time-delta
            # 页面html显示需要的数据+分页效果
            page = int(request.GET.get('page'))
            number = request.GET.get('num')
            dict_data = {}
            count = PartItem.objects.order_by("Id").filter(TrnDate__range=(start_time, end_time),UseStatus='normal').count()
            # 页面显示需要的数据
            if number == "All":
                data = PartItem.objects.order_by("Id").filter(TrnDate__range=(start_time, end_time),UseStatus='normal')
                limit_value = Configuration.objects.filter(Type="NG_range").values("Max", "Min", "Id")
                limit_value = list(limit_value)
                data = list(data.values())
                dict_data['data'] = data
                dict_data['limit_value'] = limit_value
                dict_data['page_count'] = count
                return restful.ok(data=dict_data)
            if number != "All":
                number = int(number)
                page_num = count // number  # 总共多少页
                if count % number > 0:
                    page_num = page_num + 1
                if page_num >= page:
                    data = PartItem.objects.order_by("Id").filter(TrnDate__range=(start_time, end_time),UseStatus='normal')[
                           (page - 1) * number:number * page]
                    limit_value = Configuration.objects.filter(Type="NG_range").values("Max", "Min", "Id")
                    limit_value = list(limit_value)
                    data = list(data.values())
                    dict_data['data'] = data
                    dict_data['limit_value'] = limit_value
                    dict_data['page_count'] = page_num
                    return restful.ok(data=dict_data)
                else:
                    return restful.params_error(message='it had no others page')
        except Exception as e:
            return restful.params_error(message=repr(e))
#查询函数的获取页码的查询数据
def monitor_equipment_query_next(request):
    if request.method == "GET":
        page = int(request.GET['page'])
        number = request.GET['num']
        sn = str(request.GET['sn'])
        part_name = request.GET['part_name']
        status = request.GET['status']
        start_time = request.GET['start_tim']
        end_time = request.GET['end_tim']
        NG_monitor_type = "NG_range"
        dict_data = {}
        sql_count = 'select count(*) FROM "PartItem" where "UseStatus"=\'normal\' '
        sql = 'select "Id","SN","OSN","PN","PartName","Spec","UsedTimes","CreatedTime","UpdatedTime","CheckCycle"' \
              ',"CheckCycleCount","NextCheckCount","NextCheckDate","ErrorCounts","TrnDate","NGRate" ' \
              'FROM "PartItem" WHERE "UseStatus"=\'normal\''
        try:
            parmeter_stands = Configuration.objects.get(Type=NG_monitor_type)
            limit_value = [parmeter_stands.Min, parmeter_stands.Max]
            dict_data['limit_value'] = limit_value
            parmeter_stands_min = str(parmeter_stands.Min)
            parmeter_stands_max = str(parmeter_stands.Max)
            if sn != "":
                sql = sql + 'AND "SN" = \'' + sn + '\''
                sql_count = sql_count + 'AND "SN" = \'' + sn + '\''
            if part_name != "":
                sql = sql + 'AND "PartName" like \'%{0}%\''.format(part_name)
                sql_count = sql_count + 'AND "PartName" like \'%{0}%\''.format(part_name)
            if start_time != "":
                sql = sql + 'AND "TrnDate" >= \'%{0}%\''.format(start_time)
                sql_count = sql_count + 'AND "TrnDate" >= \'%{0}%\''.format(start_time)
            if end_time != "":
                sql = sql + 'AND "TrnDate" <= \'%{0}%\''.format(end_time)
                sql_count = sql_count + 'AND "TrnDate" <= \'%{0}%\''.format(end_time)
            if status == "正常":
                sql = sql + 'AND "NGRate" < \'' + parmeter_stands_min + '\''
                sql_count = sql_count + 'AND "NGRate" < \'' + parmeter_stands_min + '\''
            if status == "预警":
                sql = sql + 'AND "NGRate" >= \'' + parmeter_stands_min + '\'' + 'AND "NGRate" <= \'' \
                      + parmeter_stands_max + '\''
                sql_count = sql_count + 'AND "NGRate" >= \'' + parmeter_stands_min + '\'' + 'AND "NGRate" <= \'' \
                            + parmeter_stands_max + '\''
            if status == "超标":
                sql = sql + 'AND "NGRate" > \'' + parmeter_stands_max + '\''
                sql_count = sql_count + 'AND "NGRate" > \'' + parmeter_stands_max + '\''
            cur = connection.cursor()
            cur.execute(sql_count)
            count = cur.fetchall()  # 数量的总数
            if number == "All":
                cur = connection.cursor()
                cur.execute(sql)
                data = cur.fetchall()
                dict_data['data'] = data
                dict_data['page_count'] = count[0][0]
            if number != "All":
                number = int(number)
                count_page = count[0][0] // int(number)  # 总数除以一页显示多少条，得到总的页数
                if count[0][0] % number > 0:
                    count_page += 1
                if page <= count_page:
                    sql = sql + ' order by "Id" limit ' + str(number) + ' offset ' + str((page - 1)*number)
                    cur = connection.cursor()
                    cur.execute(sql)
                    data = cur.fetchall()
                    dict_data['data'] = data
                    dict_data['page_count'] = count_page
                else:
                    return restful.params_error(message="it had no other pages")
            return restful.ok(data=dict_data)
        except Exception as e:
            return restful.params_error(message=repr(e))
